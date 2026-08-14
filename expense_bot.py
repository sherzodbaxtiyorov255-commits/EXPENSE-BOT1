"""
Xarajatlar/Balans Telegram boti (to'liq versiya 2)
------------------------------------------------------
Endi tugma bosmasdan ham oddiy gapda yozishingiz mumkin:
    "500000 Alidan qarz oldim"
    "13 mln maosh oldim"
    "200 000 Valiga qarz berdim"
Ovozli xabar yuborsangiz ham, bot uni matnga aylantirib xuddi shunday
qayta ishlaydi (agar tegishli dastur o'rnatilgan bo'lsa).

Oddiy xabar formatlari:
    -50000 taksi          -> UZS xarajat
    50000 nonushta         -> belgi bo'lmasa, XARAJAT deb hisoblanadi
    13 mln maosh            -> "maosh" so'zi borligi uchun avtomatik DAROMAD
    +200000 maosh           -> aniq belgi bilan ham yozish mumkin
    13000000 yoki 13 mln    -> ikkalasi ham bir xil summa deb tushuniladi

Buyruqlar:
    /start, /help     - yo'riqnoma va pastki menyu
    /balance          - balans + qarzlar + sof holat
    /history          - sana bilan birlashtirilgan tarix
    /stat             - matnli va diagrammali statistika
    /categories       - kategoriyalar bo'yicha jami xarajat
    /kurs 12700       - USD kursini belgilash (yoki 💱 Kurs tugmasi)
    /group_new, /group_join KOD, /group_leave, /group_info
    /qarz_oldim Ism summa izoh, /qarz_berdim Ism summa izoh
    /qarzlar, /qarz_yopish Ism
    /excel, /reset

O'rnatish (asosiy):
    pip install python-telegram-bot openpyxl matplotlib --break-system-packages

Qo'shimcha (ixtiyoriy funksiyalar uchun):
    Skrinshot o'qish:  pip install pytesseract Pillow --break-system-packages
                       + Tesseract-OCR dasturi (https://github.com/UB-Mannheim/tesseract/wiki)
    Ovozli xabar:      pip install SpeechRecognition pydub --break-system-packages
                       + ffmpeg dasturi (https://www.gyan.dev/ffmpeg/builds/, PATH'ga qo'shing)
    Bular o'rnatilmasa ham, bot qolgan hamma funksiyalar bilan ishlayveradi.

Ishga tushirish:
    export BOT_TOKEN="sizning_bot_tokeningiz"
    python3 expense_bot.py

Bot tokenini olish: Telegram-da @BotFather ga yozing -> /newbot
"""

import os
import re
import random
import string
import tempfile
import sqlite3
from datetime import datetime, timedelta, time as dt_time

from telegram import (
    Update,
    ReplyKeyboardMarkup,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
)
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters,
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import pytesseract
    from PIL import Image

    OCR_AVAILABLE = True
    _tcmd = os.environ.get("TESSERACT_CMD")
    if _tcmd:
        pytesseract.pytesseract.tesseract_cmd = _tcmd
    elif os.name == "nt":
        _default_win_path = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        if os.path.exists(_default_win_path):
            pytesseract.pytesseract.tesseract_cmd = _default_win_path
except ImportError as e:
    OCR_AVAILABLE = False
    print(f"[OGOHLANTIRISH] OCR o'rnatilmagan: {e}")

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    MATPLOTLIB_AVAILABLE = True
except ImportError as e:
    MATPLOTLIB_AVAILABLE = False
    print(f"[OGOHLANTIRISH] matplotlib o'rnatilmagan: {e}")

try:
    import speech_recognition as sr
    from pydub import AudioSegment

    VOICE_AVAILABLE = True
except ImportError as e:
    VOICE_AVAILABLE = False
    print(f"[OGOHLANTIRISH] Ovoz kutubxonalari o'rnatilmagan: {e}")

try:
    import anthropic

    _anthropic_key = os.environ.get("ANTHROPIC_API_KEY")
    if _anthropic_key:
        anthropic_client = anthropic.Anthropic(api_key=_anthropic_key)
        AI_AVAILABLE = True
    else:
        anthropic_client = None
        AI_AVAILABLE = False
except ImportError as e:
    anthropic_client = None
    AI_AVAILABLE = False
    print(f"[OGOHLANTIRISH] AI (anthropic) kutubxonasi o'rnatilmagan: {e}")

print(f"[STATUS] OCR_AVAILABLE={OCR_AVAILABLE}, MATPLOTLIB_AVAILABLE={MATPLOTLIB_AVAILABLE}, VOICE_AVAILABLE={VOICE_AVAILABLE}, AI_AVAILABLE={AI_AVAILABLE}")

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "expenses.db")
DEFAULT_KURS = 12700.0  # taxminiy 1 USD = necha UZS

# ---------- Pastki menyu tugmalari ----------

BTN_BALANCE = "💰 Balans"
BTN_HISTORY = "📜 Tarix"
BTN_STAT = "📊 Statistika"
BTN_CATEGORIES = "🗂 Kategoriyalar"
BTN_DEBTS = "🤝 Qarzlar"
BTN_KURS = "💱 Kurs"
BTN_CARD = "💳 Karta xarajati"
BTN_CASH = "💵 Naqt xarajat"
BTN_SALDO = "💳 Hisoblar saldosi"
BTN_EXCEL = "📥 Excel"
BTN_HELP = "❓ Yordam"

MAIN_KEYBOARD = ReplyKeyboardMarkup(
    [
        [BTN_BALANCE, BTN_HISTORY],
        [BTN_STAT, BTN_CATEGORIES],
        [BTN_DEBTS, BTN_KURS],
        [BTN_CARD, BTN_CASH],
        [BTN_SALDO, BTN_EXCEL],
        [BTN_HELP],
    ],
    resize_keyboard=True,
)

# ---------- Kategoriyalarni aniqlash ----------

CATEGORY_KEYWORDS = {
    "Oziq-ovqat": ["ovqat", "tushlik", "nonushta", "kechki", "restoran", "kafe",
                   "oshxona", "market", "magazin", "do'kon", "dokon", "un", "non",
                   "gosht", "go'sht", "sabzavot", "meva"],
    "Transport": ["taksi", "avtobus", "metro", "benzin", "yoqilg'i", "yoqilgi",
                  "mashina", "yo'l", "yol", "parkovka"],
    "Kommunal": ["svet", "elektr", "gaz", "suv", "kommunal", "ijara", "kvartira",
                 "kommunalka"],
    "Aloqa": ["mobil", "telefon", "internet", "wifi", "balans", "uzmobile",
              "beeline", "ucell"],
    "Sog'liq": ["dori", "shifokor", "klinika", "gospital", "dorixona", "vrach"],
    "Kiyim": ["kiyim", "futbolka", "poyabzal", "shim", "куртка", "kurtka"],
    "Ta'lim": ["kurs", "kitob", "maktab", "universitet", "repetitor", "darslik"],
    "O'yin-kulgi": ["kino", "konsert", "o'yin", "oyin", "bar", "klub", "disko"],
    "Daromad": ["maosh", "oylik", "bonus", "sovg'a", "sovga", "foyda", "kirim"],
}


def fmt_num(value: float, decimals: int = 0) -> str:
    """Katta raqamlarni bo'shliq bilan ajratib chiqaradi: 13 000 000
    (vergul bilan emas — ko'p nol bo'lganda chalkashtirmasin uchun)."""
    s = f"{value:,.{decimals}f}"
    return s.replace(",", " ")


def detect_category(note: str, is_income: bool) -> str:
    note_lower = note.lower()
    for category, keywords in CATEGORY_KEYWORDS.items():
        for kw in keywords:
            if kw in note_lower:
                return category
    # Ro'yxatdagi kategoriyalarga mos kelmasa, izohning o'zidagi so'zni
    # kategoriya sifatida olamiz (masalan "suv" -> "Suv")
    skip_words = {"karta", "kart", "naqt", "naqd", "uzcard", "humo", "visa",
                  "uydagi", "oldimdagi", "izohsiz"}
    words = re.findall(r"[A-Za-zА-Яа-яЎўЎ'ʼ]+", note)
    for w in words:
        if w.lower() not in skip_words and len(w) > 1:
            return w.capitalize()
    return "Daromad (boshqa)" if is_income else "Boshqa"


def is_income_text(note: str) -> bool:
    note_lower = note.lower()
    return any(kw in note_lower for kw in CATEGORY_KEYWORDS["Daromad"])


# ---------- Ma'lumotlar bazasi ----------

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            space_key TEXT NOT NULL,
            amount REAL NOT NULL,
            currency TEXT NOT NULL,
            category TEXT,
            note TEXT,
            created_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL,
            created_by INTEGER NOT NULL,
            created_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS user_group (
            user_id INTEGER PRIMARY KEY,
            group_id INTEGER NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS settings (
            space_key TEXT NOT NULL,
            key TEXT NOT NULL,
            value TEXT NOT NULL,
            PRIMARY KEY (space_key, key)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS debts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            space_key TEXT NOT NULL,
            person TEXT NOT NULL,
            direction TEXT NOT NULL,
            amount REAL NOT NULL,
            currency TEXT NOT NULL,
            note TEXT,
            created_at TEXT NOT NULL,
            settled INTEGER NOT NULL DEFAULT 0
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS known_chats (
            chat_id INTEGER PRIMARY KEY,
            space_key TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS accounts (
            space_key TEXT NOT NULL,
            name TEXT NOT NULL,
            balance REAL NOT NULL DEFAULT 0,
            PRIMARY KEY (space_key, name)
        )
        """
    )
    _migrate_old_schema(conn)
    return conn


def _migrate_old_schema(conn):
    cols = {row[1] for row in conn.execute("PRAGMA table_info(transactions)").fetchall()}
    if "space_key" not in cols:
        conn.execute("ALTER TABLE transactions ADD COLUMN space_key TEXT")
        conn.execute("UPDATE transactions SET space_key='u' || user_id WHERE space_key IS NULL")
    if "category" not in cols:
        conn.execute("ALTER TABLE transactions ADD COLUMN category TEXT")
        conn.execute("UPDATE transactions SET category='Boshqa' WHERE category IS NULL")
    if "method" not in cols:
        conn.execute("ALTER TABLE transactions ADD COLUMN method TEXT")
    if "account" not in cols:
        conn.execute("ALTER TABLE transactions ADD COLUMN account TEXT")
    conn.commit()


ACCOUNT_NAMES = ["Uzcard", "Humo", "Visa", "Uydagi naqt", "Oldimdagi naqt"]
ACCOUNT_ALIASES = {
    "uzcard": "Uzcard", "uz card": "Uzcard",
    "humo": "Humo",
    "visa": "Visa",
    "uydagi naqt": "Uydagi naqt", "uyda": "Uydagi naqt", "uydagi": "Uydagi naqt",
    "oldimdagi naqt": "Oldimdagi naqt", "cepdagi": "Oldimdagi naqt",
    "hamyondagi": "Oldimdagi naqt", "yonimdagi": "Oldimdagi naqt", "cho'ntakdagi": "Oldimdagi naqt",
}


def detect_account(note: str):
    note_lower = note.lower()
    for alias, name in sorted(ACCOUNT_ALIASES.items(), key=lambda kv: -len(kv[0])):
        if alias in note_lower:
            return name
    return None


def ensure_default_accounts(space_key: str):
    conn = get_db()
    for name in ACCOUNT_NAMES:
        conn.execute(
            "INSERT OR IGNORE INTO accounts (space_key, name, balance) VALUES (?, ?, 0)",
            (space_key, name),
        )
    conn.commit()
    conn.close()


def get_accounts(space_key: str):
    ensure_default_accounts(space_key)
    conn = get_db()
    rows = conn.execute(
        "SELECT name, balance FROM accounts WHERE space_key=? ORDER BY rowid", (space_key,)
    ).fetchall()
    conn.close()
    return rows


def set_account_balance(space_key: str, name: str, value: float):
    ensure_default_accounts(space_key)
    conn = get_db()
    conn.execute(
        "UPDATE accounts SET balance=? WHERE space_key=? AND name=?", (value, space_key, name)
    )
    conn.commit()
    conn.close()


def apply_account_delta(space_key: str, name: str, delta: float):
    ensure_default_accounts(space_key)
    conn = get_db()
    conn.execute(
        "UPDATE accounts SET balance = balance + ? WHERE space_key=? AND name=?",
        (delta, space_key, name),
    )
    conn.commit()
    conn.close()


def get_today_account_expense(space_key: str):
    """Bugun har bir hisobdan qancha xarajat qilinganini (UZS ekvivalentida) qaytaradi."""
    now = datetime.now()
    start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    kurs = get_kurs(space_key)
    conn = get_db()
    rows = conn.execute(
        "SELECT account, currency, SUM(amount) FROM transactions "
        "WHERE space_key=? AND account IS NOT NULL AND amount<0 AND created_at>=? "
        "GROUP BY account, currency",
        (space_key, start.isoformat(timespec="seconds")),
    ).fetchall()
    conn.close()
    totals = {}
    for account, currency, total in rows:
        val_uzs = abs(total) * (kurs if currency == "USD" else 1)
        totals[account] = totals.get(account, 0.0) + val_uzs
    return totals


def build_accounts_summary_text(space_key: str) -> str:
    rows = get_accounts(space_key)
    today_expense = get_today_account_expense(space_key)
    today_str = datetime.now().strftime("%d.%m.%Y")
    lines = [f"🏦 Hisoblar bo'yicha saldo ({today_str}):"]
    total_balance = 0.0
    total_today_expense = 0.0
    for name, balance in rows:
        total_balance += balance
        exp_today = today_expense.get(name, 0.0)
        total_today_expense += exp_today
        if exp_today > 0:
            lines.append(f"  {name}: {fmt_num(balance)} so'm  (bugun xarajat: {fmt_num(exp_today)} so'm)")
        else:
            lines.append(f"  {name}: {fmt_num(balance)} so'm")
    lines.append("  —")
    lines.append(f"  Umumiy (barcha hisoblar): {fmt_num(total_balance)} so'm")
    if total_today_expense > 0:
        lines.append(f"  Bugungi umumiy xarajat: {fmt_num(total_today_expense)} so'm")
    return "\n".join(lines)


def parse_saldo_input(text: str):
    text_stripped = text.strip()
    lower = text_stripped.lower()
    for name in sorted(ACCOUNT_NAMES, key=lambda n: -len(n)):
        if lower.startswith(name.lower()):
            rest = text_stripped[len(name):].strip()
            m = re.search(r"(\d[\d\s.,]*)\s*(mln|million|млн|ming|минг)?\b", rest, re.IGNORECASE)
            if m:
                value = parse_amount_token(m.group(1), m.group(2))
                if value is not None:
                    return name, value
    return None


def get_space_key(user_id: int) -> str:
    conn = get_db()
    row = conn.execute(
        "SELECT group_id FROM user_group WHERE user_id=?", (user_id,)
    ).fetchone()
    conn.close()
    if row:
        return f"g{row[0]}"
    return f"u{user_id}"


def get_space_key_for_update(update: Update) -> str:
    """Guruh chatida yozilgan xabarlar shu guruhning umumiy balansiga
    tushadi (kim yozishidan qat'iy nazar); shaxsiy chatda esa odatdagi
    shaxsiy/jamoaviy (kod orqali qo'shilgan) balans ishlatiladi."""
    chat = update.effective_chat
    if chat and chat.type in ("group", "supergroup"):
        space_key = f"c{chat.id}"
    else:
        space_key = get_space_key(update.effective_user.id)
    remember_chat(chat.id if chat else update.effective_user.id, space_key)
    return space_key


def remember_chat(chat_id: int, space_key: str):
    """Kun oxiridagi avtomatik xabar yuborish uchun qaysi chat qaysi
    balansga tegishli ekanini yodda saqlaydi."""
    conn = get_db()
    conn.execute(
        "INSERT INTO known_chats (chat_id, space_key, updated_at) VALUES (?, ?, ?) "
        "ON CONFLICT(chat_id) DO UPDATE SET space_key=excluded.space_key, updated_at=excluded.updated_at",
        (chat_id, space_key, datetime.now().isoformat(timespec="seconds")),
    )
    conn.commit()
    conn.close()


def get_all_known_chats():
    conn = get_db()
    rows = conn.execute("SELECT chat_id, space_key FROM known_chats").fetchall()
    conn.close()
    return rows


def add_transaction(user_id: int, space_key: str, amount: float, currency: str, category: str, note: str, method: str = None, account: str = None):
    conn = get_db()
    conn.execute(
        "INSERT INTO transactions (user_id, space_key, amount, currency, category, note, method, account, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (user_id, space_key, amount, currency, category, note, method, account,
         datetime.now().isoformat(timespec="seconds")),
    )
    conn.commit()
    conn.close()


def detect_method(note: str):
    note_lower = note.lower()
    if "karta" in note_lower or "kart" in note_lower or "plastik" in note_lower:
        return "karta"
    if "naqt" in note_lower or "naqd" in note_lower or "nakt" in note_lower:
        return "naqt"
    return None


def get_method_totals(space_key: str, method: str):
    conn = get_db()
    rows = conn.execute(
        "SELECT currency, SUM(amount) FROM transactions "
        "WHERE space_key=? AND method=? AND amount<0 GROUP BY currency",
        (space_key, method),
    ).fetchall()
    recent = conn.execute(
        "SELECT amount, currency, category, note, created_at FROM transactions "
        "WHERE space_key=? AND method=? ORDER BY id DESC LIMIT 10",
        (space_key, method),
    ).fetchall()
    conn.close()
    totals = {"UZS": 0.0, "USD": 0.0}
    for currency, total in rows:
        totals[currency] = abs(total or 0.0)
    return totals, recent


def get_balance(space_key: str):
    conn = get_db()
    cur = conn.execute(
        "SELECT currency, SUM(amount) FROM transactions WHERE space_key=? GROUP BY currency",
        (space_key,),
    )
    rows = cur.fetchall()
    conn.close()
    balances = {"UZS": 0.0, "USD": 0.0}
    for currency, total in rows:
        balances[currency] = total or 0.0
    return balances


def get_kurs(space_key: str) -> float:
    conn = get_db()
    row = conn.execute(
        "SELECT value FROM settings WHERE space_key=? AND key='kurs'", (space_key,)
    ).fetchone()
    conn.close()
    return float(row[0]) if row else DEFAULT_KURS


def set_kurs(space_key: str, value: float):
    conn = get_db()
    conn.execute(
        "INSERT INTO settings (space_key, key, value) VALUES (?, 'kurs', ?) "
        "ON CONFLICT(space_key, key) DO UPDATE SET value=excluded.value",
        (space_key, str(value)),
    )
    conn.commit()
    conn.close()


def get_history(space_key: str, limit: int = 20):
    conn = get_db()
    cur = conn.execute(
        "SELECT amount, currency, category, note, created_at FROM transactions "
        "WHERE space_key=? ORDER BY id DESC LIMIT ?",
        (space_key, limit),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def get_period_stats(space_key: str, start_dt: datetime):
    conn = get_db()
    cur = conn.execute(
        "SELECT amount, currency, category FROM transactions "
        "WHERE space_key=? AND created_at>=?",
        (space_key, start_dt.isoformat(timespec="seconds")),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def get_period_stats_with_date(space_key: str, start_dt: datetime):
    conn = get_db()
    cur = conn.execute(
        "SELECT amount, currency, created_at FROM transactions "
        "WHERE space_key=? AND created_at>=?",
        (space_key, start_dt.isoformat(timespec="seconds")),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def get_combined_history(space_key: str, limit: int = 20):
    conn = get_db()
    tx_rows = conn.execute(
        "SELECT amount, currency, category, note, created_at FROM transactions "
        "WHERE space_key=? ORDER BY id DESC LIMIT ?",
        (space_key, limit),
    ).fetchall()
    debt_rows = conn.execute(
        "SELECT person, direction, amount, currency, note, created_at FROM debts "
        "WHERE space_key=? ORDER BY id DESC LIMIT ?",
        (space_key, limit),
    ).fetchall()
    conn.close()

    combined = []
    for amount, currency, category, note, created_at in tx_rows:
        combined.append({
            "kind": "tx", "created_at": created_at, "amount": amount,
            "currency": currency, "category": category, "note": note,
        })
    for person, direction, amount, currency, note, created_at in debt_rows:
        combined.append({
            "kind": "debt", "created_at": created_at, "person": person,
            "direction": direction, "amount": amount, "currency": currency, "note": note,
        })
    combined.sort(key=lambda x: x["created_at"], reverse=True)
    return combined[:limit]


def get_category_totals(space_key: str):
    conn = get_db()
    cur = conn.execute(
        "SELECT category, currency, SUM(amount) FROM transactions "
        "WHERE space_key=? AND amount<0 GROUP BY category, currency ORDER BY SUM(amount) ASC",
        (space_key,),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def reset_space(space_key: str):
    conn = get_db()
    conn.execute("DELETE FROM transactions WHERE space_key=?", (space_key,))
    conn.commit()
    conn.close()


def create_group(user_id: int) -> str:
    code = "".join(random.choices(string.ascii_uppercase + string.digits, k=6))
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO groups (code, created_by, created_at) VALUES (?, ?, ?)",
        (code, user_id, datetime.now().isoformat(timespec="seconds")),
    )
    group_id = cur.lastrowid
    conn.execute(
        "INSERT INTO user_group (user_id, group_id) VALUES (?, ?) "
        "ON CONFLICT(user_id) DO UPDATE SET group_id=excluded.group_id",
        (user_id, group_id),
    )
    conn.commit()
    conn.close()
    return code


def join_group(user_id: int, code: str) -> bool:
    conn = get_db()
    row = conn.execute("SELECT id FROM groups WHERE code=?", (code.upper(),)).fetchone()
    if not row:
        conn.close()
        return False
    group_id = row[0]
    conn.execute(
        "INSERT INTO user_group (user_id, group_id) VALUES (?, ?) "
        "ON CONFLICT(user_id) DO UPDATE SET group_id=excluded.group_id",
        (user_id, group_id),
    )
    conn.commit()
    conn.close()
    return True


def leave_group(user_id: int):
    conn = get_db()
    conn.execute("DELETE FROM user_group WHERE user_id=?", (user_id,))
    conn.commit()
    conn.close()


def get_group_info(user_id: int):
    conn = get_db()
    row = conn.execute(
        "SELECT g.code, g.id, (SELECT COUNT(*) FROM user_group WHERE group_id=g.id) "
        "FROM groups g JOIN user_group ug ON ug.group_id=g.id WHERE ug.user_id=?",
        (user_id,),
    ).fetchone()
    conn.close()
    return row


# ---------- Qarz (debitor/kreditor) ----------
# direction='oldim'  -> foydalanuvchi shu odamdan qarz oldi (u qarzdor)
# direction='berdim' -> foydalanuvchi shu odamga qarz berdi (u sizga qarzdor)

def add_debt(space_key: str, person: str, direction: str, amount: float, currency: str, note: str):
    conn = get_db()
    conn.execute(
        "INSERT INTO debts (space_key, person, direction, amount, currency, note, created_at, settled) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, 0)",
        (space_key, person, direction, amount, currency, note,
         datetime.now().isoformat(timespec="seconds")),
    )
    conn.commit()
    conn.close()


def get_debt_summary(space_key: str):
    conn = get_db()
    rows = conn.execute(
        "SELECT person, direction, currency, SUM(amount) FROM debts "
        "WHERE space_key=? AND settled=0 GROUP BY person, direction, currency",
        (space_key,),
    ).fetchall()
    conn.close()
    summary = {}
    for person, direction, currency, total in rows:
        summary.setdefault(person, {}).setdefault(currency, 0.0)
        if direction == "berdim":
            summary[person][currency] += total
        else:
            summary[person][currency] -= total
    return summary


def build_debts_summary_text(space_key: str) -> str:
    summary = get_debt_summary(space_key)
    lines = ["🤝 Qarzlar holati:"]
    has_open = False
    for person, cur_map in summary.items():
        parts = []
        for cur, net in cur_map.items():
            if abs(net) < 0.01:
                continue
            unit = "so'm" if cur == "UZS" else "$"
            if net > 0:
                parts.append(f"sizga {fmt_num(net)} {unit} qarzdor")
            else:
                parts.append(f"siz {fmt_num(abs(net))} {unit} qarzdorsiz")
        if parts:
            has_open = True
            lines.append(f"  {person}: " + ", ".join(parts))
    if not has_open:
        return "Qarzlar yo'q. 🎉"
    lines.append("\nYopish uchun: /qarz_yopish Ism")
    return "\n".join(lines)


def settle_person(space_key: str, person: str) -> int:
    conn = get_db()
    cur = conn.execute(
        "UPDATE debts SET settled=1 WHERE space_key=? AND LOWER(person)=LOWER(?) AND settled=0",
        (space_key, person),
    )
    changed = cur.rowcount
    conn.commit()
    conn.close()
    return changed


# ---------- Summani tahlil qilish (13 mln, 13 000 000, mingchisiz h.k.) ----------

def parse_amount_token(raw_amount: str, mult_word: str):
    digits = re.sub(r"[\s.,]", "", raw_amount)
    if not digits:
        return None
    try:
        value = float(digits)
    except ValueError:
        return None
    if mult_word:
        mw = mult_word.lower()
        if mw in ("mln", "million", "млн"):
            value *= 1_000_000
        elif mw in ("ming", "минг"):
            value *= 1_000
    return value


PATTERN = re.compile(
    r"^([+-]?)\s*(\d[\d\s.,]*)\s*(mln|million|млн|ming|минг)?\b\s*(\$|usd|so'?m|sum|uzs)?\s*(.*)$",
    re.IGNORECASE,
)


def parse_message(text: str):
    text = text.strip()
    match = PATTERN.match(text)
    if not match:
        return None

    sign, raw_amount, mult_word, currency_raw, note = match.groups()
    amount = parse_amount_token(raw_amount, mult_word)
    if amount is None:
        return None

    currency = "UZS" if (not currency_raw or currency_raw.lower() in ("sum", "so'm", "som", "uzs")) else "USD"
    note = note.strip() or "(izohsiz)"

    if sign == "+":
        is_income = True
    elif sign == "-":
        is_income = False
    else:
        is_income = is_income_text(note)

    amount = abs(amount) if is_income else -abs(amount)
    category = detect_category(note, is_income)
    return amount, currency, category, note


def format_balance_text(space_key: str) -> str:
    bal = get_balance(space_key)
    kurs = get_kurs(space_key)
    total_uzs = bal["UZS"] + bal["USD"] * kurs
    return (
        f"💰 Balans:\n"
        f"UZS: {fmt_num(bal['UZS'])} so'm\n"
        f"USD: {fmt_num(bal['USD'], 2)} $\n"
        f"—\n"
        f"Umumiy (taxminan): {fmt_num(total_uzs)} so'm  (kurs: 1$ = {fmt_num(kurs)} so'm)"
    )


def build_full_status_text(space_key: str) -> str:
    kurs = get_kurs(space_key)
    bal = get_balance(space_key)
    balance_uzs = bal["UZS"] + bal["USD"] * kurs

    debt_summary = get_debt_summary(space_key)
    receivable_uzs = 0.0
    payable_uzs = 0.0
    for person, cur_map in debt_summary.items():
        for cur, net in cur_map.items():
            val_uzs = net * (kurs if cur == "USD" else 1)
            if val_uzs > 0:
                receivable_uzs += val_uzs
            else:
                payable_uzs += -val_uzs

    net_worth = balance_uzs + receivable_uzs - payable_uzs

    lines = [format_balance_text(space_key), ""]
    lines.append(build_debts_summary_text(space_key))
    lines.append("")
    lines.append(
        f"📌 Sof holat (balans + sizga qarzdorlar − siz qarzdorsiz):\n"
        f"{fmt_num(net_worth)} so'm (taxminan)"
    )
    lines.append("")
    lines.append(build_accounts_summary_text(space_key))
    return "\n".join(lines)


# ---------- Tabiiy tildagi qarz yozuvini aniqlash (tugmasiz) ----------
# Masalan: "500000 Alidan qarz oldim" yoki "Valiga 200000 qarz berdim"

FREEFORM_AMOUNT_RE = re.compile(r"(\d[\d\s.,]*\d|\d)\s*(mln|million|млн|ming|минг)?\b", re.IGNORECASE)
PERSON_FROM_RE = re.compile(r"\b([A-Za-zА-Яа-яЎўҚқҒғҲҳ']+?)dan\b", re.IGNORECASE)
PERSON_TO_RE = re.compile(r"\b([A-Za-zА-Яа-яЎўҚқҒғҲҳ']+?)ga\b", re.IGNORECASE)


def try_parse_freeform_debt(text: str):
    if "qarz" not in text.lower():
        return None
    text_lower = text.lower()
    if re.search(r"qarz\s*old", text_lower):
        direction = "oldim"
    elif re.search(r"qarz\s*ber", text_lower):
        direction = "berdim"
    else:
        return None

    m = FREEFORM_AMOUNT_RE.search(text)
    if not m:
        return None
    amount = parse_amount_token(m.group(1), m.group(2))
    if not amount:
        return None

    currency = "USD" if re.search(r"\$|\busd\b", text_lower) else "UZS"

    person = None
    pm = PERSON_FROM_RE.search(text) if direction == "oldim" else PERSON_TO_RE.search(text)
    if pm:
        candidate = pm.group(1)
        if candidate.lower() not in ("qarz", "shu", "u", "meni", "sen"):
            person = candidate

    note = text.strip()
    return person, direction, amount, currency, note


# ---------- Skrinshotdan summani o'qish (OCR) ----------

def extract_amount_candidates(text: str):
    candidates = []

    for m in re.finditer(r"\$\s?(\d{1,5}(?:[.,]\d{1,2})?)", text):
        candidates.append((float(m.group(1).replace(",", ".")), "USD"))
    for m in re.finditer(r"(\d{1,5}(?:[.,]\d{1,2})?)\s?\$", text):
        candidates.append((float(m.group(1).replace(",", ".")), "USD"))

    for m in re.finditer(r"\d{1,3}(?:[ ,.]\d{3}){1,4}", text):
        digits = re.sub(r"[ ,.]", "", m.group(0))
        if digits.isdigit():
            candidates.append((float(digits), "UZS"))
    for m in re.finditer(r"(?<!\d)(\d{4,9})(?!\d)", text):
        candidates.append((float(m.group(1)), "UZS"))

    seen = set()
    unique = []
    for amt, cur in candidates:
        key = (round(amt, 2), cur)
        if key not in seen:
            seen.add(key)
            unique.append((amt, cur))

    unique.sort(key=lambda x: -x[0])
    return unique[:4]


# ---------- Bot buyruqlari ----------

HELP_TEXT = (
    "Salom! Men xarajat/balans botiman.\n\n"
    "Endi tugma bosmasdan ham oddiy gapda yozishingiz mumkin:\n"
    "  -50000 taksi\n"
    "  13 mln maosh oldim\n"
    "  500000 Alidan qarz oldim\n"
    "  200 000 Valiga qarz berdim\n"
    "  50000 taksi karta   (yoki \"naqt\" — to'lov usulini ham belgilaydi)\n\n"
    "Ovozli xabar yuborsangiz ham tushunaman (agar o'rnatilgan bo'lsa).\n"
    "To'lov skrinshotini izoh bilan yuborsangiz, summani o'zim o'qiyman.\n\n"
    "Meni guruhga qo'shsangiz, guruhdagi barcha xarajatlar bitta umumiy "
    "balansga yoziladi (oilaviy hisob uchun qulay).\n"
    "Har kuni soat 23:59 da kunlik balans holatini o'zim yuboraman.\n\n"
    "Pastdagi tugmalar orqali tez foydalaning."
)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(HELP_TEXT, reply_markup=MAIN_KEYBOARD)


async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(HELP_TEXT, reply_markup=MAIN_KEYBOARD)


async def balance(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    await update.message.reply_text(build_full_status_text(space_key))


async def history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    entries = get_combined_history(space_key)
    if not entries:
        await update.message.reply_text("Hali yozuvlar yo'q.")
        return
    lines = ["📜 Oxirgi yozuvlar (xarajat/daromad va qarzlar):"]
    for e in entries:
        try:
            date_str = datetime.fromisoformat(e["created_at"]).strftime("%d.%m.%Y")
        except ValueError:
            date_str = e["created_at"].split("T")[0]
        if e["kind"] == "tx":
            sign = "+" if e["amount"] >= 0 else ""
            unit = "so'm" if e["currency"] == "UZS" else "$"
            lines.append(f"{date_str}  {sign}{fmt_num(e['amount'])} {unit}  [{e['category']}]  — {e['note']}")
        else:
            unit = "so'm" if e["currency"] == "UZS" else "$"
            if e["direction"] == "oldim":
                verb = f"{e['person']} dan {fmt_num(e['amount'])} {unit} qarz oldingiz"
            else:
                verb = f"{e['person']} ga {fmt_num(e['amount'])} {unit} qarz berdingiz"
            lines.append(f"{date_str}  🤝 {verb}  — {e['note']}")
    await update.message.reply_text("\n".join(lines))


def build_stat_chart(space_key: str) -> str:
    kurs = get_kurs(space_key)
    days = 7
    now = datetime.now()
    start = (now - timedelta(days=days - 1)).replace(hour=0, minute=0, second=0, microsecond=0)

    rows = get_period_stats_with_date(space_key, start)
    date_list = [(start + timedelta(days=i)).date() for i in range(days)]
    daily_income = {d: 0.0 for d in date_list}
    daily_expense = {d: 0.0 for d in date_list}

    for amount, currency, created_at in rows:
        d = datetime.fromisoformat(created_at).date()
        if d not in daily_income:
            continue
        val_uzs = amount * (kurs if currency == "USD" else 1)
        if val_uzs >= 0:
            daily_income[d] += val_uzs
        else:
            daily_expense[d] += -val_uzs

    labels = [d.strftime("%d.%m") for d in date_list]
    income_vals = [daily_income[d] for d in date_list]
    expense_vals = [daily_expense[d] for d in date_list]

    debt_summary = get_debt_summary(space_key)
    receivable = 0.0
    payable = 0.0
    for person, cur_map in debt_summary.items():
        for cur, net in cur_map.items():
            val = net * (kurs if cur == "USD" else 1)
            if val > 0:
                receivable += val
            else:
                payable += -val

    cat_rows = get_category_totals(space_key)
    cat_totals_uzs = {}
    for category, currency, total in cat_rows:
        val = abs(total) * (kurs if currency == "USD" else 1)
        cat_totals_uzs[category] = cat_totals_uzs.get(category, 0) + val
    top_cats = sorted(cat_totals_uzs.items(), key=lambda x: -x[1])[:5]

    fig, axes = plt.subplots(1, 3, figsize=(15, 4.8))
    ax1, ax2, ax3 = axes

    x = list(range(len(labels)))
    width = 0.35
    bars_income = ax1.bar([i - width / 2 for i in x], income_vals, width, label="Daromad", color="#2e7d32")
    bars_expense = ax1.bar([i + width / 2 for i in x], expense_vals, width, label="Xarajat", color="#c62828")
    ax1.set_xticks(x)
    ax1.set_xticklabels(labels, rotation=45)
    ax1.set_title("So'nggi 7 kun (so'mda)")
    ax1.legend()
    ax1.grid(axis="y", alpha=0.3)
    ax1.bar_label(bars_income, fmt="%.0f", fontsize=7, rotation=90, padding=2)
    ax1.bar_label(bars_expense, fmt="%.0f", fontsize=7, rotation=90, padding=2)

    bars_debt = ax2.bar(["Sizga\nqarzdor", "Siz\nqarzdor"], [receivable, payable],
                         color=["#2e7d32", "#c62828"])
    ax2.set_title("Qarz holati (so'mda)")
    ax2.grid(axis="y", alpha=0.3)
    ax2.bar_label(bars_debt, fmt="%.0f", fontsize=9, padding=3)

    if top_cats:
        cat_labels = [c for c, v in top_cats]
        cat_values = [v for c, v in top_cats]
        colors = ["#c62828", "#ef6c00", "#f9a825", "#6a1b9a", "#0277bd"][:len(cat_labels)]
        wedges, texts, autotexts = ax3.pie(
            cat_values, labels=cat_labels, autopct=lambda p: f"{p:.0f}%\n({fmt_num(p/100*sum(cat_values))})",
            colors=colors, textprops={"fontsize": 7},
        )
        ax3.set_title("Xarajat kategoriyalari")
    else:
        ax3.axis("off")
        ax3.text(0.5, 0.5, "Xarajat yo'q", ha="center", va="center")

    fig.suptitle("Moliyaviy holat")
    fig.tight_layout()

    out_dir = tempfile.gettempdir()
    filepath = os.path.join(out_dir, f"stat_{space_key}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
    fig.savefig(filepath, dpi=120)
    plt.close(fig)
    return filepath


async def stat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    now = datetime.now()

    periods = [
        ("Bugun", now.replace(hour=0, minute=0, second=0, microsecond=0)),
        ("Shu hafta", (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)),
        ("Shu oy", now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)),
    ]

    lines = []
    for label, start_dt in periods:
        rows = get_period_stats(space_key, start_dt)
        income = {"UZS": 0.0, "USD": 0.0}
        expense = {"UZS": 0.0, "USD": 0.0}
        cat_totals = {}
        for amount, currency, category in rows:
            if amount >= 0:
                income[currency] += amount
            else:
                expense[currency] += -amount
                if currency == "UZS":
                    cat_totals[category] = cat_totals.get(category, 0) + (-amount)

        lines.append(f"📅 {label}:")
        lines.append(f"  Daromad: {fmt_num(income['UZS'])} so'm, {fmt_num(income['USD'], 2)} $")
        lines.append(f"  Xarajat: {fmt_num(expense['UZS'])} so'm, {fmt_num(expense['USD'], 2)} $")
        if cat_totals:
            top = sorted(cat_totals.items(), key=lambda x: -x[1])[:3]
            top_str = ", ".join(f"{c}: {fmt_num(v)}" for c, v in top if v > 0)
            if top_str:
                lines.append(f"  Ko'p sarflangan: {top_str}")
        lines.append("")

    await update.message.reply_text("\n".join(lines))

    if MATPLOTLIB_AVAILABLE:
        try:
            path = build_stat_chart(space_key)
            with open(path, "rb") as f:
                await update.message.reply_photo(
                    photo=f, caption="📊 Daromad/xarajat, qarz holati va kategoriyalar diagrammasi"
                )
            os.remove(path)
        except Exception:
            pass
    else:
        await update.message.reply_text(
            "📊 Diagramma ko'rish uchun: python -m pip install matplotlib --break-system-packages"
        )


async def categories(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    rows = get_category_totals(space_key)
    if not rows:
        await update.message.reply_text("Hali xarajatlar yo'q.")
        return
    lines = ["📊 Kategoriyalar bo'yicha jami xarajatlar:"]
    for category, currency, total in rows:
        unit = "so'm" if currency == "UZS" else "$"
        lines.append(f"  {category}: {fmt_num(abs(total))} {unit}")
    await update.message.reply_text("\n".join(lines))


async def show_method_totals(update: Update, context: ContextTypes.DEFAULT_TYPE, method: str, title: str):
    space_key = get_space_key_for_update(update)
    totals, recent = get_method_totals(space_key, method)
    lines = [f"{title}:"]
    lines.append(f"  UZS: {fmt_num(totals['UZS'])} so'm")
    lines.append(f"  USD: {fmt_num(totals['USD'], 2)} $")
    if recent:
        lines.append("")
        lines.append("So'nggi yozuvlar:")
        for amount, currency, category, note, created_at in recent:
            try:
                date_str = datetime.fromisoformat(created_at).strftime("%d.%m.%Y")
            except ValueError:
                date_str = created_at.split("T")[0]
            unit = "so'm" if currency == "UZS" else "$"
            lines.append(f"  {date_str}  {fmt_num(abs(amount))} {unit}  [{category}] — {note}")
    else:
        lines.append("")
        lines.append("Hali bu usulda yozuv yo'q. Xarajat yozganingizda \"karta\" yoki \"naqt\" so'zini qo'shsangiz, shu yerda ko'rinadi.")
    await update.message.reply_text("\n".join(lines))


async def card_expenses(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await show_method_totals(update, context, "karta", "💳 Karta orqali xarajatlar")


async def cash_expenses(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await show_method_totals(update, context, "naqt", "💵 Naqt xarajatlar")


async def saldo_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    context.chat_data["pending_saldo_input"] = True
    await update.message.reply_text(
        f"{build_accounts_summary_text(space_key)}\n\n"
        f"Yangilash uchun: Hisob nomi va summa yozing, masalan: Uzcard 500000\n"
        f"Hisoblar: {', '.join(ACCOUNT_NAMES)}"
    )


async def kurs_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    cur_kurs = get_kurs(space_key)
    context.chat_data["pending_kurs_input"] = True
    await update.message.reply_text(
        f"Joriy kurs: 1$ = {fmt_num(cur_kurs)} so'm\n\n"
        f"Yangi kursni raqam bilan yozing, masalan: 12700"
    )


async def kurs_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    if not context.args:
        await kurs_menu(update, context)
        return
    try:
        value = float(context.args[0].replace(",", ""))
        if value <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("Noto'g'ri qiymat. Masalan: /kurs 12700")
        return
    set_kurs(space_key, value)
    await update.message.reply_text(f"Kurs yangilandi: 1$ = {fmt_num(value)} so'm ✅")


async def group_new(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    code = create_group(user_id)
    await update.message.reply_text(
        f"👨‍👩‍👧 Jamoaviy balans yaratildi!\n\n"
        f"Taklif kodi: {code}\n\n"
        f"Oila a'zolaringiz shu kodni /group_join {code} bilan kiritsa, "
        f"hammangiz bitta umumiy balansni yuritasiz."
    )


async def group_join(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Kodni kiriting. Masalan: /group_join AB12CD")
        return
    code = context.args[0]
    ok = join_group(update.effective_user.id, code)
    if ok:
        await update.message.reply_text(
            f"✅ Jamoaviy balansga qo'shildingiz!\n\n"
            f"{format_balance_text(get_space_key_for_update(update))}"
        )
    else:
        await update.message.reply_text("Bunday kod topilmadi. Qaytadan tekshirib ko'ring.")


async def group_leave(update: Update, context: ContextTypes.DEFAULT_TYPE):
    leave_group(update.effective_user.id)
    await update.message.reply_text("Jamoaviy balansdan chiqdingiz. Endi shaxsiy balansingiz ishlaydi. 👤")


async def group_info(update: Update, context: ContextTypes.DEFAULT_TYPE):
    info = get_group_info(update.effective_user.id)
    if not info:
        await update.message.reply_text("Siz hozir shaxsiy balansdasiz (jamoaga qo'shilmagansiz).")
        return
    code, group_id, member_count = info
    await update.message.reply_text(
        f"👨‍👩‍👧 Jamoa ma'lumoti:\nKod: {code}\nA'zolar soni: {member_count}"
    )


async def reset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    reset_space(space_key)
    await update.message.reply_text("Balans va tarix tozalandi. 🧹")


# --- Qarz buyruqlari (tugma va matnli) ---

async def debts_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    buttons = [
        [InlineKeyboardButton("➕ Men qarz oldim", callback_data="debtdir:oldim")],
        [InlineKeyboardButton("➕ Men qarz berdim", callback_data="debtdir:berdim")],
        [InlineKeyboardButton("📋 Ro'yxat", callback_data="debtlist")],
    ]
    await update.message.reply_text(
        "Qarzlar bo'limi — tugma bosing, YOKI to'g'ridan-to'g'ri shu tarzda yozing:\n"
        "\"500000 Alidan qarz oldim\" yoki \"200000 Valiga qarz berdim\"",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def debt_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    if data == "debtlist":
        space_key = get_space_key_for_update(update)
        await query.edit_message_text(build_debts_summary_text(space_key))
        return
    direction = data.split(":")[1]
    context.chat_data["pending_debt_direction"] = direction
    if direction == "oldim":
        prompt = "Kimdan va qancha qarz oldingiz? Masalan: Ali 500000 taksi uchun"
    else:
        prompt = "Kimga va qancha qarz berdingiz? Masalan: Vali 200000"
    await query.edit_message_text(prompt)


DEBT_PATTERN = re.compile(
    r"^(\S+)\s+(\d[\d\s.,]*)\s*(mln|million|млн|ming|минг)?\b\s*(\$|usd|so'?m|sum|uzs)?\s*(.*)$",
    re.IGNORECASE,
)


def parse_debt_args(text: str):
    text = text.strip()
    match = DEBT_PATTERN.match(text)
    if not match:
        return None
    person, raw_amount, mult_word, currency_raw, note = match.groups()
    amount = parse_amount_token(raw_amount, mult_word)
    if amount is None:
        return None
    currency = "UZS" if (not currency_raw or currency_raw.lower() in ("sum", "so'm", "som", "uzs")) else "USD"
    note = note.strip() or "(izohsiz)"
    return person, amount, currency, note


async def qarz_oldim_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = " ".join(context.args)
    parsed = parse_debt_args(text) if text else None
    if not parsed:
        await update.message.reply_text("Masalan: /qarz_oldim Ali 500000 taksi uchun")
        return
    person, amount, currency, note = parsed
    add_debt(get_space_key_for_update(update), person, "oldim", amount, currency, note)
    unit = "so'm" if currency == "UZS" else "$"
    await update.message.reply_text(f"Qayd etildi: {person} dan {fmt_num(amount)} {unit} qarz oldingiz.")


async def qarz_berdim_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = " ".join(context.args)
    parsed = parse_debt_args(text) if text else None
    if not parsed:
        await update.message.reply_text("Masalan: /qarz_berdim Vali 200000")
        return
    person, amount, currency, note = parsed
    add_debt(get_space_key_for_update(update), person, "berdim", amount, currency, note)
    unit = "so'm" if currency == "UZS" else "$"
    await update.message.reply_text(f"Qayd etildi: {person} ga {fmt_num(amount)} {unit} qarz berdingiz.")


async def debts_list_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    await update.message.reply_text(build_debts_summary_text(space_key))


async def qarz_yopish_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Ism kiriting. Masalan: /qarz_yopish Ali")
        return
    person = " ".join(context.args)
    space_key = get_space_key_for_update(update)
    changed = settle_person(space_key, person)
    if changed:
        await update.message.reply_text(f"{person} bilan hisob-kitob yopildi. ✅")
    else:
        await update.message.reply_text(f"{person} bilan ochiq qarz topilmadi.")


# --- Excel ---

def build_excel_report(space_key: str) -> str:
    conn = get_db()
    rows = conn.execute(
        "SELECT amount, currency, category, note, created_at FROM transactions "
        "WHERE space_key=? ORDER BY id ASC",
        (space_key,),
    ).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Xarajatlar"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    normal_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", bold=True, size=10)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    green_font = Font(name="Arial", size=10, color="1E7B34")
    red_font = Font(name="Arial", size=10, color="C00000")

    headers = ["Sana", "Turi", "Kategoriya", "Summa", "Valyuta", "Izoh"]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    r = 2
    for amount, currency, category, note, created_at in rows:
        date_str = created_at.split("T")[0]
        kind = "Daromad" if amount >= 0 else "Xarajat"
        font = green_font if amount >= 0 else red_font
        ws.cell(row=r, column=1, value=date_str).font = normal_font
        ws.cell(row=r, column=2, value=kind).font = font
        ws.cell(row=r, column=3, value=category or "").font = normal_font
        amount_cell = ws.cell(row=r, column=4, value=amount)
        amount_cell.font = font
        amount_cell.number_format = "#,##0.00"
        ws.cell(row=r, column=5, value=currency).font = normal_font
        ws.cell(row=r, column=6, value=note).font = normal_font
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = border
        r += 1

    last_data_row = r - 1
    r += 1
    if last_data_row >= 2:
        for currency, offset in (("UZS", 0), ("USD", 1)):
            row_num = r + offset
            ws.cell(row=row_num, column=3, value=f"Jami ({currency})").font = bold_font
            formula = f'=SUMIFS(D2:D{last_data_row},E2:E{last_data_row},"{currency}")'
            total_cell = ws.cell(row=row_num, column=4, value=formula)
            total_cell.font = bold_font
            total_cell.number_format = "#,##0.00"
            for col in range(1, 7):
                ws.cell(row=row_num, column=col).border = border

    widths = {1: 12, 2: 10, 3: 14, 4: 16, 5: 9, 6: 28}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"

    out_dir = tempfile.gettempdir()
    filename = f"balans_{space_key}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(out_dir, filename)
    wb.save(filepath)
    return filepath


# --- Xarajat/daromadni yozib qo'yish (kerak bo'lsa hisob so'raydi) ---

async def commit_expense(update: Update, context: ContextTypes.DEFAULT_TYPE, user_id: int, space_key: str,
                          amount: float, currency: str, category: str, note: str, method: str = None,
                          query=None):
    account = detect_account(note)
    if amount < 0 and not account:
        context.chat_data["pending_account_tx"] = {
            "user_id": user_id, "space_key": space_key, "amount": amount,
            "currency": currency, "category": category, "note": note, "method": method,
        }
        buttons = [[InlineKeyboardButton(name, callback_data=f"acct:{name}")] for name in ACCOUNT_NAMES]
        buttons.append([InlineKeyboardButton("O'tkazib yuborish", callback_data="acct:skip")])
        prompt = "Qaysi hisobdan xarajat qildingiz?"
        if query:
            await query.edit_message_text(prompt, reply_markup=InlineKeyboardMarkup(buttons))
        else:
            await update.message.reply_text(prompt, reply_markup=InlineKeyboardMarkup(buttons))
        return
    await _finalize_commit(update, user_id, space_key, amount, currency, category, note, method, account, query=query)


async def _finalize_commit(update: Update, user_id: int, space_key: str, amount: float, currency: str,
                            category: str, note: str, method: str, account: str, query=None):
    add_transaction(user_id, space_key, amount, currency, category, note, method, account)
    if account:
        # Hisoblar doim UZS ekvivalentida yuritiladi — agar tranzaksiya
        # USD bo'lsa, joriy kursga aylantirib keyin hisobdan ayiramiz/qo'shamiz
        kurs = get_kurs(space_key)
        delta_uzs = amount if currency == "UZS" else amount * kurs
        apply_account_delta(space_key, account, delta_uzs)

    sign = "+" if amount >= 0 else ""
    unit = "so'm" if currency == "UZS" else "$"
    text = f"Qabul qilindi: {sign}{fmt_num(amount)} {unit}  [{category}]"
    if account:
        text += f"  ({account})"
    text += f"  ({note})\n\n{format_balance_text(space_key)}"

    if query:
        await query.edit_message_text(text)
    else:
        await update.message.reply_text(text)


async def account_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    pending = context.chat_data.pop("pending_account_tx", None)
    if not pending:
        await query.edit_message_text("Bu so'rovning muddati o'tgan. Qaytadan yozib ko'ring.")
        return
    account = None if data == "acct:skip" else data.split(":", 1)[1]
    await _finalize_commit(
        update, pending["user_id"], pending["space_key"], pending["amount"], pending["currency"],
        pending["category"], pending["note"], pending["method"], account, query=query,
    )


async def excel_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    rows = get_history(space_key, limit=1)
    if not rows:
        await update.message.reply_text("Hali yozuvlar yo'q, Excel yaratib bo'lmaydi.")
        return

    await update.message.reply_text("📊 Excel fayl tayyorlanmoqda...")
    filepath = build_excel_report(space_key)
    with open(filepath, "rb") as f:
        await update.message.reply_document(
            document=f,
            filename=os.path.basename(filepath),
            caption="Balansingiz bo'yicha to'liq hisobot 📈",
        )
    os.remove(filepath)


# --- Skrinshot (rasm) qabul qilish ---

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not OCR_AVAILABLE:
        await update.message.reply_text(
            "Rasm o'qish funksiyasi hozircha o'rnatilmagan. Yordamdagi yo'riqnomaga qarang, "
            "yoki summani qo'lda yozing."
        )
        return

    photo = update.message.photo[-1]
    file = await photo.get_file()
    tmp_path = os.path.join(
        tempfile.gettempdir(), f"ocr_{update.effective_user.id}_{datetime.now().strftime('%H%M%S%f')}.jpg"
    )
    await file.download_to_drive(tmp_path)

    try:
        img = Image.open(tmp_path)
        text = pytesseract.image_to_string(img)
    except Exception as e:
        print(f"[OCR XATO] {e}")
        await update.message.reply_text(
            "Rasmni o'qib bo'lmadi. Tesseract-OCR to'g'ri o'rnatilganini tekshiring, "
            "yoki summani qo'lda yozing."
        )
        return
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

    candidates = extract_amount_candidates(text)
    caption = (update.message.caption or "").strip()

    if not candidates:
        await update.message.reply_text(
            "Rasmda summani aniqlay olmadim 😕 Iltimos summani qo'lda yozing, masalan: -50000 taksi"
        )
        return

    space_key = get_space_key_for_update(update)
    method = detect_method(caption)

    if len(candidates) == 1 and caption:
        amt, cur = candidates[0]
        category = detect_category(caption, False)
        await commit_expense(update, context, update.effective_user.id, space_key, -abs(amt), cur, category, caption, method)
        return

    buttons = []
    for amt, cur in candidates:
        unit = "so'm" if cur == "UZS" else "$"
        buttons.append([InlineKeyboardButton(f"{fmt_num(amt)} {unit}", callback_data=f"amtpick:{amt}:{cur}")])
    buttons.append([InlineKeyboardButton("✏️ Qo'lda kiritish", callback_data="amtmanual")])
    context.chat_data["pending_photo_caption"] = caption
    await update.message.reply_text(
        "Rasmda bir nechta summa topildi, qaysi biri to'g'ri?",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def photo_amount_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    user_id = query.from_user.id

    if data == "amtmanual":
        context.chat_data.pop("pending_photo_caption", None)
        await query.edit_message_text("Yaxshi, summani qo'lda yozing, masalan: -50000 taksi")
        return

    _, amt_str, cur = data.split(":")
    amt = float(amt_str)
    caption = context.chat_data.pop("pending_photo_caption", "") or ""
    space_key = get_space_key_for_update(update)

    if caption:
        category = detect_category(caption, False)
        method = detect_method(caption)
        await commit_expense(update, context, user_id, space_key, -abs(amt), cur, category, caption, method, query=query)
    else:
        context.chat_data["pending_amount"] = (amt, cur)
        unit = "so'm" if cur == "UZS" else "$"
        await query.edit_message_text(
            f"Summasi: {fmt_num(amt)} {unit}\nEndi nima uchun ekanini yozing (masalan: taksi)"
        )


# --- Salomlashish, botga qaratilganlik va AI yordamchi ---

GREETING_WORDS = [
    "salom", "assalomu alaykum", "assalomu alaykum!", "salomlar", "hi", "hello",
    "hey", "xayrli tong", "xayrli kun", "xayrli kech", "qalaysiz", "yaxshimisiz",
    "salom bot", "привет",
]


def is_greeting(text: str) -> bool:
    t = text.strip().lower().rstrip("!.? ")
    return any(t == g or t.startswith(g) for g in GREETING_WORDS)


QUESTION_WORDS = ["qanaqa", "qancha", "nima", "nimaga", "nega", "kim", "qachon",
                   "qayer", "qayerda", "necha", "qaysi", "qalay", "qanday", "nechta"]


def is_question(text: str) -> bool:
    t = text.lower()
    if "?" in t:
        return True
    return any(re.search(rf"\b{w}\b", t) for w in QUESTION_WORDS)


def is_directed_at_bot(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    chat = update.effective_chat
    if not chat or chat.type not in ("group", "supergroup"):
        return True
    msg = update.message
    text = (msg.text or msg.caption or "") if msg else ""
    bot_username = context.bot.username if context.bot else None
    if bot_username and f"@{bot_username.lower()}" in text.lower():
        return True
    if msg and msg.reply_to_message and msg.reply_to_message.from_user and context.bot:
        if msg.reply_to_message.from_user.id == context.bot.id:
            return True
    return False


async def ask_ai(question: str) -> str:
    """Anthropic API orqali savolga javob oladi (agar ANTHROPIC_API_KEY
    sozlangan bo'lsa). Sekin/bloklovchi chaqiruv bo'lgani uchun alohida
    oqimda (thread) bajariladi."""
    if not AI_AVAILABLE:
        return None
    try:
        import asyncio

        def _call():
            resp = anthropic_client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=600,
                messages=[{"role": "user", "content": question}],
            )
            return resp.content[0].text

        return await asyncio.to_thread(_call)
    except Exception as e:
        print(f"[AI XATO] {e}")
        return None


# --- Markaziy matn qayta ishlash (matn xabar va ovozli xabar shu yerga tushadi) ---

async def process_free_text(update: Update, context: ContextTypes.DEFAULT_TYPE, text: str, silent_if_unrecognized: bool = False):
    text = text.strip()
    user_id = update.effective_user.id
    space_key = get_space_key_for_update(update)

    # 1) Kurs kiritilishi kutilayotgan bo'lsa (faqat "sof raqam" kelsa)
    if context.chat_data.get("pending_kurs_input"):
        cleaned = re.sub(r"[\s,]", "", text)
        if re.fullmatch(r"\d+(\.\d+)?", cleaned):
            value = float(cleaned)
            if value <= 0:
                await update.message.reply_text("Noto'g'ri raqam. Masalan: 12700")
                return
            set_kurs(space_key, value)
            context.chat_data.pop("pending_kurs_input", None)
            await update.message.reply_text(f"Kurs yangilandi: 1$ = {fmt_num(value)} so'm ✅")
            return
        # Raqam emas -> foydalanuvchi boshqa narsa yozgan, kurs so'rovini
        # bekor qilib, xabarni oddiy tarzda qayta ishlaymiz
        context.chat_data.pop("pending_kurs_input", None)

    # 1b) Hisob saldosini kiritish kutilayotgan bo'lsa (faqat hisob nomi bilan boshlansa)
    if context.chat_data.get("pending_saldo_input"):
        if any(text.strip().lower().startswith(name.lower()) for name in ACCOUNT_NAMES):
            parsed_saldo = parse_saldo_input(text)
            if not parsed_saldo:
                await update.message.reply_text(
                    "Tushunmadim. Masalan: Uzcard 500000, yoki Uydagi naqt 200000\n"
                    f"Hisoblar: {', '.join(ACCOUNT_NAMES)}"
                )
                return
            name, value = parsed_saldo
            set_account_balance(space_key, name, value)
            context.chat_data.pop("pending_saldo_input", None)
            await update.message.reply_text(f"{name}: {fmt_num(value)} so'm qilib belgilandi ✅\n\n{build_accounts_summary_text(space_key)}")
            return
        # Hisob nomi bilan boshlanmagan -> saldo so'rovini bekor qilib,
        # xabarni oddiy tarzda qayta ishlaymiz
        context.chat_data.pop("pending_saldo_input", None)

    # 1c) Ertalabki saldo tekshiruviga javob kutilayotgan bo'lsa
    if context.chat_data.get("pending_morning_check"):
        result_text = process_morning_check_reply(user_id, space_key, text)
        if result_text:
            context.chat_data.pop("pending_morning_check", None)
            await update.message.reply_text(result_text)
            return
        # Hech qanday hisob nomi topilmadi -> bu ertalabki javob emas,
        # oddiy xabar sifatida davom ettiramiz (so'rovni bekor qilmaymiz,
        # chunki ehtimol keyinroq javob berishadi)

    # 2) Tugma orqali boshlangan qarz kiritish jarayoni
    pending_dir = context.chat_data.get("pending_debt_direction")
    if pending_dir:
        parsed = parse_debt_args(text)
        if not parsed:
            await update.message.reply_text("Tushunmadim. Masalan: Ali 500000 taksi uchun")
            return
        person, amount, currency, note = parsed
        add_debt(space_key, person, pending_dir, amount, currency, note)
        context.chat_data.pop("pending_debt_direction", None)
        unit = "so'm" if currency == "UZS" else "$"
        verb = "siz qarzdorsiz" if pending_dir == "oldim" else "sizga qarzdor"
        await update.message.reply_text(f"Qayd etildi: {person} — {fmt_num(amount)} {unit} ({verb}) ✅")
        return

    # 3) Tabiiy tilda qarz aniqlandi, lekin ism topilmadi -> ism kutilmoqda
    pending_person_info = context.chat_data.get("pending_debt_awaiting_person")
    if pending_person_info:
        direction, amount, currency, note = pending_person_info
        person = text.strip()
        if not person:
            await update.message.reply_text("Ismni yozing, masalan: Ali")
            return
        add_debt(space_key, person, direction, amount, currency, note)
        context.chat_data.pop("pending_debt_awaiting_person", None)
        unit = "so'm" if currency == "UZS" else "$"
        verb = "siz qarzdorsiz" if direction == "oldim" else "sizga qarzdor"
        await update.message.reply_text(f"Qayd etildi: {person} — {fmt_num(amount)} {unit} ({verb}) ✅")
        return

    # 4) Rasmdan summa tanlangandan keyin izoh kutilayotgan bo'lsa
    pending_amount = context.chat_data.get("pending_amount")
    if pending_amount:
        amt, cur = pending_amount
        note = text or "(izohsiz)"
        category = detect_category(note, False)
        method = detect_method(note)
        context.chat_data.pop("pending_amount", None)
        await commit_expense(update, context, user_id, space_key, -abs(amt), cur, category, note, method)
        return

    # 5) Tabiiy tilda yozilgan qarz xabari ("500000 Alidan qarz oldim")
    freeform_debt = try_parse_freeform_debt(text)
    if freeform_debt:
        person, direction, amount, currency, note = freeform_debt
        if person:
            add_debt(space_key, person, direction, amount, currency, note)
            unit = "so'm" if currency == "UZS" else "$"
            verb = "siz qarzdorsiz" if direction == "oldim" else "sizga qarzdor"
            await update.message.reply_text(f"Qayd etildi: {person} — {fmt_num(amount)} {unit} ({verb}) ✅")
        else:
            context.chat_data["pending_debt_awaiting_person"] = (direction, amount, currency, note)
            question = "Kimdan qarz oldingiz? Ismini yozing." if direction == "oldim" else "Kimga qarz berdingiz? Ismini yozing."
            await update.message.reply_text(question)
        return

    # 6) Oddiy xarajat/daromad yozuvi
    parsed = parse_message(text)
    if not parsed:
        is_q = is_question(text)

        if is_greeting(text) and not is_q:
            # Guruhda oddiy "salom"ga botga aloqasi yo'q bo'lsa jim turamiz
            # (keraksiz javob bermaslik uchun), shaxsiy chatda esa qisqa salom beramiz
            if not silent_if_unrecognized:
                await update.message.reply_text(
                    "Salom! 👋 Xarajat yozmoqchi bo'lsangiz, masalan: -50000 taksi kabi yozing, "
                    "yoki savolingizni yozing."
                )
            return

        # Guruhda botga bevosita murojaat qilinmagan bo'lsa ham, savolga
        # o'xshasa (masalan "ob-havo qanaqa?") AI orqali javob berishga harakat qilamiz
        should_respond = (not silent_if_unrecognized) or is_q
        if not should_respond:
            return

        if AI_AVAILABLE:
            answer = await ask_ai(text)
            if answer:
                await update.message.reply_text(answer)
                return

        if not silent_if_unrecognized:
            await update.message.reply_text(
                "Tushunmadim 🤔 Masalan shunday yozing: -50000 taksi, 13 mln maosh, yoki 500000 Alidan qarz oldim"
            )
        # Guruhda savolga o'xshasa-yu AI yoq/javob berolmasa ham, keraksiz
        # "Tushunmadim" bilan spam qilmaymiz — jim qolamiz
        return

    amount, currency, category, note = parsed
    method = detect_method(note)
    await commit_expense(update, context, user_id, space_key, amount, currency, category, note, method)


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()

    button_routes = {
        BTN_BALANCE: balance,
        BTN_HISTORY: history,
        BTN_STAT: stat,
        BTN_CATEGORIES: categories,
        BTN_DEBTS: debts_menu,
        BTN_KURS: kurs_menu,
        BTN_CARD: card_expenses,
        BTN_CASH: cash_expenses,
        BTN_SALDO: saldo_menu,
        BTN_EXCEL: excel_export,
        BTN_HELP: help_cmd,
    }
    if text in button_routes:
        await button_routes[text](update, context)
        return

    directed = is_directed_at_bot(update, context)
    await process_free_text(update, context, text, silent_if_unrecognized=not directed)


# --- Ovozli xabar ---

async def handle_voice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not VOICE_AVAILABLE:
        await update.message.reply_text(
            "Ovozli xabarni tushunish funksiyasi hozircha o'rnatilmagan. "
            "Yordamdagi yo'riqnomaga qarang, yoki matn bilan yozing."
        )
        return

    voice = update.message.voice
    file = await voice.get_file()
    base = os.path.join(tempfile.gettempdir(), f"voice_{update.effective_user.id}_{datetime.now().strftime('%H%M%S%f')}")
    ogg_path = base + ".ogg"
    wav_path = base + ".wav"
    await file.download_to_drive(ogg_path)

    text = None
    try:
        audio = AudioSegment.from_file(ogg_path)
        # Ovoz balandligini me'yorlashtirish (juda past yoki baland ovozni tekislaydi)
        # va aniqlikni oshirish uchun mono + 16kHz formatga o'tkazish
        audio = audio.normalize() if hasattr(audio, "normalize") else audio
        audio = audio.set_channels(1).set_frame_rate(16000)
        audio.export(wav_path, format="wav")
        recognizer = sr.Recognizer()
        recognizer.energy_threshold = 300
        recognizer.dynamic_energy_threshold = True
        with sr.AudioFile(wav_path) as source:
            recognizer.adjust_for_ambient_noise(source, duration=0.3)
            audio_data = recognizer.record(source)
        try:
            text = recognizer.recognize_google(audio_data, language="uz-UZ")
        except Exception as e:
            print(f"[OVOZ XATO] uz-UZ tanib bo'lmadi: {e}")
            try:
                text = recognizer.recognize_google(audio_data, language="ru-RU")
            except Exception as e2:
                print(f"[OVOZ XATO] ru-RU ham tanib bo'lmadi: {e2}")
                text = None
    except Exception as e:
        print(f"[OVOZ XATO] audio qayta ishlashda xato: {e}")
        text = None
    finally:
        for p in (ogg_path, wav_path):
            if os.path.exists(p):
                os.remove(p)

    if not text:
        await update.message.reply_text(
            "Ovozli xabarni tanib bo'lmadi 😕 Iloji bo'lsa jimroq joyda, mikrofonga yaqinroq turib "
            "qayta gapirib ko'ring, yoki matn bilan yozing."
        )
        return

    await update.message.reply_text(f"🎙 Eshitdim: \"{text}\"")
    await process_free_text(update, context, text)


async def send_daily_status(context: ContextTypes.DEFAULT_TYPE):
    """Har kuni 23:59 da (Toshkent vaqti bilan) har bir ma'lum chatga
    o'sha kunning balans holatini avtomatik yuboradi."""
    for chat_id, space_key in get_all_known_chats():
        try:
            text = "🌙 Kun yakuni — bugungi balans holati:\n\n" + build_full_status_text(space_key)
            await context.bot.send_message(chat_id=chat_id, text=text)
        except Exception as e:
            print(f"[KUNLIK XABAR XATO] chat_id={chat_id}: {e}")


async def send_morning_check(context: ContextTypes.DEFAULT_TYPE):
    """Har kuni ertalab har bir ma'lum chatdan haqiqiy hisob (Uzcard, Humo,
    Visa, naqt) qoldiqlarini so'raydi, javob kelganda bot yozgan hisob
    bilan solishtirib, farqni "Prochee rasxod/daromad" sifatida qayd etadi."""
    for chat_id, space_key in get_all_known_chats():
        try:
            await context.bot.send_message(
                chat_id=chat_id,
                text=(
                    "🌅 Xayrli tong! Hisoblaringizdagi HAQIQIY qoldiqni tekshirib, "
                    "har birini alohida qatorda yozing, masalan:\n\n"
                    "Uzcard 350000\nHumo 120000\nVisa 0\nUydagi naqt 200000\nOldimdagi naqt 50000\n\n"
                    "Agar mendagi hisob bilan farq bo'lsa, avtomatik \"Prochee rasxod/daromad\" "
                    "sifatida qayd etib, hisobni to'g'rilab qo'yaman."
                ),
            )
            context.application.chat_data[chat_id]["pending_morning_check"] = True
        except Exception as e:
            print(f"[ERTALABKI SO'ROV XATO] chat_id={chat_id}: {e}")


def process_morning_check_reply(user_id: int, space_key: str, text: str):
    """Ertalabki xabarga javobni qator-qator tahlil qiladi, farqlarni
    yozib, hisoblarni haqiqiy qiymatga tenglashtiradi. Xabar matnini qaytaradi."""
    lines_in = [ln.strip() for ln in text.splitlines() if ln.strip()]
    results = []
    matched_any = False
    for line in lines_in:
        parsed = parse_saldo_input(line)
        if not parsed:
            continue
        matched_any = True
        name, actual = parsed
        current_rows = dict(get_accounts(space_key))
        expected = current_rows.get(name, 0.0)
        diff = expected - actual  # musbat -> pulingiz yo'q bo'lib qolgan (xarajat)
        if abs(diff) > 0.5:
            if diff > 0:
                amount = -diff
                category = "Prochee rasxod"
            else:
                amount = -diff
                category = "Prochee daromad"
            add_transaction(user_id, space_key, amount, "UZS", category,
                             f"Avtomatik: {name} saldo farqi", None, name)
            results.append(f"  {name}: kutilgan {fmt_num(expected)}, haqiqiy {fmt_num(actual)} — "
                            f"farq {fmt_num(abs(diff))} so'm [{category}] sifatida yozildi")
        else:
            results.append(f"  {name}: mos keladi ({fmt_num(actual)} so'm)")
        set_account_balance(space_key, name, actual)

    if not matched_any:
        return None
    return "✅ Ertalabki tekshiruv natijasi:\n" + "\n".join(results) + f"\n\n{build_accounts_summary_text(space_key)}"


def main():
    token = os.environ.get("BOT_TOKEN")
    if not token:
        raise SystemExit("BOT_TOKEN muhit o'zgaruvchisi topilmadi. export BOT_TOKEN=... qiling.")

    app = ApplicationBuilder().token(token).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_cmd))
    app.add_handler(CommandHandler("balance", balance))
    app.add_handler(CommandHandler("history", history))
    app.add_handler(CommandHandler("stat", stat))
    app.add_handler(CommandHandler("categories", categories))
    app.add_handler(CommandHandler("kurs", kurs_cmd))
    app.add_handler(CommandHandler("group_new", group_new))
    app.add_handler(CommandHandler("group_join", group_join))
    app.add_handler(CommandHandler("group_leave", group_leave))
    app.add_handler(CommandHandler("group_info", group_info))
    app.add_handler(CommandHandler("qarz_oldim", qarz_oldim_cmd))
    app.add_handler(CommandHandler("qarz_berdim", qarz_berdim_cmd))
    app.add_handler(CommandHandler("qarzlar", debts_list_cmd))
    app.add_handler(CommandHandler("qarz_yopish", qarz_yopish_cmd))
    app.add_handler(CommandHandler("reset", reset))
    app.add_handler(CommandHandler("excel", excel_export))
    app.add_handler(CallbackQueryHandler(debt_callback, pattern=r"^debtdir:|^debtlist$"))
    app.add_handler(CallbackQueryHandler(photo_amount_callback, pattern=r"^amtpick:|^amtmanual$"))
    app.add_handler(CallbackQueryHandler(account_callback, pattern=r"^acct:"))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.VOICE, handle_voice))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    if app.job_queue is not None:
        try:
            from zoneinfo import ZoneInfo
            tz = ZoneInfo("Asia/Tashkent")
            app.job_queue.run_daily(send_daily_status, time=dt_time(23, 59, tzinfo=tz))
            app.job_queue.run_daily(send_morning_check, time=dt_time(8, 0, tzinfo=tz))
            print("Kunlik xabar (23:59) va ertalabki so'rov (08:00, Toshkent vaqti) sozlandi.")
        except Exception as e:
            print(f"[OGOHLANTIRISH] Kunlik/ertalabki xabarni sozlab bo'lmadi: {e}")
    else:
        print("[OGOHLANTIRISH] job_queue mavjud emas — 'python-telegram-bot[job-queue]' o'rnatilganini tekshiring.")

    print("Bot ishga tushdi...")
    app.run_polling()


if __name__ == "__main__":
    main()
