"""
Xarajatlar/Balans Telegram boti (to'liq versiya 2)
------------------------------------------------------
Endi tugma bosmasdan ham oddiy gapda yozishingiz mumkin:
    "500000 Alidan qarz oldim"
    "13 mln maosh oldim"
    "200 000 Valiga qarz berdim"
Ovozli xabar yuborsangiz ham, bot uni matnga aylantirib xuddi shunday
qayta ishlaydi (agar tegishli dastur o'rnatilgan bo'lsa).

Oddiy xabar formatlari:
    -50000 taksi          -> UZS xarajat
    50000 nonushta         -> belgi bo'lmasa, XARAJAT deb hisoblanadi
    13 mln maosh            -> "maosh" so'zi borligi uchun avtomatik DAROMAD
    +200000 maosh           -> aniq belgi bilan ham yozish mumkin
    13000000 yoki 13 mln    -> ikkalasi ham bir xil summa deb tushuniladi

Buyruqlar:
    /start, /help     - yo'riqnoma va pastki menyu
    /balance          - balans + qarzlar + sof holat
    /history          - sana bilan birlashtirilgan tarix
    /stat             - matnli va diagrammali statistika
    /categories       - kategoriyalar bo'yicha jami xarajat
    /kurs 12700       - USD kursini belgilash (yoki 💱 Kurs tugmasi)
    /group_new, /group_join KOD, /group_leave, /group_info
    /qarz_oldim Ism summa izoh, /qarz_berdim Ism summa izoh
    /qarzlar, /qarz_yopish Ism
    /excel, /reset

O'rnatish (asosiy):
    pip install python-telegram-bot openpyxl matplotlib --break-system-packages

Qo'shimcha (ixtiyoriy funksiyalar uchun):
    Skrinshot o'qish:  pip install pytesseract Pillow --break-system-packages
                       + Tesseract-OCR dasturi (https://github.com/UB-Mannheim/tesseract/wiki)
    Ovozli xabar:      pip install SpeechRecognition pydub --break-system-packages
                       + ffmpeg dasturi (https://www.gyan.dev/ffmpeg/builds/, PATH'ga qo'shing)
    Bular o'rnatilmasa ham, bot qolgan hamma funksiyalar bilan ishlayveradi.

Ishga tushirish:
    export BOT_TOKEN="sizning_bot_tokeningiz"
    python3 expense_bot.py

Bot tokenini olish: Telegram-da @BotFather ga yozing -> /newbot
"""

import os
import re
import random
import string
import tempfile
import sqlite3
import html as html_module
from datetime import datetime, timedelta, time as dt_time

from telegram import (
    Update,
    ReplyKeyboardMarkup,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
)
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    Defaults,
    filters,
)
from telegram.constants import ParseMode

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import pytesseract
    from PIL import Image

    OCR_AVAILABLE = True
    _tcmd = os.environ.get("TESSERACT_CMD")
    if _tcmd:
        pytesseract.pytesseract.tesseract_cmd = _tcmd
    elif os.name == "nt":
        _default_win_path = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        if os.path.exists(_default_win_path):
            pytesseract.pytesseract.tesseract_cmd = _default_win_path
except ImportError as e:
    OCR_AVAILABLE = False
    print(f"[OGOHLANTIRISH] OCR o'rnatilmagan: {e}")

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    MATPLOTLIB_AVAILABLE = True
except ImportError as e:
    MATPLOTLIB_AVAILABLE = False
    print(f"[OGOHLANTIRISH] matplotlib o'rnatilmagan: {e}")

try:
    import speech_recognition as sr
    from pydub import AudioSegment

    VOICE_AVAILABLE = True
except ImportError as e:
    VOICE_AVAILABLE = False
    print(f"[OGOHLANTIRISH] Ovoz kutubxonalari o'rnatilmagan: {e}")

try:
    import anthropic

    _anthropic_key = os.environ.get("ANTHROPIC_API_KEY")
    if _anthropic_key:
        anthropic_client = anthropic.Anthropic(api_key=_anthropic_key)
        AI_AVAILABLE = True
    else:
        anthropic_client = None
        AI_AVAILABLE = False
except ImportError as e:
    anthropic_client = None
    AI_AVAILABLE = False
    print(f"[OGOHLANTIRISH] AI (anthropic) kutubxonasi o'rnatilmagan: {e}")

print(f"[STATUS] OCR_AVAILABLE={OCR_AVAILABLE}, MATPLOTLIB_AVAILABLE={MATPLOTLIB_AVAILABLE}, VOICE_AVAILABLE={VOICE_AVAILABLE}, AI_AVAILABLE={AI_AVAILABLE}")

_data_dir = os.environ.get("DATA_DIR") or os.path.dirname(os.path.abspath(__file__))
os.makedirs(_data_dir, exist_ok=True)
DB_PATH = os.path.join(_data_dir, "expenses.db")
print(f"[BAZA] DB_PATH = {DB_PATH}  (mavjud: {os.path.exists(DB_PATH)})")
print("[BAZA] DIQQAT: bu fayl DATA_DIR muhit o'zgaruvchisi ko'rsatgan doimiy "
      "(persistent) diskda turishi shart, aks holda har bir yangi deploy'da "
      "server konteyneri qayta yaratilib, shu fayl bilan birga hisoblar ham yo'qoladi.")
DEFAULT_KURS = 12700.0  # taxminiy 1 USD = necha UZS

try:
    from zoneinfo import ZoneInfo
    TASHKENT_TZ = ZoneInfo("Asia/Tashkent")
except Exception:
    TASHKENT_TZ = None


def now_tz() -> datetime:
    """Bot har doim O'zbekiston (Toshkent) vaqti bilan ishlashi uchun."""
    if TASHKENT_TZ:
        return datetime.now(TASHKENT_TZ)
    return datetime.now()

# ---------- Pastki menyu tugmalari ----------

BTN_BALANCE = "💰 Balans"
BTN_HISTORY = "📜 Tarix"
BTN_STAT = "📊 Statistika"
BTN_CATEGORIES = "🗂 Kategoriyalar"
BTN_DEBTS = "🤝 Qarzlar"
BTN_KURS = "💱 Kurs"
BTN_CARD = "💳 Karta xarajati"
BTN_CASH = "💵 Naqt xarajat"
BTN_SALDO = "💳 Hisoblar saldosi"
BTN_TRANSFER = "🔁 Hisobdan hisobga"
BTN_MANAGE_ACCOUNTS = "⚙️ Hisoblarni boshqarish"
BTN_CALENDAR = "📅 Kalendar"
BTN_EXCEL = "📥 Excel"
BTN_RESET = "🧹 Tozalash"
BTN_EDIT = "✏️ Tuzatish"
BTN_START = "▶️ Start"
BTN_HELP = "❓ Yordam"

MAIN_KEYBOARD = ReplyKeyboardMarkup(
    [
        [BTN_BALANCE, BTN_HISTORY],
        [BTN_STAT, BTN_CATEGORIES],
        [BTN_DEBTS, BTN_KURS],
        [BTN_CARD, BTN_CASH],
        [BTN_SALDO, BTN_TRANSFER],
        [BTN_MANAGE_ACCOUNTS, BTN_CALENDAR],
        [BTN_EXCEL, BTN_EDIT],
        [BTN_RESET],
        [BTN_START, BTN_HELP],
    ],
    resize_keyboard=True,
)

# ---------- Kategoriyalarni aniqlash ----------

CATEGORY_KEYWORDS = {
    "Oziq-ovqat": ["ovqat", "tushlik", "nonushta", "kechki", "restoran", "kafe",
                   "oshxona", "market", "magazin", "do'kon", "dokon", "un", "non",
                   "gosht", "go'sht", "sabzavot", "meva"],
    "Transport": ["taksi", "avtobus", "metro", "benzin", "yoqilg'i", "yoqilgi",
                  "mashina", "yo'l", "yol", "parkovka"],
    "Kommunal": ["svet", "elektr", "gaz", "suv", "kommunal", "ijara", "kvartira",
                 "kommunalka"],
    "Aloqa": ["mobil", "telefon", "internet", "wifi", "balans", "uzmobile",
              "beeline", "ucell"],
    "Sog'liq": ["dori", "shifokor", "klinika", "gospital", "dorixona", "vrach"],
    "Kiyim": ["kiyim", "futbolka", "poyabzal", "shim", "куртка", "kurtka"],
    "Ta'lim": ["kurs", "kitob", "maktab", "universitet", "repetitor", "darslik"],
    "O'yin-kulgi": ["kino", "konsert", "o'yin", "oyin", "bar", "klub", "disko"],
    "Daromad": ["maosh", "oylik", "bonus", "sovg'a", "sovga", "foyda", "kirim"],
}


def esc(text) -> str:
    """HTML formatlashda foydalanuvchi yozgan matn (izoh, ism) xavfsiz
    ko'rinishi uchun maxsus belgilarni ekranlaydi."""
    return html_module.escape(str(text))


def fmt_num(value: float, decimals: int = 0) -> str:
    """Katta raqamlarni bo'shliq bilan ajratib chiqaradi: 13 000 000
    (vergul bilan emas — ko'p nol bo'lganda chalkashtirmasin uchun)."""
    s = f"{value:,.{decimals}f}"
    return s.replace(",", " ")


def detect_category(note: str, is_income: bool) -> str:
    note_lower = note.lower()
    for category, keywords in CATEGORY_KEYWORDS.items():
        for kw in keywords:
            if kw in note_lower:
                return category
    # Ro'yxatdagi kategoriyalarga mos kelmasa, izohning o'zidagi so'zni
    # kategoriya sifatida olamiz (masalan "suv" -> "Suv")
    skip_words = {"karta", "kart", "naqt", "naqd", "uzcard", "humo", "visa",
                  "uydagi", "oldimdagi", "izohsiz"}
    words = re.findall(r"[A-Za-zА-Яа-яЎўЎ'ʼ]+", note)
    for w in words:
        if w.lower() not in skip_words and len(w) > 1:
            return w.capitalize()
    return "Daromad (boshqa)" if is_income else "Boshqa"


def is_income_text(note: str) -> bool:
    note_lower = note.lower()
    return any(kw in note_lower for kw in CATEGORY_KEYWORDS["Daromad"])


# ---------- Ma'lumotlar bazasi ----------

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            space_key TEXT NOT NULL,
            amount REAL NOT NULL,
            currency TEXT NOT NULL,
            category TEXT,
            note TEXT,
            created_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL,
            created_by INTEGER NOT NULL,
            created_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS user_group (
            user_id INTEGER PRIMARY KEY,
            group_id INTEGER NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS settings (
            space_key TEXT NOT NULL,
            key TEXT NOT NULL,
            value TEXT NOT NULL,
            PRIMARY KEY (space_key, key)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS debts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            space_key TEXT NOT NULL,
            person TEXT NOT NULL,
            direction TEXT NOT NULL,
            amount REAL NOT NULL,
            currency TEXT NOT NULL,
            note TEXT,
            created_at TEXT NOT NULL,
            settled INTEGER NOT NULL DEFAULT 0
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS known_chats (
            chat_id INTEGER PRIMARY KEY,
            space_key TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS accounts (
            space_key TEXT NOT NULL,
            name TEXT NOT NULL,
            balance REAL NOT NULL DEFAULT 0,
            PRIMARY KEY (space_key, name)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS account_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            space_key TEXT NOT NULL,
            name TEXT NOT NULL,
            balance REAL NOT NULL,
            created_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS known_users (
            user_id INTEGER PRIMARY KEY,
            full_name TEXT,
            updated_at TEXT NOT NULL
        )
        """
    )
    _migrate_old_schema(conn)
    return conn


def remember_user(user_id: int, full_name: str):
    """Guruhda kim qanday xarajat yozganini keyinroq (Excel'da) ism bilan
    ko'rsatish uchun foydalanuvchi ismini eslab qoladi."""
    if not user_id:
        return
    conn = get_db()
    conn.execute(
        "INSERT INTO known_users (user_id, full_name, updated_at) VALUES (?, ?, ?) "
        "ON CONFLICT(user_id) DO UPDATE SET full_name=excluded.full_name, updated_at=excluded.updated_at",
        (user_id, full_name or f"Foydalanuvchi {user_id}", now_tz().isoformat(timespec="seconds")),
    )
    conn.commit()
    conn.close()


def get_user_totals(space_key: str):
    """Shu balans (space_key) ichida har bir foydalanuvchi qancha xarajat
    va daromad qilganini qaytaradi — guruhda kim qancha sarflaganini
    ko'rsatish uchun."""
    conn = get_db()
    rows = conn.execute(
        "SELECT t.user_id, u.full_name, t.currency, "
        "SUM(CASE WHEN t.amount<0 THEN -t.amount ELSE 0 END) AS expense, "
        "SUM(CASE WHEN t.amount>=0 THEN t.amount ELSE 0 END) AS income "
        "FROM transactions t LEFT JOIN known_users u ON u.user_id=t.user_id "
        "WHERE t.space_key=? GROUP BY t.user_id, t.currency",
        (space_key,),
    ).fetchall()
    conn.close()
    return rows


def _migrate_old_schema(conn):
    cols = {row[1] for row in conn.execute("PRAGMA table_info(transactions)").fetchall()}
    if "space_key" not in cols:
        conn.execute("ALTER TABLE transactions ADD COLUMN space_key TEXT")
        conn.execute("UPDATE transactions SET space_key='u' || user_id WHERE space_key IS NULL")
    if "category" not in cols:
        conn.execute("ALTER TABLE transactions ADD COLUMN category TEXT")
        conn.execute("UPDATE transactions SET category='Boshqa' WHERE category IS NULL")
    if "method" not in cols:
        conn.execute("ALTER TABLE transactions ADD COLUMN method TEXT")
    if "account" not in cols:
        conn.execute("ALTER TABLE transactions ADD COLUMN account TEXT")
    if "user_name" not in cols:
        conn.execute("ALTER TABLE transactions ADD COLUMN user_name TEXT")
    conn.commit()


ACCOUNT_NAMES = ["Uzcard", "Humo", "Visa", "Uydagi naqt", "Oldimdagi naqt"]
ACCOUNT_ALIASES = {
    "uzcard": "Uzcard", "uz card": "Uzcard",
    "humo": "Humo",
    "visa": "Visa",
    "uydagi naqt": "Uydagi naqt", "uyda": "Uydagi naqt", "uydagi": "Uydagi naqt",
    "oldimdagi naqt": "Oldimdagi naqt", "cepdagi": "Oldimdagi naqt",
    "hamyondagi": "Oldimdagi naqt", "yonimdagi": "Oldimdagi naqt", "cho'ntakdagi": "Oldimdagi naqt",
}


def detect_account(note: str, space_key: str = None):
    note_lower = note.lower()
    # Avval joriy (foydalanuvchi o'zi nomlagan/qo'shgan) hisoblarni tekshiramiz
    if space_key:
        for name in get_account_names(space_key):
            if name.lower() in note_lower:
                return name
    # Keyin standart qisqartmalarni (uzcard/humo/naqt va h.k.) tekshiramiz
    for alias, name in sorted(ACCOUNT_ALIASES.items(), key=lambda kv: -len(kv[0])):
        if alias in note_lower:
            return name
    return None


def ensure_default_accounts(space_key: str):
    """Standart 5 ta hisobni FAQAT shu balans hali umuman hisobga ega
    bo'lmagan holatda yaratadi — aks holda, foydalanuvchi bir hisobni
    nomini o'zgartirganda yoki o'chirganda, u avtomatik qayta tiklanib
    qolmasin."""
    conn = get_db()
    count = conn.execute(
        "SELECT COUNT(*) FROM accounts WHERE space_key=?", (space_key,)
    ).fetchone()[0]
    if count == 0:
        for name in ACCOUNT_NAMES:
            conn.execute(
                "INSERT OR IGNORE INTO accounts (space_key, name, balance) VALUES (?, ?, 0)",
                (space_key, name),
            )
        conn.commit()
    conn.close()


def get_accounts(space_key: str):
    ensure_default_accounts(space_key)
    conn = get_db()
    rows = conn.execute(
        "SELECT name, balance FROM accounts WHERE space_key=? ORDER BY rowid", (space_key,)
    ).fetchall()
    conn.close()
    return rows


def get_account_names(space_key: str):
    return [name for name, _ in get_accounts(space_key)]


def rename_account(space_key: str, old_name: str, new_name: str) -> bool:
    """Foydalanuvchi hisob nomini o'zi xohlagancha o'zgartirishi uchun —
    kod yoki Railway/GitHub'ga kirish shart emas."""
    new_name = new_name.strip()
    if not new_name:
        return False
    conn = get_db()
    exists = conn.execute(
        "SELECT 1 FROM accounts WHERE space_key=? AND name=?", (space_key, old_name)
    ).fetchone()
    if not exists:
        conn.close()
        return False
    conn.execute(
        "UPDATE accounts SET name=? WHERE space_key=? AND name=?", (new_name, space_key, old_name)
    )
    conn.execute(
        "UPDATE transactions SET account=? WHERE space_key=? AND account=?",
        (new_name, space_key, old_name),
    )
    conn.execute(
        "UPDATE account_history SET name=? WHERE space_key=? AND name=?",
        (new_name, space_key, old_name),
    )
    conn.commit()
    conn.close()
    return True


def add_custom_account(space_key: str, name: str) -> bool:
    """Foydalanuvchi o'zi xohlagan yangi hisob (masalan boshqa bank kartasi)
    qo'shishi uchun — kod yozish shart emas."""
    name = name.strip()
    if not name:
        return False
    ensure_default_accounts(space_key)
    conn = get_db()
    existing = conn.execute(
        "SELECT 1 FROM accounts WHERE space_key=? AND name=?", (space_key, name)
    ).fetchone()
    if existing:
        conn.close()
        return False
    conn.execute(
        "INSERT INTO accounts (space_key, name, balance) VALUES (?, ?, 0)", (space_key, name)
    )
    conn.commit()
    conn.close()
    log_account_snapshot(space_key, name, 0.0)
    return True


def delete_account(space_key: str, name: str) -> bool:
    conn = get_db()
    cur = conn.execute("DELETE FROM accounts WHERE space_key=? AND name=?", (space_key, name))
    changed = cur.rowcount
    conn.commit()
    conn.close()
    return changed > 0


def log_account_snapshot(space_key: str, name: str, balance: float):
    """Har bir hisob o'zgarishini vaqti bilan yozib boradi — shu orqali
    istalgan o'tgan sanadagi hisob qoldig'ini qayta tiklash mumkin bo'ladi."""
    conn = get_db()
    conn.execute(
        "INSERT INTO account_history (space_key, name, balance, created_at) VALUES (?, ?, ?, ?)",
        (space_key, name, balance, now_tz().isoformat(timespec="seconds")),
    )
    conn.commit()
    conn.close()


def set_account_balance(space_key: str, name: str, value: float):
    ensure_default_accounts(space_key)
    conn = get_db()
    conn.execute(
        "UPDATE accounts SET balance=? WHERE space_key=? AND name=?", (value, space_key, name)
    )
    conn.commit()
    conn.close()
    log_account_snapshot(space_key, name, value)


def apply_account_delta(space_key: str, name: str, delta: float):
    ensure_default_accounts(space_key)
    conn = get_db()
    conn.execute(
        "UPDATE accounts SET balance = balance + ? WHERE space_key=? AND name=?",
        (delta, space_key, name),
    )
    conn.commit()
    row = conn.execute(
        "SELECT balance FROM accounts WHERE space_key=? AND name=?", (space_key, name)
    ).fetchone()
    conn.close()
    if row:
        log_account_snapshot(space_key, name, row[0])


def get_account_balances_as_of(space_key: str, date_str: str):
    """<date_str> kunining OXIRIGACHA bo'lgan har bir hisob qoldig'ini
    (eng so'nggi ma'lum qiymatni) qaytaradi."""
    end_of_day = date_str + "T23:59:59"
    result = {}
    conn = get_db()
    for name in get_account_names(space_key):
        row = conn.execute(
            "SELECT balance FROM account_history WHERE space_key=? AND name=? AND created_at<=? "
            "ORDER BY created_at DESC, id DESC LIMIT 1",
            (space_key, name, end_of_day),
        ).fetchone()
        result[name] = row[0] if row else 0.0
    conn.close()
    return result


def get_total_expense(space_key: str):
    conn = get_db()
    rows = conn.execute(
        "SELECT currency, SUM(amount) FROM transactions WHERE space_key=? AND amount<0 GROUP BY currency",
        (space_key,),
    ).fetchall()
    conn.close()
    totals = {"UZS": 0.0, "USD": 0.0}
    for currency, total in rows:
        totals[currency] = abs(total or 0.0)
    return totals


def get_total_income(space_key: str):
    conn = get_db()
    rows = conn.execute(
        "SELECT currency, SUM(amount) FROM transactions WHERE space_key=? AND amount>=0 GROUP BY currency",
        (space_key,),
    ).fetchall()
    conn.close()
    totals = {"UZS": 0.0, "USD": 0.0}
    for currency, total in rows:
        totals[currency] = total or 0.0
    return totals


def get_today_total_expense(space_key: str):
    now = now_tz()
    start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    conn = get_db()
    rows = conn.execute(
        "SELECT currency, SUM(amount) FROM transactions WHERE space_key=? AND amount<0 AND created_at>=? GROUP BY currency",
        (space_key, start.isoformat(timespec="seconds")),
    ).fetchall()
    conn.close()
    totals = {"UZS": 0.0, "USD": 0.0}
    for currency, total in rows:
        totals[currency] = abs(total or 0.0)
    return totals


def build_post_transaction_summary(space_key: str) -> str:
    """Xarajat/daromad yozilgandan keyin ko'rsatiladigan qisqa xabar —
    umumiy xarajat, bugungi xarajat, umumiy daromad, va hisoblar qoldig'i."""
    total_exp = get_total_expense(space_key)
    today_exp = get_today_total_expense(space_key)
    total_inc = get_total_income(space_key)

    lines = []
    line1 = f"💸 Umumiy xarajat: <b>{fmt_num(total_exp['UZS'])}</b> so'm"
    if total_exp["USD"] > 0:
        line1 += f", <b>{fmt_num(total_exp['USD'], 2)}</b> $"
    lines.append(line1)

    line2 = f"📆 Bugungi xarajat: <b>{fmt_num(today_exp['UZS'])}</b> so'm"
    if today_exp["USD"] > 0:
        line2 += f", <b>{fmt_num(today_exp['USD'], 2)}</b> $"
    lines.append(line2)

    line3 = f"📈 Umumiy daromad: <b>{fmt_num(total_inc['UZS'])}</b> so'm"
    if total_inc["USD"] > 0:
        line3 += f", <b>{fmt_num(total_inc['USD'], 2)}</b> $"
    lines.append(line3)

    lines.append("")
    lines.append(build_accounts_inline_text(space_key))
    return "\n".join(lines)


def build_balance_view_text(space_key: str) -> str:
    total_exp = get_total_expense(space_key)
    today_exp = get_today_total_expense(space_key)

    lines = ["💸 <b>Xarajatlar</b>"]
    total_line = f"▫️ Umumiy: <b>{fmt_num(total_exp['UZS'])}</b> so'm"
    if total_exp["USD"] > 0:
        total_line += f", <b>{fmt_num(total_exp['USD'], 2)}</b> $"
    lines.append(total_line)

    today_line = f"▫️ Bugungi: <b>{fmt_num(today_exp['UZS'])}</b> so'm"
    if today_exp["USD"] > 0:
        today_line += f", <b>{fmt_num(today_exp['USD'], 2)}</b> $"
    lines.append(today_line)

    lines.append("")
    lines.append(build_accounts_summary_text(space_key))
    return "\n".join(lines)


def get_today_account_expense(space_key: str):
    """Bugun har bir hisobdan qancha xarajat qilinganini (UZS ekvivalentida) qaytaradi."""
    now = now_tz()
    start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    kurs = get_kurs(space_key)
    conn = get_db()
    rows = conn.execute(
        "SELECT account, currency, SUM(amount) FROM transactions "
        "WHERE space_key=? AND account IS NOT NULL AND amount<0 AND created_at>=? "
        "GROUP BY account, currency",
        (space_key, start.isoformat(timespec="seconds")),
    ).fetchall()
    conn.close()
    totals = {}
    for account, currency, total in rows:
        val_uzs = abs(total) * (kurs if currency == "USD" else 1)
        totals[account] = totals.get(account, 0.0) + val_uzs
    return totals


def build_accounts_summary_text(space_key: str) -> str:
    rows = get_accounts(space_key)
    today_expense = get_today_account_expense(space_key)
    today_str = now_tz().strftime("%d.%m.%Y")
    lines = [f"🏦 <b>Hisoblar bo'yicha saldo</b> <i>({today_str})</i>"]
    total_balance = 0.0
    total_today_expense = 0.0
    for name, balance in rows:
        total_balance += balance
        marker = "🔴 " if balance < 0 else "▫️ "
        exp_today = today_expense.get(name, 0.0)
        total_today_expense += exp_today
        if exp_today > 0:
            lines.append(f"{marker}<b>{name}</b>: {fmt_num(balance, 2)} so'm  <i>(bugun xarajat: {fmt_num(exp_today, 2)} so'm)</i>")
        else:
            lines.append(f"{marker}<b>{name}</b>: {fmt_num(balance, 2)} so'm")
    lines.append("—")
    lines.append(f"<b>Umumiy (barcha hisoblar): {fmt_num(total_balance, 2)} so'm</b>")
    if total_today_expense > 0:
        lines.append(f"<i>Bugungi umumiy xarajat: {fmt_num(total_today_expense, 2)} so'm</i>")
    return "\n".join(lines)


def build_accounts_inline_text(space_key: str) -> str:
    """Xarajat/daromad yozilgandan keyin darhol ko'rsatiladigan qisqa
    hisoblar qoldig'i — manfiy hisob 🔴 bilan belgilanadi."""
    rows = get_accounts(space_key)
    lines = ["🏦 <b>Hisoblar qoldig'i</b>"]
    for name, balance in rows:
        marker = "🔴 " if balance < 0 else "▫️ "
        lines.append(f"{marker}<b>{name}</b>: {fmt_num(balance, 2)} so'm")
    return "\n".join(lines)


def parse_saldo_input(text: str, space_key: str = None):
    text_stripped = text.strip()
    lower = text_stripped.lower()
    names = get_account_names(space_key) if space_key else ACCOUNT_NAMES
    for name in sorted(names, key=lambda n: -len(n)):
        if lower.startswith(name.lower()):
            rest = text_stripped[len(name):].strip()
            m = re.search(r"(\d[\d\s.,]*)\s*(mln|million|млн|ming|минг)?\b", rest, re.IGNORECASE)
            if m:
                value = parse_amount_token(m.group(1), m.group(2))
                if value is not None:
                    return name, value
    return None


def get_space_key(user_id: int) -> str:
    conn = get_db()
    row = conn.execute(
        "SELECT group_id FROM user_group WHERE user_id=?", (user_id,)
    ).fetchone()
    conn.close()
    if row:
        return f"g{row[0]}"
    return f"u{user_id}"


def get_space_key_for_update(update: Update) -> str:
    """Guruh chatida yozilgan xabarlar shu guruhning umumiy balansiga
    tushadi (kim yozishidan qat'iy nazar); shaxsiy chatda esa odatdagi
    shaxsiy/jamoaviy (kod orqali qo'shilgan) balans ishlatiladi."""
    chat = update.effective_chat
    if chat and chat.type in ("group", "supergroup"):
        space_key = f"c{chat.id}"
    else:
        space_key = get_space_key(update.effective_user.id)
    remember_chat(chat.id if chat else update.effective_user.id, space_key)
    if update.effective_user:
        remember_user(update.effective_user.id, get_user_display_name(update))
    return space_key


def get_user_display_name(update: Update) -> str:
    """Excel hisobotida kim qancha xarajat qilganini ko'rsatish uchun
    foydalanuvchining ismini oladi."""
    user = update.effective_user
    if not user:
        return "Noma'lum"
    return user.full_name or (f"@{user.username}" if user.username else str(user.id))


def remember_chat(chat_id: int, space_key: str):
    """Kun oxiridagi avtomatik xabar yuborish uchun qaysi chat qaysi
    balansga tegishli ekanini yodda saqlaydi."""
    conn = get_db()
    conn.execute(
        "INSERT INTO known_chats (chat_id, space_key, updated_at) VALUES (?, ?, ?) "
        "ON CONFLICT(chat_id) DO UPDATE SET space_key=excluded.space_key, updated_at=excluded.updated_at",
        (chat_id, space_key, now_tz().isoformat(timespec="seconds")),
    )
    conn.commit()
    conn.close()


def get_all_known_chats():
    conn = get_db()
    rows = conn.execute("SELECT chat_id, space_key FROM known_chats").fetchall()
    conn.close()
    return rows


def add_transaction(user_id: int, space_key: str, amount: float, currency: str, category: str, note: str, method: str = None, account: str = None, user_name: str = None):
    conn = get_db()
    conn.execute(
        "INSERT INTO transactions (user_id, space_key, amount, currency, category, note, method, account, user_name, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (user_id, space_key, amount, currency, category, note, method, account, user_name,
         now_tz().isoformat(timespec="seconds")),
    )
    conn.commit()
    conn.close()


def detect_method(note: str):
    note_lower = note.lower()
    if "karta" in note_lower or "kart" in note_lower or "plastik" in note_lower:
        return "karta"
    if "naqt" in note_lower or "naqd" in note_lower or "nakt" in note_lower:
        return "naqt"
    return None


def get_method_totals(space_key: str, method: str):
    conn = get_db()
    rows = conn.execute(
        "SELECT currency, SUM(amount) FROM transactions "
        "WHERE space_key=? AND method=? AND amount<0 GROUP BY currency",
        (space_key, method),
    ).fetchall()
    recent = conn.execute(
        "SELECT amount, currency, category, note, created_at FROM transactions "
        "WHERE space_key=? AND method=? ORDER BY id DESC LIMIT 10",
        (space_key, method),
    ).fetchall()
    conn.close()
    totals = {"UZS": 0.0, "USD": 0.0}
    for currency, total in rows:
        totals[currency] = abs(total or 0.0)
    return totals, recent


def get_balance(space_key: str):
    conn = get_db()
    cur = conn.execute(
        "SELECT currency, SUM(amount) FROM transactions WHERE space_key=? GROUP BY currency",
        (space_key,),
    )
    rows = cur.fetchall()
    conn.close()
    balances = {"UZS": 0.0, "USD": 0.0}
    for currency, total in rows:
        balances[currency] = total or 0.0
    return balances


def get_kurs(space_key: str) -> float:
    conn = get_db()
    row = conn.execute(
        "SELECT value FROM settings WHERE space_key=? AND key='kurs'", (space_key,)
    ).fetchone()
    conn.close()
    return float(row[0]) if row else DEFAULT_KURS


def set_kurs(space_key: str, value: float):
    conn = get_db()
    conn.execute(
        "INSERT INTO settings (space_key, key, value) VALUES (?, 'kurs', ?) "
        "ON CONFLICT(space_key, key) DO UPDATE SET value=excluded.value",
        (space_key, str(value)),
    )
    conn.commit()
    conn.close()


def get_method_accounts(space_key: str, method: str):
    """\"💳 Karta xarajati\" / \"💵 Naqt xarajat\" tugmasiga bir marta ulangan
    hisoblar ro'yxatini qaytaradi. Hali ulanmagan bo'lsa None qaytaradi —
    shunda tugma bosilganda hisoblarni tanlash oynasi ko'rsatiladi."""
    conn = get_db()
    row = conn.execute(
        "SELECT value FROM settings WHERE space_key=? AND key=?",
        (space_key, f"method_accounts_{method}"),
    ).fetchone()
    conn.close()
    if not row or not row[0]:
        return None
    existing = set(get_account_names(space_key))
    names = [n for n in row[0].split("|") if n and n in existing]
    return names


def set_method_accounts(space_key: str, method: str, names: list):
    conn = get_db()
    value = "|".join(names)
    conn.execute(
        "INSERT INTO settings (space_key, key, value) VALUES (?, ?, ?) "
        "ON CONFLICT(space_key, key) DO UPDATE SET value=excluded.value",
        (space_key, f"method_accounts_{method}", value),
    )
    conn.commit()
    conn.close()


def remove_account_from_method_links(space_key: str, name: str):
    """Bir hisob o'chirilganda, uni Karta/Naqt ulanishlaridan ham olib tashlaydi."""
    for method in ("karta", "naqt"):
        linked = get_method_accounts(space_key, method)
        if linked and name in linked:
            linked = [n for n in linked if n != name]
            set_method_accounts(space_key, method, linked)


METHOD_PERIODS = [
    ("today", "📅 Kunlik"),
    ("week", "📆 Haftalik"),
    ("month", "🗓 Oylik"),
    ("all", "♾ Umumiy"),
]


def get_method_period_start(period: str):
    """Berilgan davr kodi uchun (bugun/hafta/oy/umumiy) boshlanish vaqti va
    ko'rinadigan nomini qaytaradi. \"all\" uchun boshlanish vaqti None
    (cheklovsiz, butun tarix)."""
    now = now_tz()
    if period == "today":
        return now.replace(hour=0, minute=0, second=0, microsecond=0), "Bugun"
    if period == "week":
        return now - timedelta(days=7), "So'nggi 7 kun"
    if period == "month":
        return now.replace(day=1, hour=0, minute=0, second=0, microsecond=0), "Shu oy"
    return None, "Umumiy (barcha vaqt)"


def get_account_expense_totals(space_key: str, names: list, start_dt=None):
    """Berilgan hisoblar ro'yxati bo'yicha (va ixtiyoriy boshlanish vaqtidan
    buyon) HAR BIR hisobning alohida xarajat jamini qaytaradi:
    {hisob_nomi: {"UZS": summa, "USD": summa}}"""
    result = {n: {"UZS": 0.0, "USD": 0.0} for n in names}
    if not names:
        return result
    conn = get_db()
    placeholders = ",".join("?" for _ in names)
    query = (
        f"SELECT account, currency, SUM(-amount) FROM transactions "
        f"WHERE space_key=? AND amount<0 AND account IN ({placeholders})"
    )
    params = [space_key] + names
    if start_dt is not None:
        query += " AND created_at>=?"
        params.append(start_dt.isoformat(timespec="seconds"))
    query += " GROUP BY account, currency"
    rows = conn.execute(query, params).fetchall()
    conn.close()
    for account, currency, total in rows:
        if account in result:
            result[account][currency] = total or 0.0
    return result


def build_method_expense_text(space_key: str, method: str, title: str, period: str) -> str:
    names = get_method_accounts(space_key, method) or []
    start_dt, period_label = get_method_period_start(period)
    totals = get_account_expense_totals(space_key, names, start_dt)

    lines = [f"<b>{title}</b>", f"<i>Davr: {period_label}</i>", ""]
    grand_uzs = 0.0
    grand_usd = 0.0
    for n in names:
        t = totals.get(n, {"UZS": 0.0, "USD": 0.0})
        grand_uzs += t["UZS"]
        grand_usd += t["USD"]
        line = f"▫️ <b>{n}</b>: {fmt_num(t['UZS'], 2)} so'm"
        if t["USD"] > 0:
            line += f", {fmt_num(t['USD'], 2)} $"
        lines.append(line)
    lines.append("—")
    total_line = f"<b>Jami xarajat: {fmt_num(grand_uzs, 2)} so'm</b>"
    if grand_usd > 0:
        total_line += f", <b>{fmt_num(grand_usd, 2)} $</b>"
    lines.append(total_line)
    return "\n".join(lines)


def build_method_period_keyboard(method: str, selected_period: str) -> InlineKeyboardMarkup:
    rows = []
    row = []
    for key, label in METHOD_PERIODS:
        text = f"✅ {label}" if key == selected_period else label
        row.append(InlineKeyboardButton(text, callback_data=f"methodperiod:{method}:{key}"))
        if len(row) == 2:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton("🔄 Hisoblarni qayta ulash", callback_data=f"methodlink:{method}:relink")])
    return InlineKeyboardMarkup(rows)


def build_method_link_keyboard(method: str, names: list, selected: set) -> InlineKeyboardMarkup:
    buttons = []
    for n in names:
        mark = "✅ " if n in selected else "▫️ "
        buttons.append([InlineKeyboardButton(f"{mark}{n}", callback_data=f"methodlink:{method}:toggle:{n}")])
    buttons.append([InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"methodlink:{method}:confirm")])
    buttons.append([InlineKeyboardButton("❌ Bekor qilish", callback_data=f"methodlink:{method}:cancel")])
    return InlineKeyboardMarkup(buttons)


def get_history(space_key: str, limit: int = 20):
    conn = get_db()
    cur = conn.execute(
        "SELECT amount, currency, category, note, created_at FROM transactions "
        "WHERE space_key=? ORDER BY id DESC LIMIT ?",
        (space_key, limit),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def get_period_stats(space_key: str, start_dt: datetime):
    conn = get_db()
    cur = conn.execute(
        "SELECT amount, currency, category FROM transactions "
        "WHERE space_key=? AND created_at>=?",
        (space_key, start_dt.isoformat(timespec="seconds")),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def get_period_stats_with_date(space_key: str, start_dt: datetime):
    conn = get_db()
    cur = conn.execute(
        "SELECT amount, currency, created_at FROM transactions "
        "WHERE space_key=? AND created_at>=?",
        (space_key, start_dt.isoformat(timespec="seconds")),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def get_combined_history(space_key: str, limit: int = 20):
    conn = get_db()
    tx_rows = conn.execute(
        "SELECT amount, currency, category, note, created_at FROM transactions "
        "WHERE space_key=? ORDER BY id DESC LIMIT ?",
        (space_key, limit),
    ).fetchall()
    debt_rows = conn.execute(
        "SELECT person, direction, amount, currency, note, created_at FROM debts "
        "WHERE space_key=? ORDER BY id DESC LIMIT ?",
        (space_key, limit),
    ).fetchall()
    conn.close()

    combined = []
    for amount, currency, category, note, created_at in tx_rows:
        combined.append({
            "kind": "tx", "created_at": created_at, "amount": amount,
            "currency": currency, "category": category, "note": note,
        })
    for person, direction, amount, currency, note, created_at in debt_rows:
        combined.append({
            "kind": "debt", "created_at": created_at, "person": person,
            "direction": direction, "amount": amount, "currency": currency, "note": note,
        })
    combined.sort(key=lambda x: x["created_at"], reverse=True)
    return combined[:limit]


def get_category_totals(space_key: str):
    conn = get_db()
    cur = conn.execute(
        "SELECT category, currency, SUM(amount) FROM transactions "
        "WHERE space_key=? AND amount<0 GROUP BY category, currency ORDER BY SUM(amount) ASC",
        (space_key,),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def get_category_totals_period(space_key: str, start_dt, end_dt=None):
    conn = get_db()
    if start_dt is None:
        rows = conn.execute(
            "SELECT category, currency, SUM(amount) FROM transactions "
            "WHERE space_key=? AND amount<0 GROUP BY category, currency ORDER BY SUM(amount) ASC",
            (space_key,),
        ).fetchall()
    elif end_dt is None:
        rows = conn.execute(
            "SELECT category, currency, SUM(amount) FROM transactions "
            "WHERE space_key=? AND amount<0 AND created_at>=? GROUP BY category, currency ORDER BY SUM(amount) ASC",
            (space_key, start_dt.isoformat(timespec="seconds")),
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT category, currency, SUM(amount) FROM transactions "
            "WHERE space_key=? AND amount<0 AND created_at>=? AND created_at<=? "
            "GROUP BY category, currency ORDER BY SUM(amount) ASC",
            (space_key, start_dt.isoformat(timespec="seconds"), end_dt.isoformat(timespec="seconds")),
        ).fetchall()
    conn.close()
    return rows


def reset_space(space_key: str):
    """Faqat \"🧹 Tozalash\" tugmasi bosilganda chaqiriladi. Oddiy kod
    yangilanishi/qayta ishga tushganda bu funksiya CHAQIRILMAYDI — shuning
    uchun eski yozuvlar (tranzaksiyalar) baza fayli saqlanib qolar ekan,
    hech qachon o'z-o'zidan uchib ketmaydi. Faqat shu tugma bosilgandagina
    tarix o'chiriladi VA hisoblar balansi 0'ga tushiriladi."""
    conn = get_db()
    conn.execute("DELETE FROM transactions WHERE space_key=?", (space_key,))
    conn.execute("DELETE FROM debts WHERE space_key=?", (space_key,))
    conn.commit()
    conn.close()
    # Hisoblar ro'yxati (nomlari) saqlanib qoladi, faqat balanslar 0'ga tushadi
    for name in get_account_names(space_key):
        set_account_balance(space_key, name, 0.0)


def create_group(user_id: int) -> str:
    code = "".join(random.choices(string.ascii_uppercase + string.digits, k=6))
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO groups (code, created_by, created_at) VALUES (?, ?, ?)",
        (code, user_id, now_tz().isoformat(timespec="seconds")),
    )
    group_id = cur.lastrowid
    conn.execute(
        "INSERT INTO user_group (user_id, group_id) VALUES (?, ?) "
        "ON CONFLICT(user_id) DO UPDATE SET group_id=excluded.group_id",
        (user_id, group_id),
    )
    conn.commit()
    conn.close()
    return code


def join_group(user_id: int, code: str) -> bool:
    conn = get_db()
    row = conn.execute("SELECT id FROM groups WHERE code=?", (code.upper(),)).fetchone()
    if not row:
        conn.close()
        return False
    group_id = row[0]
    conn.execute(
        "INSERT INTO user_group (user_id, group_id) VALUES (?, ?) "
        "ON CONFLICT(user_id) DO UPDATE SET group_id=excluded.group_id",
        (user_id, group_id),
    )
    conn.commit()
    conn.close()
    return True


def leave_group(user_id: int):
    conn = get_db()
    conn.execute("DELETE FROM user_group WHERE user_id=?", (user_id,))
    conn.commit()
    conn.close()


def get_group_info(user_id: int):
    conn = get_db()
    row = conn.execute(
        "SELECT g.code, g.id, (SELECT COUNT(*) FROM user_group WHERE group_id=g.id) "
        "FROM groups g JOIN user_group ug ON ug.group_id=g.id WHERE ug.user_id=?",
        (user_id,),
    ).fetchone()
    conn.close()
    return row


# ---------- Qarz (debitor/kreditor) ----------
# direction='oldim'  -> foydalanuvchi shu odamdan qarz oldi (u qarzdor)
# direction='berdim' -> foydalanuvchi shu odamga qarz berdi (u sizga qarzdor)

def add_debt(space_key: str, person: str, direction: str, amount: float, currency: str, note: str):
    conn = get_db()
    conn.execute(
        "INSERT INTO debts (space_key, person, direction, amount, currency, note, created_at, settled) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, 0)",
        (space_key, person, direction, amount, currency, note,
         now_tz().isoformat(timespec="seconds")),
    )
    conn.commit()
    conn.close()


def get_debt_summary(space_key: str):
    conn = get_db()
    rows = conn.execute(
        "SELECT person, direction, currency, SUM(amount) FROM debts "
        "WHERE space_key=? AND settled=0 GROUP BY person, direction, currency",
        (space_key,),
    ).fetchall()
    conn.close()
    summary = {}
    for person, direction, currency, total in rows:
        summary.setdefault(person, {}).setdefault(currency, 0.0)
        if direction == "berdim":
            summary[person][currency] += total
        else:
            summary[person][currency] -= total
    return summary


def build_debts_summary_text(space_key: str) -> str:
    summary = get_debt_summary(space_key)
    lines = ["🤝 <b>Qarzlar holati</b>"]
    has_open = False
    for person, cur_map in summary.items():
        parts = []
        for cur, net in cur_map.items():
            if abs(net) < 0.01:
                continue
            unit = "so'm" if cur == "UZS" else "$"
            if net > 0:
                parts.append(f"sizga {fmt_num(net)} {unit} qarzdor")
            else:
                parts.append(f"siz {fmt_num(abs(net))} {unit} qarzdorsiz")
        if parts:
            has_open = True
            lines.append(f"▫️ <b>{esc(person)}</b>: " + ", ".join(parts))
    if not has_open:
        return "Qarzlar yo'q. 🎉"
    lines.append("\n<i>Yopish uchun: /qarz_yopish Ism</i>")
    return "\n".join(lines)


def settle_person(space_key: str, person: str) -> int:
    conn = get_db()
    cur = conn.execute(
        "UPDATE debts SET settled=1 WHERE space_key=? AND LOWER(person)=LOWER(?) AND settled=0",
        (space_key, person),
    )
    changed = cur.rowcount
    conn.commit()
    conn.close()
    return changed


# ---------- Summani tahlil qilish (13 mln, 13 000 000, mingchisiz h.k.) ----------

def parse_amount_token(raw_amount: str, mult_word: str):
    s = raw_amount.strip()
    if not s:
        return None

    # Agar oxirida .XX yoki ,XX (1-2 xonali kasr/tiyin) bo'lsa, buni KASR
    # QISM deb hisoblaymiz, ming ajratuvchi emas — masalan "22 830.49"
    decimal_match = re.search(r"[.,](\d{1,2})$", s)
    if decimal_match:
        frac = decimal_match.group(1)
        integer_part = s[: decimal_match.start()]
        integer_digits = re.sub(r"[\s.,]", "", integer_part) or "0"
        try:
            value = float(f"{integer_digits}.{frac}")
        except ValueError:
            return None
    else:
        digits = re.sub(r"[\s.,]", "", s)
        if not digits:
            return None
        try:
            value = float(digits)
        except ValueError:
            return None

    if mult_word:
        mw = mult_word.lower()
        if mw in ("mln", "million", "млн"):
            value *= 1_000_000
        elif mw in ("ming", "минг"):
            value *= 1_000
    return value


def parse_plain_amount(text: str):
    """Faqat raqamdan iborat matnni (ishorasiz, +/- belgilarsiz) miqdorga
    aylantiradi — hisobdan hisobga o'tkazma miqdorini o'qish uchun.
    Mos kelmasa None qaytaradi."""
    stripped = text.strip()
    match = re.fullmatch(r"(\d[\d\s.,]*)\s*(mln|million|млн|ming|минг)?", stripped, re.IGNORECASE)
    if not match:
        return None
    return parse_amount_token(match.group(1), match.group(2))


PATTERN = re.compile(
    r"^([+-]?)\s*(\d[\d\s.,]*)\s*(mln|million|млн|ming|минг)?\b\s*(\$|usd|so'?m|sum|uzs)?\s*(.*)$",
    re.IGNORECASE,
)


def parse_message(text: str):
    text = text.strip()
    match = PATTERN.match(text)
    if not match:
        return None

    sign, raw_amount, mult_word, currency_raw, note = match.groups()
    amount = parse_amount_token(raw_amount, mult_word)
    if amount is None:
        return None

    currency = "UZS" if (not currency_raw or currency_raw.lower() in ("sum", "so'm", "som", "uzs")) else "USD"
    note = note.strip() or "(izohsiz)"

    if sign == "+":
        is_income = True
    elif sign == "-":
        is_income = False
    else:
        is_income = is_income_text(note)

    amount = abs(amount) if is_income else -abs(amount)
    category = detect_category(note, is_income)
    return amount, currency, category, note


def try_parse_bulk_lines(text: str):
    """Xabarda 2 yoki undan ko'p qator bo'lsa va HAR BIRI mustaqil
    xarajat/daromad sifatida tushunilsa, ro'yxat qilib qaytaradi.
    Aks holda None (bitta oddiy xabar sifatida qayta ishlanadi)."""
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if len(lines) < 2:
        return None
    parsed_list = []
    for line in lines:
        parsed = parse_message(line)
        if not parsed:
            return None
        parsed_list.append(parsed)
    return parsed_list


def format_balance_text(space_key: str) -> str:
    bal = get_balance(space_key)
    kurs = get_kurs(space_key)
    total_uzs = bal["UZS"] + bal["USD"] * kurs
    return (
        f"💰 <b>Balans</b>\n"
        f"UZS: <b>{fmt_num(bal['UZS'])}</b> so'm\n"
        f"USD: <b>{fmt_num(bal['USD'], 2)}</b> $\n"
        f"—\n"
        f"<i>Umumiy (taxminan): {fmt_num(total_uzs)} so'm  (kurs: 1$ = {fmt_num(kurs)} so'm)</i>"
    )


def build_full_status_text(space_key: str) -> str:
    kurs = get_kurs(space_key)
    bal = get_balance(space_key)
    balance_uzs = bal["UZS"] + bal["USD"] * kurs

    debt_summary = get_debt_summary(space_key)
    receivable_uzs = 0.0
    payable_uzs = 0.0
    for person, cur_map in debt_summary.items():
        for cur, net in cur_map.items():
            val_uzs = net * (kurs if cur == "USD" else 1)
            if val_uzs > 0:
                receivable_uzs += val_uzs
            else:
                payable_uzs += -val_uzs

    net_worth = balance_uzs + receivable_uzs - payable_uzs

    lines = [format_balance_text(space_key), ""]
    lines.append(build_debts_summary_text(space_key))
    lines.append("")
    lines.append(
        f"📌 Sof holat (balans + sizga qarzdorlar − siz qarzdorsiz):\n"
        f"{fmt_num(net_worth)} so'm (taxminan)"
    )
    lines.append("")
    lines.append(build_accounts_summary_text(space_key))
    return "\n".join(lines)


# ---------- Tabiiy tildagi qarz yozuvini aniqlash (tugmasiz) ----------
# Masalan: "500000 Alidan qarz oldim" yoki "Valiga 200000 qarz berdim"

FREEFORM_AMOUNT_RE = re.compile(r"(\d[\d\s.,]*\d|\d)\s*(mln|million|млн|ming|минг)?\b", re.IGNORECASE)
PERSON_FROM_RE = re.compile(r"\b([A-Za-zА-Яа-яЎўҚқҒғҲҳ']+?)dan\b", re.IGNORECASE)
PERSON_TO_RE = re.compile(r"\b([A-Za-zА-Яа-яЎўҚқҒғҲҳ']+?)ga\b", re.IGNORECASE)


def try_parse_freeform_debt(text: str):
    if "qarz" not in text.lower():
        return None
    text_lower = text.lower()
    if re.search(r"qarz\s*old", text_lower):
        direction = "oldim"
    elif re.search(r"qarz\s*ber", text_lower):
        direction = "berdim"
    else:
        return None

    m = FREEFORM_AMOUNT_RE.search(text)
    if not m:
        return None
    amount = parse_amount_token(m.group(1), m.group(2))
    if not amount:
        return None

    currency = "USD" if re.search(r"\$|\busd\b", text_lower) else "UZS"

    person = None
    pm = PERSON_FROM_RE.search(text) if direction == "oldim" else PERSON_TO_RE.search(text)
    if pm:
        candidate = pm.group(1)
        if candidate.lower() not in ("qarz", "shu", "u", "meni", "sen"):
            person = candidate

    note = text.strip()
    return person, direction, amount, currency, note


# ---------- Skrinshotdan summani o'qish (OCR) ----------

def extract_amount_candidates(text: str):
    candidates = []

    for m in re.finditer(r"\$\s?(\d{1,5}(?:[.,]\d{1,2})?)", text):
        candidates.append((float(m.group(1).replace(",", ".")), "USD"))
    for m in re.finditer(r"(\d{1,5}(?:[.,]\d{1,2})?)\s?\$", text):
        candidates.append((float(m.group(1).replace(",", ".")), "USD"))

    for m in re.finditer(r"\d{1,3}(?:[ ,.]\d{3}){1,4}", text):
        digits = re.sub(r"[ ,.]", "", m.group(0))
        if digits.isdigit():
            candidates.append((float(digits), "UZS"))
    for m in re.finditer(r"(?<!\d)(\d{4,9})(?!\d)", text):
        candidates.append((float(m.group(1)), "UZS"))

    seen = set()
    unique = []
    for amt, cur in candidates:
        key = (round(amt, 2), cur)
        if key not in seen:
            seen.add(key)
            unique.append((amt, cur))

    unique.sort(key=lambda x: -x[0])
    return unique[:4]


# ---------- Bot buyruqlari ----------

START_TEXT = "👋 <b>Salom! Men sizning hisobchingizman.</b> 🧮"

HELP_TEXT = (
    "<b>Men xarajat/balans botiman.</b>\n\n"
    "Tugma bosmasdan ham oddiy gapda yozishingiz mumkin:\n"
    "🔹 <code>-50000 taksi</code>\n"
    "🔹 <code>13 mln maosh oldim</code>\n"
    "🔹 <code>500000 Alidan qarz oldim</code>\n"
    "🔹 <code>200 000 Valiga qarz berdim</code>\n"
    "🔹 <code>50000 taksi karta</code>   <i>(yoki \"naqt\" — to'lov usulini ham belgilaydi)</i>\n\n"
    "🎙 Ovozli xabar yuborsangiz ham tushunaman.\n"
    "📸 To'lov skrinshotini izoh bilan yuborsangiz, summani o'zim o'qiyman.\n\n"
    "👨‍👩‍👧 Meni guruhga qo'shsangiz, guruhdagi barcha xarajatlar bitta umumiy "
    "balansga yoziladi <i>(oilaviy hisob uchun qulay)</i>.\n"
    "🌙 Har kuni soat 23:59 da kunlik balans holatini o'zim yuboraman.\n\n"
    "<i>Pastdagi tugmalar orqali tez foydalaning.</i>"
)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(START_TEXT, reply_markup=MAIN_KEYBOARD)
    buttons = [
        [InlineKeyboardButton("✅ Ha, o'chirish", callback_data="reset:confirm")],
        [InlineKeyboardButton("❌ Yo'q, saqlab qolish", callback_data="reset:cancel")],
    ]
    await update.message.reply_text(
        "🗑 Eski xarajatlaringizni o'chirib, yangidan boshlaymizmi?",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(HELP_TEXT, reply_markup=MAIN_KEYBOARD)


async def balance(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    await update.message.reply_text(build_balance_view_text(space_key), reply_markup=MAIN_KEYBOARD)


async def history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    entries = get_combined_history(space_key)
    if not entries:
        await update.message.reply_text("Hali yozuvlar yo'q.")
        return
    lines = ["📜 <b>Oxirgi yozuvlar</b> <i>(xarajat/daromad va qarzlar)</i>"]
    for e in entries:
        try:
            date_str = datetime.fromisoformat(e["created_at"]).strftime("%d.%m.%Y")
        except ValueError:
            date_str = e["created_at"].split("T")[0]
        if e["kind"] == "tx":
            sign = "+" if e["amount"] >= 0 else ""
            unit = "so'm" if e["currency"] == "UZS" else "$"
            lines.append(f"🕐 {date_str}  <b>{sign}{fmt_num(e['amount'])} {unit}</b>  <i>[{esc(e['category'])}]</i>  — {esc(e['note'])}")
        else:
            unit = "so'm" if e["currency"] == "UZS" else "$"
            if e["direction"] == "oldim":
                verb = f"{esc(e['person'])} dan {fmt_num(e['amount'])} {unit} qarz oldingiz"
            else:
                verb = f"{esc(e['person'])} ga {fmt_num(e['amount'])} {unit} qarz berdingiz"
            lines.append(f"🤝 {date_str}  {verb}  — {esc(e['note'])}")
    await update.message.reply_text("\n".join(lines))


def build_stat_chart(space_key: str) -> str:
    kurs = get_kurs(space_key)
    days = 7
    now = now_tz()
    start = (now - timedelta(days=days - 1)).replace(hour=0, minute=0, second=0, microsecond=0)

    rows = get_period_stats_with_date(space_key, start)
    date_list = [(start + timedelta(days=i)).date() for i in range(days)]
    daily_income = {d: 0.0 for d in date_list}
    daily_expense = {d: 0.0 for d in date_list}

    for amount, currency, created_at in rows:
        d = datetime.fromisoformat(created_at).date()
        if d not in daily_income:
            continue
        val_uzs = amount * (kurs if currency == "USD" else 1)
        if val_uzs >= 0:
            daily_income[d] += val_uzs
        else:
            daily_expense[d] += -val_uzs

    labels = [d.strftime("%d.%m") for d in date_list]
    income_vals = [daily_income[d] for d in date_list]
    expense_vals = [daily_expense[d] for d in date_list]

    debt_summary = get_debt_summary(space_key)
    receivable = 0.0
    payable = 0.0
    for person, cur_map in debt_summary.items():
        for cur, net in cur_map.items():
            val = net * (kurs if cur == "USD" else 1)
            if val > 0:
                receivable += val
            else:
                payable += -val

    cat_rows = get_category_totals(space_key)
    cat_totals_uzs = {}
    for category, currency, total in cat_rows:
        val = abs(total) * (kurs if currency == "USD" else 1)
        cat_totals_uzs[category] = cat_totals_uzs.get(category, 0) + val
    top_cats = sorted(cat_totals_uzs.items(), key=lambda x: -x[1])[:5]

    fig, axes = plt.subplots(1, 3, figsize=(15, 4.8))
    ax1, ax2, ax3 = axes

    x = list(range(len(labels)))
    width = 0.35
    bars_income = ax1.bar([i - width / 2 for i in x], income_vals, width, label="Daromad", color="#2e7d32")
    bars_expense = ax1.bar([i + width / 2 for i in x], expense_vals, width, label="Xarajat", color="#c62828")
    ax1.set_xticks(x)
    ax1.set_xticklabels(labels, rotation=45)
    ax1.set_title("So'nggi 7 kun (so'mda)")
    ax1.legend()
    ax1.grid(axis="y", alpha=0.3)
    ax1.bar_label(bars_income, fmt="%.0f", fontsize=7, rotation=90, padding=2)
    ax1.bar_label(bars_expense, fmt="%.0f", fontsize=7, rotation=90, padding=2)

    bars_debt = ax2.bar(["Sizga\nqarzdor", "Siz\nqarzdor"], [receivable, payable],
                         color=["#2e7d32", "#c62828"])
    ax2.set_title("Qarz holati (so'mda)")
    ax2.grid(axis="y", alpha=0.3)
    ax2.bar_label(bars_debt, fmt="%.0f", fontsize=9, padding=3)

    if top_cats:
        cat_labels = [c for c, v in top_cats]
        cat_values = [v for c, v in top_cats]
        colors = ["#c62828", "#ef6c00", "#f9a825", "#6a1b9a", "#0277bd"][:len(cat_labels)]
        wedges, texts, autotexts = ax3.pie(
            cat_values, labels=cat_labels, autopct=lambda p: f"{p:.0f}%\n({fmt_num(p/100*sum(cat_values))})",
            colors=colors, textprops={"fontsize": 7},
        )
        ax3.set_title("Xarajat kategoriyalari")
    else:
        ax3.axis("off")
        ax3.text(0.5, 0.5, "Xarajat yo'q", ha="center", va="center")

    fig.suptitle("Moliyaviy holat")
    fig.tight_layout()

    out_dir = tempfile.gettempdir()
    filepath = os.path.join(out_dir, f"stat_{space_key}_{now_tz().strftime('%Y%m%d_%H%M%S')}.png")
    fig.savefig(filepath, dpi=120)
    plt.close(fig)
    return filepath


async def stat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    now = now_tz()

    periods = [
        ("Bugun", now.replace(hour=0, minute=0, second=0, microsecond=0)),
        ("Shu hafta", (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)),
        ("Shu oy", now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)),
    ]

    lines = []
    for label, start_dt in periods:
        rows = get_period_stats(space_key, start_dt)
        income = {"UZS": 0.0, "USD": 0.0}
        expense = {"UZS": 0.0, "USD": 0.0}
        cat_totals = {}
        for amount, currency, category in rows:
            if amount >= 0:
                income[currency] += amount
            else:
                expense[currency] += -amount
                if currency == "UZS":
                    cat_totals[category] = cat_totals.get(category, 0) + (-amount)

        lines.append(f"📅 {label}:")
        lines.append(f"  Daromad: {fmt_num(income['UZS'])} so'm, {fmt_num(income['USD'], 2)} $")
        lines.append(f"  Xarajat: {fmt_num(expense['UZS'])} so'm, {fmt_num(expense['USD'], 2)} $")
        if cat_totals:
            top = sorted(cat_totals.items(), key=lambda x: -x[1])[:3]
            top_str = ", ".join(f"{c}: {fmt_num(v)}" for c, v in top if v > 0)
            if top_str:
                lines.append(f"  Ko'p sarflangan: {top_str}")
        lines.append("")

    await update.message.reply_text("\n".join(lines))

    if MATPLOTLIB_AVAILABLE:
        try:
            path = build_stat_chart(space_key)
            with open(path, "rb") as f:
                await update.message.reply_photo(
                    photo=f, caption="📊 Daromad/xarajat, qarz holati va kategoriyalar diagrammasi"
                )
            os.remove(path)
        except Exception:
            pass
    else:
        await update.message.reply_text(
            "📊 Diagramma ko'rish uchun: python -m pip install matplotlib --break-system-packages"
        )


async def categories(update: Update, context: ContextTypes.DEFAULT_TYPE):
    buttons = [
        [InlineKeyboardButton("Bugun", callback_data="catperiod:day"),
         InlineKeyboardButton("Shu hafta", callback_data="catperiod:week")],
        [InlineKeyboardButton("Shu oy", callback_data="catperiod:month"),
         InlineKeyboardButton("Hammasi", callback_data="catperiod:all")],
    ]
    await update.message.reply_text(
        "🗂 <b>Qaysi davr uchun ko'rsatilsin?</b>",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


def build_category_period_text(space_key: str, period: str) -> str:
    now = now_tz()
    if period == "day":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        label = "Bugun"
    elif period == "week":
        start = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        label = "Shu hafta"
    elif period == "month":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        label = "Shu oy"
    else:
        start = None
        label = "Hammasi (jami)"

    rows = get_category_totals_period(space_key, start)
    if not rows:
        return f"🗂 <b>{label}</b>\n\nBu davrda xarajat yo'q."

    lines = [f"🗂 <b>Kategoriyalar bo'yicha xarajat</b> <i>({label})</i>"]
    for category, currency, total in rows:
        unit = "so'm" if currency == "UZS" else "$"
        lines.append(f"▫️ <b>{esc(category)}</b>: {fmt_num(abs(total))} {unit}")
    return "\n".join(lines)


async def category_period_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    period = query.data.split(":")[1]
    space_key = get_space_key_for_update(update)
    await query.edit_message_text(build_category_period_text(space_key, period))


async def show_method_totals(update: Update, context: ContextTypes.DEFAULT_TYPE, method: str, title: str):
    space_key = get_space_key_for_update(update)
    totals, recent = get_method_totals(space_key, method)
    lines = [f"{title}:"]
    lines.append(f"  UZS: {fmt_num(totals['UZS'])} so'm")
    lines.append(f"  USD: {fmt_num(totals['USD'], 2)} $")
    if recent:
        lines.append("")
        lines.append("So'nggi yozuvlar:")
        for amount, currency, category, note, created_at in recent:
            try:
                date_str = datetime.fromisoformat(created_at).strftime("%d.%m.%Y")
            except ValueError:
                date_str = created_at.split("T")[0]
            unit = "so'm" if currency == "UZS" else "$"
            lines.append(f"  {date_str}  {fmt_num(abs(amount))} {unit}  [{category}] — {note}")
    else:
        lines.append("")
        lines.append("Hali bu usulda yozuv yo'q. Xarajat yozganingizda \"karta\" yoki \"naqt\" so'zini qo'shsangiz, shu yerda ko'rinadi.")
    await update.message.reply_text("\n".join(lines))


PENDING_TEXT_FLAGS = [
    "pending_kurs_input",
    "pending_saldo_input",
    "pending_new_account",
    "pending_new_account_name",
    "pending_rename_account",
    "pending_rename_new_name",
]


def clear_conflicting_pending_flags(context: ContextTypes.DEFAULT_TYPE):
    """Yangi matn kutiladigan jarayon (saldo, kurs, hisob qo'shish/nomini
    o'zgartirish) boshlanganda, oldingi tugallanmagan jarayonlar
    (masalan eski \"hisob saldosi\" so'rovi) yangisiga xalaqit bermasligi
    uchun barcha ziddiyatli pending bayroqlarni tozalaydi."""
    for key in PENDING_TEXT_FLAGS:
        context.chat_data.pop(key, None)


async def maybe_handle_transaction_during_pending(update: Update, context: ContextTypes.DEFAULT_TYPE,
                                                    user_id: int, space_key: str, text: str,
                                                    pending_key: str, label: str) -> bool:
    """Foydalanuvchi biror jarayon (masalan yangi hisob nomini kiritish yoki
    hisobni qayta nomlash) kutilayotgan paytda, kutilgan matn o'rniga
    xarajat/daromad yozuvi (masalan \"-50000 taksi\") kiritsa — bu yozuvni
    alohida xarajat sifatida qayd etadi, so'ng foydalanuvchidan avvalgi
    tugallanmagan jarayonni davom ettirish yoki bekor qilishni so'raydi.
    True qaytarsa — chaqiruvchi darhol return qilishi kerak (jarayon hali
    tugallanmagan, chat_data'dagi bayroq o'chirilmagan)."""
    parsed = parse_message(text)
    if not parsed:
        return False
    amount, currency, category, note = parsed
    method = detect_method(note)
    await commit_expense(update, context, user_id, space_key, amount, currency, category, note, method)
    buttons = [
        [InlineKeyboardButton("✅ Davom ettiraman", callback_data=f"pendingcont:{pending_key}:keep")],
        [InlineKeyboardButton("❌ Bekor qilish", callback_data=f"pendingcont:{pending_key}:cancel")],
    ]
    await update.message.reply_text(
        f"☝️ Bu yozuv xarajat/daromad sifatida qayd etildi.\n\n"
        f"Sizda hali tugallanmagan amal bor edi: <b>{esc(label)}</b>. Uni davom ettirasizmi yoki bekor qilasizmi?",
        reply_markup=InlineKeyboardMarkup(buttons),
    )
    return True


async def pending_continuation_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    parts = query.data.split(":")  # pendingcont:<key>:<action>
    key = parts[1]
    action = parts[2]
    if action == "cancel":
        context.chat_data.pop(key, None)
        if key == "pending_new_account":
            context.chat_data.pop("pending_new_account_name", None)
        if key == "pending_rename_account":
            context.chat_data.pop("pending_rename_new_name", None)
        await query.edit_message_text("Bekor qilindi. ✅")
    else:
        await query.edit_message_text("✅ Davom eting — kutilgan ma'lumotni yozing.")
    await send_main_menu(query.message)


async def send_main_menu(message_obj):
    """ESLATMA: foydalanuvchi so'roviga ko'ra bu funksiya endi alohida
    \"🏠 Asosiy menyu\" xabari yubormaydi — pastdagi doimiy klaviatura
    (ReplyKeyboardMarkup) allaqachon ekranda turadi, shuning uchun har bir
    amaldan keyin qo'shimcha xabar keraksiz. Funksiya chaqiruvlari kodda
    qolgani uchun bo'sh (no-op) qilib qo'yildi — kelajakda kerak bo'lsa,
    shu yerga qaytarib qo'yish kifoya."""
    return


async def start_method_link(update: Update, context: ContextTypes.DEFAULT_TYPE, method: str, title: str):
    space_key = get_space_key_for_update(update)
    context.chat_data[f"pending_method_link_{method}"] = set()
    names = get_account_names(space_key)
    await update.message.reply_text(
        f"<b>{esc(title)}</b>\n\nBu tugmaga qaysi hisob(lar)ni ulaymiz? Bir yoki bir nechta "
        f"hisobni tanlang, so'ng \"✅ Tasdiqlash\"ni bosing. Bu ulanish faqat BIR MARTA "
        f"so'raladi — keyin shu tugmani bosganingizda to'g'ridan-to'g'ri shu hisoblar "
        f"bo'yicha balans chiqadi.",
        reply_markup=build_method_link_keyboard(method, names, set()),
    )


async def show_linked_method(update: Update, context: ContextTypes.DEFAULT_TYPE, method: str, title: str):
    space_key = get_space_key_for_update(update)
    linked = get_method_accounts(space_key, method)
    if not linked:
        await start_method_link(update, context, method, title)
        return
    full_title = f"{title} bo'yicha ulangan hisoblar"
    text = build_method_expense_text(space_key, method, full_title, "all")
    await update.message.reply_text(text, reply_markup=build_method_period_keyboard(method, "all"))


async def card_expenses(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await show_linked_method(update, context, "karta", "💳 Karta xarajati")


async def cash_expenses(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await show_linked_method(update, context, "naqt", "💵 Naqt xarajat")


async def method_link_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    space_key = get_space_key_for_update(update)
    parts = query.data.split(":")
    method = parts[1]
    action = parts[2]
    key = f"pending_method_link_{method}"
    title = "💳 Karta xarajati" if method == "karta" else "💵 Naqt xarajat"

    if action == "relink":
        context.chat_data[key] = set()
        names = get_account_names(space_key)
        await query.edit_message_text(
            f"<b>{esc(title)}</b>\n\nQaysi hisob(lar)ni ulaymiz? Tanlang va \"✅ Tasdiqlash\"ni bosing.",
            reply_markup=build_method_link_keyboard(method, names, set()),
        )
        return

    if action == "toggle":
        name = parts[3]
        selected = context.chat_data.setdefault(key, set())
        if name in selected:
            selected.discard(name)
        else:
            selected.add(name)
        names = get_account_names(space_key)
        await query.edit_message_text(
            f"<b>{esc(title)}</b>\n\nQaysi hisob(lar)ni ulaymiz? Tanlang va \"✅ Tasdiqlash\"ni bosing.",
            reply_markup=build_method_link_keyboard(method, names, selected),
        )
        return

    if action == "confirm":
        selected = context.chat_data.pop(key, set())
        if not selected:
            names = get_account_names(space_key)
            await query.edit_message_text(
                f"<b>{esc(title)}</b>\n\n⚠️ Kamida bitta hisobni tanlang, keyin \"✅ Tasdiqlash\"ni bosing.",
                reply_markup=build_method_link_keyboard(method, names, set()),
            )
            return
        ordered = [n for n in get_account_names(space_key) if n in selected]
        set_method_accounts(space_key, method, ordered)
        full_title = f"{title} bo'yicha ulangan hisoblar"
        text = build_method_expense_text(space_key, method, full_title, "all")
        await query.edit_message_text(text, reply_markup=build_method_period_keyboard(method, "all"))
        await send_main_menu(query.message)
        return

    if action == "cancel":
        context.chat_data.pop(key, None)
        await query.edit_message_text("Bekor qilindi. ✅")
        await send_main_menu(query.message)
        return


async def method_period_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    space_key = get_space_key_for_update(update)
    parts = query.data.split(":")  # methodperiod:<method>:<period>
    method = parts[1]
    period = parts[2]
    title = "💳 Karta xarajati bo'yicha ulangan hisoblar" if method == "karta" else "💵 Naqt xarajat bo'yicha ulangan hisoblar"
    text = build_method_expense_text(space_key, method, title, period)
    await query.edit_message_text(text, reply_markup=build_method_period_keyboard(method, period))


async def saldo_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    clear_conflicting_pending_flags(context)
    context.chat_data["pending_saldo_input"] = True
    await update.message.reply_text(build_accounts_summary_text(space_key))


# --- Hisobdan hisobga o'tkazma ---

async def transfer_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    clear_conflicting_pending_flags(context)
    for k in ("transfer_from_account", "transfer_amount", "pending_transfer_amount"):
        context.chat_data.pop(k, None)
    names = get_account_names(space_key)
    if len(names) < 2:
        await update.message.reply_text(
            "O'tkazma qilish uchun kamida 2 ta hisob kerak. "
            "\"⚙️ Hisoblarni boshqarish\" orqali yana hisob qo'shing."
        )
        return
    buttons = [[InlineKeyboardButton(n, callback_data=f"transfer:from:{n}")] for n in names]
    buttons.append([InlineKeyboardButton("❌ Bekor qilish", callback_data="transfer:cancel")])
    await update.message.reply_text(
        "🔁 <b>Hisobdan hisobga o'tkazma</b>\n\nQaysi hisobdan pul o'tkazamiz?",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def transfer_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    space_key = get_space_key_for_update(update)
    data = query.data

    if data == "transfer:cancel":
        for k in ("transfer_from_account", "transfer_amount", "pending_transfer_amount"):
            context.chat_data.pop(k, None)
        await query.edit_message_text("Bekor qilindi. ✅")
        await send_main_menu(query.message)
        return

    if data.startswith("transfer:from:"):
        name = data.split(":", 2)[2]
        accounts = dict(get_accounts(space_key))
        balance = accounts.get(name, 0.0)
        context.chat_data["transfer_from_account"] = name
        context.chat_data["pending_transfer_amount"] = True
        await query.edit_message_text(
            f"🏦 <b>{esc(name)}</b> joriy balans: <b>{fmt_num(balance, 2)} so'm</b>\n\n"
            f"Miqdorni yozing (maksimal: {fmt_num(balance, 2)} so'm):"
        )
        return

    if data.startswith("transfer:to:"):
        to_name = data.split(":", 2)[2]
        from_name = context.chat_data.get("transfer_from_account")
        amount = context.chat_data.get("transfer_amount")
        if not from_name or not amount:
            await query.edit_message_text("Bu so'rovning muddati o'tgan. Qaytadan boshlang.")
            await send_main_menu(query.message)
            return
        accounts = dict(get_accounts(space_key))
        from_balance = accounts.get(from_name, 0.0)
        if amount > from_balance + 0.005:
            await query.edit_message_text(
                f"⚠️ <b>{esc(from_name)}</b> hisobida faqat {fmt_num(from_balance, 2)} so'm bor. "
                f"Bundan ko'p miqdorni o'tkaza olmaysiz."
            )
            for k in ("transfer_from_account", "transfer_amount", "pending_transfer_amount"):
                context.chat_data.pop(k, None)
            await send_main_menu(query.message)
            return
        apply_account_delta(space_key, from_name, -amount)
        apply_account_delta(space_key, to_name, amount)
        for k in ("transfer_from_account", "transfer_amount", "pending_transfer_amount"):
            context.chat_data.pop(k, None)
        await query.edit_message_text(
            f"✅ <b>{fmt_num(amount, 2)} so'm</b> — <b>{esc(from_name)}</b> dan <b>{esc(to_name)}</b> ga o'tkazildi.\n\n"
            f"{build_accounts_summary_text(space_key)}"
        )
        await send_main_menu(query.message)
        return


# --- Hisoblarni boshqarish (nomini o'zgartirish / yangi qo'shish) ---

async def manage_accounts_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    buttons = [
        [InlineKeyboardButton("✏️ Hisoblarni tahrirlash", callback_data="acctmgmt:edit")],
        [InlineKeyboardButton("➕ Yangi hisob qo'shish", callback_data="acctmgmt:add")],
    ]
    await update.message.reply_text(
        "⚙️ <b>Hisoblarni boshqarish</b>\n\nBu yerda hisoblaringizni o'zingiz xohlagancha "
        "nomlashingiz yoki yangi hisob qo'shishingiz mumkin — kodga yoki Railway/GitHub'ga "
        "kirish shart emas.",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def account_management_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    space_key = get_space_key_for_update(update)

    if data == "acctmgmt:add":
        clear_conflicting_pending_flags(context)
        context.chat_data["pending_new_account"] = True
        await query.edit_message_text("➕ Yangi hisob nomini yozing (masalan: Kapitalbank):")
        return

    if data == "acctmgmt:edit":
        names = get_account_names(space_key)
        buttons = [
            [
                InlineKeyboardButton(f"✏️ {n}", callback_data=f"acctrename:{n}"),
                InlineKeyboardButton("🗑 O'chirish", callback_data=f"acctdel:{n}"),
            ]
            for n in names
        ]
        await query.edit_message_text(
            "✏️ Qaysi hisobni tahrirlamoqchisiz?\n\n✏️ — nomini o'zgartirish\n🗑 — hisobni butunlay o'chirish",
            reply_markup=InlineKeyboardMarkup(buttons),
        )
        return

    if data.startswith("acctrename:"):
        old_name = data.split(":", 1)[1]
        clear_conflicting_pending_flags(context)
        context.chat_data["pending_rename_account"] = old_name
        await query.edit_message_text(f"✏️ <b>{esc(old_name)}</b> uchun yangi nomni yozing:")
        return

    if data.startswith("acctdelyes:"):
        name = data.split(":", 1)[1]
        ok = delete_account(space_key, name)
        if ok:
            remove_account_from_method_links(space_key, name)
            remaining = get_account_names(space_key)
            if remaining:
                await query.edit_message_text(f"🗑 <b>{esc(name)}</b> hisobi o'chirildi.\n\n{build_accounts_summary_text(space_key)}")
            else:
                await query.edit_message_text(f"🗑 <b>{esc(name)}</b> hisobi o'chirildi. Hozircha boshqa hisobingiz yo'q — \"➕ Yangi hisob qo'shish\" orqali qo'shishingiz mumkin.")
        else:
            await query.edit_message_text("Bu hisobni topib bo'lmadi.")
        await send_main_menu(query.message)
        return

    if data.startswith("acctdelno:"):
        await query.edit_message_text("Bekor qilindi. Hisob saqlanib qoldi. ✅")
        await send_main_menu(query.message)
        return

    if data.startswith("acctdel:"):
        name = data.split(":", 1)[1]
        buttons = [
            [InlineKeyboardButton("✅ Ha, o'chirish", callback_data=f"acctdelyes:{name}")],
            [InlineKeyboardButton("❌ Yo'q, bekor qilish", callback_data=f"acctdelno:{name}")],
        ]
        await query.edit_message_text(
            f"⚠️ Rostdan ham <b>{esc(name)}</b> hisobini o'chirmoqchimisiz?\n"
            f"Bu hisob ro'yxatdan butunlay olib tashlanadi (eski yozuvlar tarixda qolaveradi).",
            reply_markup=InlineKeyboardMarkup(buttons),
        )
        return

    if data == "acctaddconfirm:yes":
        name = context.chat_data.pop("pending_new_account_name", None)
        if not name:
            await query.edit_message_text("Bu so'rovning muddati o'tgan. Qaytadan urinib ko'ring.")
            await send_main_menu(query.message)
            return
        ok = add_custom_account(space_key, name)
        if ok:
            await query.edit_message_text(f"✅ Yangi hisob qo'shildi: <b>{esc(name)}</b>\n\n{build_accounts_summary_text(space_key)}")
        else:
            await query.edit_message_text(f"Bu nomdagi hisob allaqachon mavjud: {esc(name)}")
        await send_main_menu(query.message)
        return

    if data == "acctaddconfirm:no":
        context.chat_data.pop("pending_new_account_name", None)
        await query.edit_message_text("Bekor qilindi. ✅")
        await send_main_menu(query.message)
        return

    if data == "acctrenameconfirm:yes":
        pending = context.chat_data.pop("pending_rename_new_name", None)
        if not pending:
            await query.edit_message_text("Bu so'rovning muddati o'tgan. Qaytadan urinib ko'ring.")
            await send_main_menu(query.message)
            return
        old_name, new_name = pending
        ok = rename_account(space_key, old_name, new_name)
        if ok:
            await query.edit_message_text(f"✅ <b>{esc(old_name)}</b> endi <b>{esc(new_name)}</b> deb nomlandi.\n\n{build_accounts_summary_text(space_key)}")
        else:
            await query.edit_message_text("Bu hisobni topib bo'lmadi.")
        await send_main_menu(query.message)
        return

    if data == "acctrenameconfirm:no":
        context.chat_data.pop("pending_rename_new_name", None)
        await query.edit_message_text("Bekor qilindi. Hisob nomi o'zgarmadi. ✅")
        await send_main_menu(query.message)
        return


# --- Kalendar ---

UZ_MONTHS = ["Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
             "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr"]
UZ_WEEKDAYS = ["Du", "Se", "Ch", "Pa", "Ju", "Sh", "Ya"]


def build_calendar_markup(year: int, month: int, range_mode: bool = False) -> InlineKeyboardMarkup:
    import calendar as _cal
    cal = _cal.Calendar(firstweekday=0)
    weeks = cal.monthdayscalendar(year, month)
    today = now_tz().date()

    prev_year, prev_month = (year - 1, 12) if month == 1 else (year, month - 1)
    next_year, next_month = (year + 1, 1) if month == 12 else (year, month + 1)

    day_prefix = "range" if range_mode else "day"

    rows = [
        [
            InlineKeyboardButton("◀️", callback_data=f"cal:nav:{prev_year}-{prev_month:02d}:{1 if range_mode else 0}"),
            InlineKeyboardButton(f"📅 {UZ_MONTHS[month-1]} {year}", callback_data="cal:ignore"),
            InlineKeyboardButton("▶️", callback_data=f"cal:nav:{next_year}-{next_month:02d}:{1 if range_mode else 0}"),
        ],
        [InlineKeyboardButton(d, callback_data="cal:ignore") for d in UZ_WEEKDAYS],
    ]
    for week in weeks:
        row = []
        for day in week:
            if day == 0:
                row.append(InlineKeyboardButton(" ", callback_data="cal:ignore"))
            else:
                is_today = (year == today.year and month == today.month and day == today.day)
                label = f"🔴{day}" if is_today else str(day)
                row.append(InlineKeyboardButton(label, callback_data=f"cal:{day_prefix}:{year}-{month:02d}-{day:02d}"))
        rows.append(row)
    if range_mode:
        rows.append([InlineKeyboardButton("❌ Bekor qilish", callback_data="cal:range_cancel")])
    else:
        rows.append([InlineKeyboardButton("🔵 Bugungi kun", callback_data=f"cal:day:{today.strftime('%Y-%m-%d')}")])
        rows.append([InlineKeyboardButton("📊 Davr bo'yicha hisobot", callback_data="cal:range_start")])
    return InlineKeyboardMarkup(rows)


async def calendar_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.chat_data.pop("range_start_date", None)
    now = now_tz()
    await update.message.reply_text(
        "📅 <b>Hisobot olish uchun kunni tanlang:</b>",
        reply_markup=build_calendar_markup(now.year, now.month),
    )


def build_range_report_text(space_key: str, start_date: str, end_date: str) -> str:
    if start_date > end_date:
        start_date, end_date = end_date, start_date

    conn = get_db()
    rows = conn.execute(
        "SELECT amount, currency, category FROM transactions "
        "WHERE space_key=? AND substr(created_at,1,10)>=? AND substr(created_at,1,10)<=?",
        (space_key, start_date, end_date),
    ).fetchall()
    conn.close()

    def pretty(d):
        try:
            return datetime.strptime(d, "%Y-%m-%d").strftime("%d.%m.%Y")
        except ValueError:
            return d

    income = {"UZS": 0.0, "USD": 0.0}
    expense = {"UZS": 0.0, "USD": 0.0}
    cat_totals = {}
    for amount, currency, category in rows:
        if amount >= 0:
            income[currency] += amount
        else:
            expense[currency] += -amount
            if currency == "UZS":
                cat_totals[category] = cat_totals.get(category, 0) + (-amount)

    lines = [f"📊 <b>{pretty(start_date)} — {pretty(end_date)}</b>", ""]
    lines.append(f"📈 Daromad: <b>{fmt_num(income['UZS'])}</b> so'm, {fmt_num(income['USD'], 2)} $")
    lines.append(f"📉 Xarajat: <b>{fmt_num(expense['UZS'])}</b> so'm, {fmt_num(expense['USD'], 2)} $")

    if cat_totals:
        lines.append("")
        lines.append("🗂 <b>Kategoriyalar bo'yicha xarajat:</b>")
        for cat, val in sorted(cat_totals.items(), key=lambda x: -x[1]):
            lines.append(f"▫️ {esc(cat)}: {fmt_num(val)} so'm")

    balances = get_account_balances_as_of(space_key, end_date)
    lines.append("")
    lines.append(f"🏦 <b>{pretty(end_date)} kuni oxiridagi hisoblar qoldig'i</b>")
    total = 0.0
    for name in get_account_names(space_key):
        bal = balances.get(name, 0.0)
        total += bal
        marker = "🔴 " if bal < 0 else "▫️ "
        lines.append(f"{marker}<b>{name}</b>: {fmt_num(bal)} so'm")
    lines.append(f"<b>Umumiy: {fmt_num(total)} so'm</b>")

    return "\n".join(lines)




def build_day_report_text(space_key: str, date_str: str) -> str:
    conn = get_db()
    rows = conn.execute(
        "SELECT amount, currency, category, note, created_at FROM transactions "
        "WHERE space_key=? AND created_at LIKE ? ORDER BY id",
        (space_key, date_str + "%"),
    ).fetchall()
    conn.close()

    try:
        pretty_date = datetime.strptime(date_str, "%Y-%m-%d").strftime("%d.%m.%Y")
    except ValueError:
        pretty_date = date_str

    if not rows:
        lines = [f"📅 <b>{pretty_date}</b>: bu kunda xarajat yo'q."]
    else:
        income = {"UZS": 0.0, "USD": 0.0}
        expense = {"UZS": 0.0, "USD": 0.0}
        lines = [f"📅 <b>{pretty_date} bo'yicha hisobot</b>", ""]
        for amount, currency, category, note, created_at in rows:
            time_str = created_at.split("T")[1][:5] if "T" in created_at else ""
            sign = "+" if amount >= 0 else ""
            unit = "so'm" if currency == "UZS" else "$"
            lines.append(f"🕐 {time_str}  <b>{sign}{fmt_num(amount)} {unit}</b>  <i>[{esc(category)}]</i> — {esc(note)}")
            if amount >= 0:
                income[currency] += amount
            else:
                expense[currency] += -amount

        lines.append("")
        lines.append(f"📈 Daromad: <b>{fmt_num(income['UZS'])}</b> so'm, {fmt_num(income['USD'], 2)} $")
        lines.append(f"📉 Xarajat: <b>{fmt_num(expense['UZS'])}</b> so'm, {fmt_num(expense['USD'], 2)} $")

    # O'sha kun oxiridagi hisob qoldiqlari (yozuv bo'lsa ham, bo'lmasa ham)
    balances = get_account_balances_as_of(space_key, date_str)
    lines.append("")
    lines.append(f"🏦 <b>{pretty_date} kuni oxiridagi hisoblar qoldig'i</b>")
    total = 0.0
    for name in get_account_names(space_key):
        bal = balances.get(name, 0.0)
        total += bal
        marker = "🔴 " if bal < 0 else "▫️ "
        lines.append(f"{marker}<b>{name}</b>: {fmt_num(bal)} so'm")
    lines.append(f"<b>Umumiy: {fmt_num(total)} so'm</b>")

    return "\n".join(lines)


async def calendar_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "cal:ignore":
        return

    if data.startswith("cal:nav:"):
        rest = data.split(":", 2)[2]
        ym, range_flag = rest.rsplit(":", 1)
        year, month = map(int, ym.split("-"))
        range_mode = range_flag == "1"
        prompt = "📅 <b>Boshlang'ich sanani tanlang:</b>" if (range_mode and "range_start_date" not in context.chat_data) else (
            "📅 <b>Oxirgi sanani tanlang:</b>" if range_mode else "📅 <b>Hisobot olish uchun kunni tanlang:</b>"
        )
        await query.edit_message_text(prompt, reply_markup=build_calendar_markup(year, month, range_mode))
        return

    if data == "cal:range_start":
        context.chat_data.pop("range_start_date", None)
        now = now_tz()
        await query.edit_message_text(
            "📅 <b>Boshlang'ich sanani tanlang:</b>",
            reply_markup=build_calendar_markup(now.year, now.month, range_mode=True),
        )
        return

    if data == "cal:range_cancel":
        context.chat_data.pop("range_start_date", None)
        now = now_tz()
        await query.edit_message_text(
            "📅 <b>Hisobot olish uchun kunni tanlang:</b>",
            reply_markup=build_calendar_markup(now.year, now.month),
        )
        return

    if data.startswith("cal:range:"):
        date_str = data.split(":", 2)[2]
        start = context.chat_data.get("range_start_date")
        if not start:
            context.chat_data["range_start_date"] = date_str
            y, m, d = map(int, date_str.split("-"))
            await query.edit_message_text(
                f"Boshlang'ich sana: <b>{date_str}</b>\n📅 <b>Endi oxirgi sanani tanlang:</b>",
                reply_markup=build_calendar_markup(y, m, range_mode=True),
            )
            return
        context.chat_data.pop("range_start_date", None)
        space_key = get_space_key_for_update(update)
        await query.edit_message_text(build_range_report_text(space_key, start, date_str))
        return

    if data.startswith("cal:day:"):
        date_str = data.split(":", 2)[2]
        space_key = get_space_key_for_update(update)
        await query.edit_message_text(build_day_report_text(space_key, date_str))
        return


async def kurs_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    cur_kurs = get_kurs(space_key)
    clear_conflicting_pending_flags(context)
    context.chat_data["pending_kurs_input"] = True
    await update.message.reply_text(
        f"Joriy kurs: 1$ = {fmt_num(cur_kurs)} so'm\n\n"
        f"Yangi kursni raqam bilan yozing, masalan: 12700"
    )


async def kurs_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    if not context.args:
        await kurs_menu(update, context)
        return
    try:
        value = float(context.args[0].replace(",", ""))
        if value <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("Noto'g'ri qiymat. Masalan: /kurs 12700")
        return
    set_kurs(space_key, value)
    await update.message.reply_text(f"Kurs yangilandi: 1$ = {fmt_num(value)} so'm ✅")


async def group_new(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    code = create_group(user_id)
    await update.message.reply_text(
        f"👨‍👩‍👧 Jamoaviy balans yaratildi!\n\n"
        f"Taklif kodi: {code}\n\n"
        f"Oila a'zolaringiz shu kodni /group_join {code} bilan kiritsa, "
        f"hammangiz bitta umumiy balansni yuritasiz."
    )


async def group_join(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Kodni kiriting. Masalan: /group_join AB12CD")
        return
    code = context.args[0]
    ok = join_group(update.effective_user.id, code)
    if ok:
        await update.message.reply_text(
            f"✅ Jamoaviy balansga qo'shildingiz!\n\n"
            f"{format_balance_text(get_space_key_for_update(update))}"
        )
    else:
        await update.message.reply_text("Bunday kod topilmadi. Qaytadan tekshirib ko'ring.")


async def group_leave(update: Update, context: ContextTypes.DEFAULT_TYPE):
    leave_group(update.effective_user.id)
    await update.message.reply_text("Jamoaviy balansdan chiqdingiz. Endi shaxsiy balansingiz ishlaydi. 👤")


async def group_info(update: Update, context: ContextTypes.DEFAULT_TYPE):
    info = get_group_info(update.effective_user.id)
    if not info:
        await update.message.reply_text("Siz hozir shaxsiy balansdasiz (jamoaga qo'shilmagansiz).")
        return
    code, group_id, member_count = info
    await update.message.reply_text(
        f"👨‍👩‍👧 Jamoa ma'lumoti:\nKod: {code}\nA'zolar soni: {member_count}"
    )


async def reset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    reset_space(space_key)
    await update.message.reply_text("Balans va tarix tozalandi, hisoblar 0'ga tushirildi. 🧹")
    await send_main_menu(update.message)


async def reset_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    buttons = [
        [InlineKeyboardButton("✅ Ha, tozalash", callback_data="reset:confirm")],
        [InlineKeyboardButton("❌ Bekor qilish", callback_data="reset:cancel")],
    ]
    await update.message.reply_text(
        "⚠️ <b>Diqqat!</b> Bu barcha xarajat/daromad, qarz tarixini o'chiradi VA "
        "hisoblar balansini 0'ga tushiradi. Bosmasangiz, eski hisob-kitob "
        "hech qayerga uchmaydi. Rostdan ham tozalashni xohlaysizmi?",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def reset_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "reset:confirm":
        space_key = get_space_key_for_update(update)
        reset_space(space_key)
        await query.edit_message_text("Balans va tarix tozalandi, hisoblar 0'ga tushirildi. 🧹")
    else:
        await query.edit_message_text("Bekor qilindi. Ma'lumotlaringiz saqlanib qoldi. ✅")
    await send_main_menu(query.message)


# --- Xato kiritilgan summani tuzatish ---

def get_recent_editable_transactions(space_key: str, limit: int = 10):
    conn = get_db()
    rows = conn.execute(
        "SELECT id, amount, currency, category, note, created_at FROM transactions "
        "WHERE space_key=? ORDER BY id DESC LIMIT ?",
        (space_key, limit),
    ).fetchall()
    conn.close()
    return rows


def get_transaction_by_id(tx_id: int):
    conn = get_db()
    row = conn.execute(
        "SELECT id, user_id, space_key, amount, currency, category, note, method, account, created_at "
        "FROM transactions WHERE id=?",
        (tx_id,),
    ).fetchone()
    conn.close()
    return row


def update_transaction_amount(tx_id: int, new_amount: float):
    row = get_transaction_by_id(tx_id)
    if not row:
        return False
    _id, user_id, space_key, old_amount, currency, category, note, method, account, created_at = row
    conn = get_db()
    conn.execute("UPDATE transactions SET amount=? WHERE id=?", (new_amount, tx_id))
    conn.commit()
    conn.close()
    if account:
        kurs = get_kurs(space_key)
        old_delta_uzs = old_amount if currency == "UZS" else old_amount * kurs
        new_delta_uzs = new_amount if currency == "UZS" else new_amount * kurs
        apply_account_delta(space_key, account, new_delta_uzs - old_delta_uzs)
    return True


def update_transaction_category(tx_id: int, new_category: str):
    conn = get_db()
    conn.execute("UPDATE transactions SET category=? WHERE id=?", (new_category, tx_id))
    conn.commit()
    conn.close()
    return True


def update_transaction_account(tx_id: int, new_account: str):
    row = get_transaction_by_id(tx_id)
    if not row:
        return False
    _id, user_id, space_key, amount, currency, category, note, method, old_account, created_at = row
    conn = get_db()
    conn.execute("UPDATE transactions SET account=? WHERE id=?", (new_account, tx_id))
    conn.commit()
    conn.close()
    kurs = get_kurs(space_key)
    delta_uzs = amount if currency == "UZS" else amount * kurs
    if old_account:
        apply_account_delta(space_key, old_account, -delta_uzs)
    if new_account:
        apply_account_delta(space_key, new_account, delta_uzs)
    return True


async def edit_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    rows = get_recent_editable_transactions(space_key, limit=10)
    if not rows:
        await update.message.reply_text("Hali yozuvlar yo'q.")
        return
    buttons = []
    for tx_id, amount, currency, category, note, created_at in rows:
        unit = "so'm" if currency == "UZS" else "$"
        sign = "+" if amount >= 0 else ""
        time_str = created_at.split("T")[1][:5] if "T" in created_at else created_at.split("T")[0]
        short_note = (note[:18] + "…") if note and len(note) > 18 else (note or "")
        label = f"{time_str}  {sign}{fmt_num(amount)} {unit} — {short_note}"
        buttons.append([InlineKeyboardButton(label, callback_data=f"edit:{tx_id}")])
    await update.message.reply_text(
        "✏️ <b>Qaysi yozuvni tuzatmoqchisiz?</b>",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def edit_pick_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    tx_id = int(query.data.split(":", 1)[1])
    row = get_transaction_by_id(tx_id)
    if not row:
        await query.edit_message_text("Bu yozuv topilmadi (ehtimol allaqachon o'chirilgan).")
        return
    _id, user_id, space_key, amount, currency, category, note, method, account, created_at = row
    unit = "so'm" if currency == "UZS" else "$"
    sign = "+" if amount >= 0 else ""

    buttons = [
        [InlineKeyboardButton("💰 Miqdorni tuzatish", callback_data=f"editfield:amount:{tx_id}")],
        [InlineKeyboardButton("🗂 Kategoriyani o'zgartirish", callback_data=f"editfield:category:{tx_id}")],
        [InlineKeyboardButton("🏦 Hisobni o'zgartirish", callback_data=f"editfield:account:{tx_id}")],
    ]
    await query.edit_message_text(
        f"Joriy yozuv: <b>{sign}{fmt_num(amount)} {unit}</b>  <i>[{esc(category)}]</i>"
        f"{f'  ({esc(account)})' if account else ''} — {esc(note)}\n\n"
        f"Nimasini tuzatmoqchisiz?",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def edit_field_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    _, field, tx_id_str = query.data.split(":", 2)
    tx_id = int(tx_id_str)
    space_key = get_space_key_for_update(update)

    if field == "amount":
        context.chat_data["pending_edit_tx_id"] = tx_id
        await query.edit_message_text("To'g'ri summani yozing (masalan: 220000):")
        return

    if field == "category":
        context.chat_data["pending_edit_category_tx_id"] = tx_id
        await query.edit_message_text("Yangi kategoriya nomini yozing (masalan: Oziq-ovqat):")
        return

    if field == "account":
        names = get_account_names(space_key)
        buttons = [[InlineKeyboardButton(n, callback_data=f"editacct:{tx_id}:{n}")] for n in names]
        await query.edit_message_text(
            "Qaysi hisobga o'zgartirmoqchisiz?", reply_markup=InlineKeyboardMarkup(buttons)
        )
        return


async def edit_account_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    _, tx_id_str, new_account = query.data.split(":", 2)
    tx_id = int(tx_id_str)
    ok = update_transaction_account(tx_id, new_account)
    if ok:
        await query.edit_message_text(f"✅ Hisob <b>{esc(new_account)}</b> ga o'zgartirildi.")
    else:
        await query.edit_message_text("Bu yozuv topilmadi.")


# --- Qarz buyruqlari (tugma va matnli) ---

async def debts_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    buttons = [
        [InlineKeyboardButton("➕ Men qarz oldim", callback_data="debtdir:oldim")],
        [InlineKeyboardButton("➕ Men qarz berdim", callback_data="debtdir:berdim")],
        [InlineKeyboardButton("📋 Ro'yxat", callback_data="debtlist")],
    ]
    await update.message.reply_text(
        "Qarzlar bo'limi:",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def debt_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    if data == "debtlist":
        space_key = get_space_key_for_update(update)
        await query.edit_message_text(build_debts_summary_text(space_key))
        return
    direction = data.split(":")[1]
    context.chat_data["pending_debt_direction"] = direction
    if direction == "oldim":
        prompt = "Kimdan va qancha qarz oldingiz?"
    else:
        prompt = "Kimga va qancha qarz berdingiz?"
    await query.edit_message_text(prompt)


DEBT_PATTERN = re.compile(
    r"^(\S+)\s+(\d[\d\s.,]*)\s*(mln|million|млн|ming|минг)?\b\s*(\$|usd|so'?m|sum|uzs)?\s*(.*)$",
    re.IGNORECASE,
)


def parse_debt_args(text: str):
    text = text.strip()
    match = DEBT_PATTERN.match(text)
    if not match:
        return None
    person, raw_amount, mult_word, currency_raw, note = match.groups()
    amount = parse_amount_token(raw_amount, mult_word)
    if amount is None:
        return None
    currency = "UZS" if (not currency_raw or currency_raw.lower() in ("sum", "so'm", "som", "uzs")) else "USD"
    note = note.strip() or "(izohsiz)"
    return person, amount, currency, note


def parse_debt_args_flexible(text: str, direction: str):
    """Tugma orqali boshlangan qarz kiritishda ikkala formatni ham
    tushunadi: "Ali 500000" (ism avval) yoki "1000 alidan" (summa avval,
    "dan"/"ga" qo'shimchasi bilan)."""
    parsed = parse_debt_args(text)
    if parsed:
        return parsed

    m = FREEFORM_AMOUNT_RE.search(text)
    if not m:
        return None
    amount = parse_amount_token(m.group(1), m.group(2))
    if not amount:
        return None
    currency = "USD" if re.search(r"\$|\busd\b", text.lower()) else "UZS"

    pm = PERSON_FROM_RE.search(text) if direction == "oldim" else PERSON_TO_RE.search(text)
    if not pm:
        return None
    person = pm.group(1)
    note = text.strip()
    return person, amount, currency, note


async def qarz_oldim_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = " ".join(context.args)
    parsed = parse_debt_args(text) if text else None
    if not parsed:
        await update.message.reply_text("Masalan: /qarz_oldim Ali 500000 taksi uchun")
        return
    person, amount, currency, note = parsed
    add_debt(get_space_key_for_update(update), person, "oldim", amount, currency, note)
    unit = "so'm" if currency == "UZS" else "$"
    await update.message.reply_text(f"✅ Qayd etildi: <b>{esc(person)}</b> dan {fmt_num(amount)} {unit} qarz oldingiz.")


async def qarz_berdim_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = " ".join(context.args)
    parsed = parse_debt_args(text) if text else None
    if not parsed:
        await update.message.reply_text("Masalan: /qarz_berdim Vali 200000")
        return
    person, amount, currency, note = parsed
    add_debt(get_space_key_for_update(update), person, "berdim", amount, currency, note)
    unit = "so'm" if currency == "UZS" else "$"
    await update.message.reply_text(f"✅ Qayd etildi: <b>{esc(person)}</b> ga {fmt_num(amount)} {unit} qarz berdingiz.")


async def debts_list_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    await update.message.reply_text(build_debts_summary_text(space_key))


async def qarz_yopish_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Ism kiriting. Masalan: /qarz_yopish Ali")
        return
    person = " ".join(context.args)
    space_key = get_space_key_for_update(update)
    changed = settle_person(space_key, person)
    if changed:
        await update.message.reply_text(f"✅ {esc(person)} bilan hisob-kitob yopildi.")
    else:
        await update.message.reply_text(f"{esc(person)} bilan ochiq qarz topilmadi.")


# --- Excel ---

def get_dates_with_activity(space_key: str):
    conn = get_db()
    rows = conn.execute(
        "SELECT DISTINCT substr(created_at,1,10) AS d FROM transactions WHERE space_key=? ORDER BY d",
        (space_key,),
    ).fetchall()
    conn.close()
    return [r[0] for r in rows]


def build_excel_report(space_key: str) -> str:
    conn = get_db()
    rows = conn.execute(
        "SELECT amount, currency, category, note, created_at, account, user_name FROM transactions "
        "WHERE space_key=? ORDER BY id ASC",
        (space_key,),
    ).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Hisobot"

    FONT_NAME = "Aptos"
    header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    normal_font = Font(name=FONT_NAME, size=10)
    bold_font = Font(name=FONT_NAME, bold=True, size=10)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    green_font = Font(name=FONT_NAME, size=10, color="1E7B34")
    red_font = Font(name=FONT_NAME, size=10, color="C00000")
    title_font = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
    advice_font = Font(name=FONT_NAME, size=10, italic=True, color="1F4E78")
    advice_fill = PatternFill(start_color="EAF1F8", end_color="EAF1F8", fill_type="solid")

    # --- Sarlavha ---
    ws.merge_cells("A1:I1")
    title_cell = ws.cell(row=1, column=1, value="💰 XARAJAT / DAROMAD HISOBOTI")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A2:I2")
    date_cell = ws.cell(row=2, column=1, value=f"Shakllantirilgan sana: {now_tz().strftime('%d.%m.%Y %H:%M')}")
    date_cell.font = Font(name=FONT_NAME, size=9, italic=True, color="808080")
    date_cell.alignment = Alignment(horizontal="center")

    headers = ["Sana", "Vaqt", "Kim", "Turi", "Kategoriya", "Hisob", "Summa", "Valyuta", "Izoh"]
    header_row = 4
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    total_exp_uzs = total_exp_usd = 0.0
    total_inc_uzs = total_inc_usd = 0.0
    category_expense = {}  # category -> UZS ekvivalent jami xarajat (taxminiy, kurssiz oddiy yig'indi)
    r = header_row + 1
    for amount, currency, category, note, created_at, account, user_name in rows:
        try:
            dt = datetime.fromisoformat(created_at)
            date_str = dt.strftime("%d.%m.%Y")
            time_str = dt.strftime("%H:%M")
        except ValueError:
            date_str = created_at.split("T")[0]
            time_str = created_at.split("T")[1][:5] if "T" in created_at else ""
        kind = "Daromad" if amount >= 0 else "Xarajat"
        font = green_font if amount >= 0 else red_font
        who = user_name or "—"

        ws.cell(row=r, column=1, value=date_str).font = normal_font
        ws.cell(row=r, column=2, value=time_str).font = normal_font
        ws.cell(row=r, column=3, value=who).font = normal_font
        ws.cell(row=r, column=4, value=kind).font = font
        ws.cell(row=r, column=5, value=category or "").font = normal_font
        ws.cell(row=r, column=6, value=account or "—").font = normal_font
        amount_cell = ws.cell(row=r, column=7, value=amount)
        amount_cell.font = font
        amount_cell.number_format = "#,##0.00"
        ws.cell(row=r, column=8, value=currency).font = normal_font
        ws.cell(row=r, column=9, value=note or "").font = normal_font
        for col in range(1, 10):
            ws.cell(row=r, column=col).border = border

        if amount < 0:
            if currency == "UZS":
                total_exp_uzs += -amount
            else:
                total_exp_usd += -amount
            cat = category or "Boshqa"
            category_expense[cat] = category_expense.get(cat, 0.0) + (-amount)
        else:
            if currency == "UZS":
                total_inc_uzs += amount
            else:
                total_inc_usd += amount
        r += 1

    last_data_row = r - 1

    # --- Jami xarajat / daromad ---
    r += 1
    summary_start = r
    summary_rows = [
        ("Jami xarajat (UZS)", total_exp_uzs, red_font),
        ("Jami xarajat (USD)", total_exp_usd, red_font),
        ("Jami daromad (UZS)", total_inc_uzs, green_font),
        ("Jami daromad (USD)", total_inc_usd, green_font),
        ("Sof holat (UZS)", total_inc_uzs - total_exp_uzs, bold_font),
    ]
    for label, value, font in summary_rows:
        ws.cell(row=r, column=5, value=label).font = bold_font
        val_cell = ws.cell(row=r, column=7, value=value)
        val_cell.font = font
        val_cell.number_format = "#,##0.00"
        for col in range(5, 8):
            ws.cell(row=r, column=col).border = border
        r += 1

    # --- Hisoblar bo'yicha joriy saldo (har doim ko'rsatiladi — tranzaksiya
    # tarixi to'liq bo'lmasa ham, joriy balans shu yerda aks etadi) ---
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    acc_title_cell = ws.cell(row=r, column=1, value="🏦 Hisoblar bo'yicha joriy saldo")
    acc_title_cell.font = Font(name=FONT_NAME, bold=True, size=12, color="1F4E78")
    r += 1

    acc_rows = get_accounts(space_key)
    total_acc_balance = 0.0
    for name, acc_balance in acc_rows:
        total_acc_balance += acc_balance
        ws.cell(row=r, column=5, value=name).font = normal_font
        bal_cell = ws.cell(row=r, column=7, value=acc_balance)
        bal_cell.font = red_font if acc_balance < 0 else normal_font
        bal_cell.number_format = "#,##0.00"
        for col in range(5, 8):
            ws.cell(row=r, column=col).border = border
        r += 1

    ws.cell(row=r, column=5, value="Jami (barcha hisoblar)").font = bold_font
    total_acc_cell = ws.cell(row=r, column=7, value=total_acc_balance)
    total_acc_cell.font = bold_font
    total_acc_cell.number_format = "#,##0.00"
    for col in range(5, 8):
        ws.cell(row=r, column=col).border = border
    r += 1

    # --- Professional moliyaviy maslahat ---
    r += 1
    advice_title_row = r
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    advice_title_cell = ws.cell(row=r, column=1, value="📊 Moliyaviy tahlil va maslahat")
    advice_title_cell.font = Font(name=FONT_NAME, bold=True, size=12, color="1F4E78")
    r += 1

    advice_lines = []
    if total_inc_uzs <= 0 and total_exp_uzs > 0:
        advice_lines.append(
            "⚠️ Hozircha daromad yozuvlari qayd etilmagan, faqat xarajatlar bor. "
            "Real moliyaviy holatni ko'rish uchun daromadlaringizni ham botga kiritib boring."
        )
    elif total_inc_uzs > 0:
        ratio = total_exp_uzs / total_inc_uzs if total_inc_uzs else 0
        percent = ratio * 100
        if ratio > 1:
            advice_lines.append(
                f"🔴 Xarajatlaringiz daromadingizdan {fmt_num(total_exp_uzs - total_inc_uzs)} so'mga ko'p "
                f"({fmt_num(percent, 0)}%). Bu — byudjet tanqisligi. Zudlik bilan keraksiz xarajatlarni "
                f"qisqartirish yoki qo'shimcha daromad manbai topish tavsiya etiladi."
            )
        elif ratio > 0.8:
            advice_lines.append(
                f"🟠 Xarajatlar daromadning {fmt_num(percent, 0)}%'ini tashkil qiladi — bu juda yuqori "
                f"ko'rsatkich. Jamg'arma uchun deyarli imkoniyat qolmayapti. Xarajatlarni kamida "
                f"10-15%'ga qisqartirishni tavsiya qilaman."
            )
        elif ratio > 0.5:
            advice_lines.append(
                f"🟡 Xarajatlar daromadning {fmt_num(percent, 0)}%'ini tashkil qiladi — bu me'yorda, "
                f"lekin jamg'arma darajasini oshirish uchun katta xarajat toifalarini qayta ko'rib chiqish foydali bo'ladi."
            )
        else:
            advice_lines.append(
                f"🟢 Ajoyib natija! Xarajatlar daromadning atigi {fmt_num(percent, 0)}%'ini tashkil qiladi. "
                f"Jamg'arma va investitsiya uchun yaxshi imkoniyatingiz bor."
            )

    if category_expense:
        top_category, top_value = max(category_expense.items(), key=lambda kv: kv[1])
        share = (top_value / total_exp_uzs * 100) if total_exp_uzs else 0
        advice_lines.append(
            f"💡 Eng katta xarajat toifasi — <b>{top_category}</b>: {fmt_num(top_value)} so'm "
            f"(barcha xarajatning {fmt_num(share, 0)}%'i). Byudjetni tejash kerak bo'lsa, "
            f"aynan shu toifadan boshlash eng samarali."
        )

    if not advice_lines:
        advice_lines.append("Hali yetarlicha ma'lumot yo'q — xarajat/daromad yozib borgan sayin bu yerda shaxsiy tahlil va tavsiyalar paydo bo'ladi.")

    for line in advice_lines:
        clean_line = re.sub(r"</?b>", "", line)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        cell = ws.cell(row=r, column=1, value=clean_line)
        cell.font = advice_font
        cell.fill = advice_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 30
        r += 1

    widths = {1: 12, 2: 8, 3: 16, 4: 10, 5: 16, 6: 14, 7: 16, 8: 9, 9: 30}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = f"A{header_row + 1}"

    out_dir = tempfile.gettempdir()
    filename = f"balans_{space_key}_{now_tz().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(out_dir, filename)
    wb.save(filepath)
    return filepath


# --- Xarajat/daromadni yozib qo'yish (kerak bo'lsa hisob so'raydi) ---

async def commit_expense(update: Update, context: ContextTypes.DEFAULT_TYPE, user_id: int, space_key: str,
                          amount: float, currency: str, category: str, note: str, method: str = None,
                          query=None):
    account = detect_account(note, space_key)
    if not account:
        context.chat_data["pending_account_tx"] = {
            "user_id": user_id, "space_key": space_key, "amount": amount,
            "currency": currency, "category": category, "note": note, "method": method,
        }
        buttons = [[InlineKeyboardButton(name, callback_data=f"acct:{name}")] for name in get_account_names(space_key)]
        prompt = "Qaysi hisobdan xarajat qildingiz?" if amount < 0 else "Qaysi hisobga kirim tushdi?"
        if query:
            await query.edit_message_text(prompt, reply_markup=InlineKeyboardMarkup(buttons))
        else:
            await update.message.reply_text(prompt, reply_markup=InlineKeyboardMarkup(buttons))
        return
    await _finalize_commit(update, user_id, space_key, amount, currency, category, note, method, account, query=query)


async def _finalize_commit(update: Update, user_id: int, space_key: str, amount: float, currency: str,
                            category: str, note: str, method: str, account: str, query=None):
    user_name = get_user_display_name(update)
    add_transaction(user_id, space_key, amount, currency, category, note, method, account, user_name)
    if account:
        # Hisoblar doim UZS ekvivalentida yuritiladi — agar tranzaksiya
        # USD bo'lsa, joriy kursga aylantirib keyin hisobdan ayiramiz/qo'shamiz
        kurs = get_kurs(space_key)
        delta_uzs = amount if currency == "UZS" else amount * kurs
        apply_account_delta(space_key, account, delta_uzs)

    sign = "+" if amount >= 0 else ""
    unit = "so'm" if currency == "UZS" else "$"
    emoji = "✅" if amount >= 0 else "🧾"
    text = f"{emoji} <b>{sign}{fmt_num(amount)} {unit}</b>  <i>[{esc(category)}]</i>"
    if account:
        text += f"  <b>({esc(account)})</b>"
    text += f"\n<i>{esc(note)}</i>\n\n{build_post_transaction_summary(space_key)}"

    if query:
        await query.edit_message_text(text)
    else:
        await update.message.reply_text(text, reply_markup=MAIN_KEYBOARD)


async def account_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    pending = context.chat_data.pop("pending_account_tx", None)
    if not pending:
        await query.edit_message_text("Bu so'rovning muddati o'tgan. Qaytadan yozib ko'ring.")
        return
    account = None if data == "acct:skip" else data.split(":", 1)[1]
    await _finalize_commit(
        update, pending["user_id"], pending["space_key"], pending["amount"], pending["currency"],
        pending["category"], pending["note"], pending["method"], account, query=query,
    )


async def excel_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    space_key = get_space_key_for_update(update)
    rows = get_history(space_key, limit=1)
    accounts = get_accounts(space_key)
    has_balance = any(balance != 0 for _, balance in accounts)
    if not rows and not has_balance:
        await update.message.reply_text("Hali yozuvlar yo'q, Excel yaratib bo'lmaydi.")
        return

    await update.message.reply_text("📊 Excel fayl tayyorlanmoqda...")
    filepath = build_excel_report(space_key)
    with open(filepath, "rb") as f:
        await update.message.reply_document(
            document=f,
            filename=os.path.basename(filepath),
            caption="Balansingiz bo'yicha to'liq hisobot 📈",
        )
    os.remove(filepath)


# --- Skrinshot (rasm) qabul qilish ---

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not OCR_AVAILABLE:
        await update.message.reply_text(
            "Rasm o'qish funksiyasi hozircha o'rnatilmagan. Yordamdagi yo'riqnomaga qarang, "
            "yoki summani qo'lda yozing."
        )
        return

    photo = update.message.photo[-1]
    file = await photo.get_file()
    tmp_path = os.path.join(
        tempfile.gettempdir(), f"ocr_{update.effective_user.id}_{now_tz().strftime('%H%M%S%f')}.jpg"
    )
    await file.download_to_drive(tmp_path)

    try:
        img = Image.open(tmp_path)
        text = pytesseract.image_to_string(img)
    except Exception as e:
        print(f"[OCR XATO] {e}")
        await update.message.reply_text(
            "Rasmni o'qib bo'lmadi. Tesseract-OCR to'g'ri o'rnatilganini tekshiring, "
            "yoki summani qo'lda yozing."
        )
        return
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

    candidates = extract_amount_candidates(text)
    caption = (update.message.caption or "").strip()

    if not candidates:
        await update.message.reply_text(
            "Rasmda summani aniqlay olmadim 😕 Iltimos summani qo'lda yozing, masalan: -50000 taksi"
        )
        return

    space_key = get_space_key_for_update(update)
    method = detect_method(caption)

    if len(candidates) == 1 and caption:
        amt, cur = candidates[0]
        category = detect_category(caption, False)
        await commit_expense(update, context, update.effective_user.id, space_key, -abs(amt), cur, category, caption, method)
        return

    buttons = []
    for amt, cur in candidates:
        unit = "so'm" if cur == "UZS" else "$"
        buttons.append([InlineKeyboardButton(f"{fmt_num(amt)} {unit}", callback_data=f"amtpick:{amt}:{cur}")])
    buttons.append([InlineKeyboardButton("✏️ Qo'lda kiritish", callback_data="amtmanual")])
    context.chat_data["pending_photo_caption"] = caption
    await update.message.reply_text(
        "Rasmda bir nechta summa topildi, qaysi biri to'g'ri?",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def photo_amount_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    user_id = query.from_user.id

    if data == "amtmanual":
        context.chat_data.pop("pending_photo_caption", None)
        await query.edit_message_text("Yaxshi, summani qo'lda yozing, masalan: -50000 taksi")
        return

    _, amt_str, cur = data.split(":")
    amt = float(amt_str)
    caption = context.chat_data.pop("pending_photo_caption", "") or ""
    space_key = get_space_key_for_update(update)

    if caption:
        category = detect_category(caption, False)
        method = detect_method(caption)
        await commit_expense(update, context, user_id, space_key, -abs(amt), cur, category, caption, method, query=query)
    else:
        context.chat_data["pending_amount"] = (amt, cur)
        unit = "so'm" if cur == "UZS" else "$"
        await query.edit_message_text(
            f"Summasi: {fmt_num(amt)} {unit}\nEndi nima uchun ekanini yozing (masalan: taksi)"
        )


# --- Salomlashish, botga qaratilganlik va AI yordamchi ---

GREETING_WORDS = [
    "salom", "assalomu alaykum", "assalomu alaykum!", "salomlar", "hi", "hello",
    "hey", "xayrli tong", "xayrli kun", "xayrli kech", "qalaysiz", "yaxshimisiz",
    "salom bot", "привет",
]


def is_greeting(text: str) -> bool:
    t = text.strip().lower().rstrip("!.? ")
    return any(t == g or t.startswith(g) for g in GREETING_WORDS)


QUESTION_WORDS = ["qanaqa", "qancha", "nima", "nimaga", "nega", "kim", "qachon",
                   "qayer", "qayerda", "necha", "qaysi", "qalay", "qanday", "nechta"]


def is_question(text: str) -> bool:
    t = text.lower()
    if "?" in t:
        return True
    return any(re.search(rf"\b{w}\b", t) for w in QUESTION_WORDS)


def is_directed_at_bot(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    chat = update.effective_chat
    if not chat or chat.type not in ("group", "supergroup"):
        return True
    msg = update.message
    text = (msg.text or msg.caption or "") if msg else ""
    bot_username = context.bot.username if context.bot else None
    if bot_username and f"@{bot_username.lower()}" in text.lower():
        return True
    if msg and msg.reply_to_message and msg.reply_to_message.from_user and context.bot:
        if msg.reply_to_message.from_user.id == context.bot.id:
            return True
    return False


async def ask_ai(question: str) -> str:
    """Anthropic API orqali savolga javob oladi (agar ANTHROPIC_API_KEY
    sozlangan bo'lsa). Sekin/bloklovchi chaqiruv bo'lgani uchun alohida
    oqimda (thread) bajariladi."""
    if not AI_AVAILABLE:
        return None
    try:
        import asyncio

        def _call():
            resp = anthropic_client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=600,
                system=(
                    "Siz Telegram botisiz. Javoblaringizni FAQAT o'zbek tilida (lotin "
                    "alifbosida) yozing, boshqa hech qanday tilda javob bermang, hatto "
                    "savol boshqa tilda yozilgan bo'lsa ham. Qisqa va aniq javob bering."
                ),
                messages=[{"role": "user", "content": question}],
            )
            return resp.content[0].text

        return await asyncio.to_thread(_call)
    except Exception as e:
        print(f"[AI XATO] {e}")
        return None


# --- Markaziy matn qayta ishlash (matn xabar va ovozli xabar shu yerga tushadi) ---

async def process_free_text(update: Update, context: ContextTypes.DEFAULT_TYPE, text: str, silent_if_unrecognized: bool = False):
    text = text.strip()
    user_id = update.effective_user.id
    space_key = get_space_key_for_update(update)

    # -1) Xato kiritilgan summani tuzatish kutilayotgan bo'lsa
    if context.chat_data.get("pending_edit_tx_id"):
        tx_id = context.chat_data["pending_edit_tx_id"]
        m = re.match(
            r"^([+-]?)\s*(\d[\d\s.,]*)\s*(mln|million|млн|ming|минг)?\b",
            text.strip(), re.IGNORECASE,
        )
        if not m:
            await update.message.reply_text("Tushunmadim. Faqat to'g'ri summani raqam bilan yozing, masalan: 220000")
            return
        sign_str, raw_amount, mult_word = m.groups()
        value = parse_amount_token(raw_amount, mult_word)
        if value is None:
            await update.message.reply_text("Noto'g'ri raqam. Masalan: 220000")
            return

        row = get_transaction_by_id(tx_id)
        if not row:
            context.chat_data.pop("pending_edit_tx_id", None)
            await update.message.reply_text("Bu yozuv topilmadi.")
            return
        old_amount = row[3]
        currency = row[4]

        if sign_str == "+":
            new_amount = abs(value)
        elif sign_str == "-":
            new_amount = -abs(value)
        else:
            new_amount = abs(value) if old_amount >= 0 else -abs(value)

        update_transaction_amount(tx_id, new_amount)
        context.chat_data.pop("pending_edit_tx_id", None)
        unit = "so'm" if currency == "UZS" else "$"
        sign = "+" if new_amount >= 0 else ""
        await update.message.reply_text(
            f"✅ Tuzatildi: <b>{sign}{fmt_num(new_amount)} {unit}</b>\n\n"
            f"{build_post_transaction_summary(space_key)}"
        )
        return

    # -1b) Kategoriyani tuzatish kutilayotgan bo'lsa
    if context.chat_data.get("pending_edit_category_tx_id"):
        tx_id = context.chat_data.pop("pending_edit_category_tx_id")
        new_category = text.strip()
        if not new_category:
            await update.message.reply_text("Kategoriya nomi bo'sh bo'lmasligi kerak.")
            return
        update_transaction_category(tx_id, new_category)
        await update.message.reply_text(f"✅ Kategoriya <b>{esc(new_category)}</b> ga o'zgartirildi.")
        return

    # 0) Bir nechta qatorli xabar — har bir qator alohida xarajat/daromad
    # sifatida tushunilsa, hammasini birdaniga yozib qo'yamiz. Bu HAR QANDAY
    # kutish holatidan (kurs, saldo, qarz va h.k.) ustun turadi — chunki
    # foydalanuvchi qaysi tugmada bo'lishidan qat'iy nazar bir zumda bir
    # nechta xarajat kiritishi kerak.
    bulk = try_parse_bulk_lines(text)
    if bulk:
        for key in ("pending_kurs_input", "pending_saldo_input", "pending_debt_direction",
                    "pending_debt_awaiting_person", "pending_amount", "pending_morning_check"):
            context.chat_data.pop(key, None)

        user_name = get_user_display_name(update)
        summary_lines = []
        for amount, currency, category, note in bulk:
            method = detect_method(note)
            account = detect_account(note, space_key)
            add_transaction(user_id, space_key, amount, currency, category, note, method, account, user_name)
            if account:
                kurs = get_kurs(space_key)
                delta_uzs = amount if currency == "UZS" else amount * kurs
                apply_account_delta(space_key, account, delta_uzs)
            sign = "+" if amount >= 0 else ""
            unit = "so'm" if currency == "UZS" else "$"
            line_text = f"{sign}{fmt_num(amount)} {unit}  <i>[{esc(category)}]</i>"
            if account:
                line_text += f"  <b>({esc(account)})</b>"
            summary_lines.append(line_text)

        header = f"✅ <b>{len(bulk)} ta yozuv qabul qilindi:</b>\n"
        text_reply = header + "\n".join(summary_lines) + f"\n\n{build_post_transaction_summary(space_key)}"
        await update.message.reply_text(text_reply)
        return

    # 1) Kurs kiritilishi kutilayotgan bo'lsa (faqat "sof raqam" kelsa)
    if context.chat_data.get("pending_kurs_input"):
        cleaned = re.sub(r"[\s,]", "", text)
        if re.fullmatch(r"\d+(\.\d+)?", cleaned):
            value = float(cleaned)
            if value <= 0:
                await update.message.reply_text("Noto'g'ri raqam. Masalan: 12700")
                return
            set_kurs(space_key, value)
            context.chat_data.pop("pending_kurs_input", None)
            await update.message.reply_text(f"Kurs yangilandi: 1$ = {fmt_num(value)} so'm ✅")
            await send_main_menu(update.message)
            return
        # Raqam emas -> foydalanuvchi boshqa narsa yozgan, kurs so'rovini
        # bekor qilib, xabarni oddiy tarzda qayta ishlaymiz
        context.chat_data.pop("pending_kurs_input", None)

    # 1b) Hisob saldosini kiritish kutilayotgan bo'lsa (faqat hisob nomi bilan boshlansa)
    if context.chat_data.get("pending_saldo_input"):
        if any(text.strip().lower().startswith(name.lower()) for name in get_account_names(space_key)):
            parsed_saldo = parse_saldo_input(text, space_key)
            if not parsed_saldo:
                await update.message.reply_text(
                    "Tushunmadim. Masalan: Uzcard 500000, yoki Uydagi naqt 200000\n"
                    f"Hisoblar: {', '.join(get_account_names(space_key))}"
                )
                return
            name, value = parsed_saldo
            set_account_balance(space_key, name, value)
            context.chat_data.pop("pending_saldo_input", None)
            await update.message.reply_text(f"{name}: {fmt_num(value)} so'm qilib belgilandi ✅\n\n{build_accounts_summary_text(space_key)}")
            await send_main_menu(update.message)
            return
        # Hisob nomi bilan boshlanmagan -> saldo so'rovini bekor qilib,
        # xabarni oddiy tarzda qayta ishlaymiz
        context.chat_data.pop("pending_saldo_input", None)

    # 1c) Ertalabki saldo tekshiruviga javob kutilayotgan bo'lsa (ko'p qatorli, tugmasiz)
    if context.chat_data.get("pending_morning_check"):
        result_text = process_morning_check_reply(user_id, space_key, text)
        if result_text:
            context.chat_data.pop("pending_morning_check", None)
            await update.message.reply_text(result_text)
            return
        # Hech qanday hisob nomi topilmadi -> bu ertalabki javob emas,
        # oddiy xabar sifatida davom ettiramiz (so'rovni bekor qilmaymiz,
        # chunki ehtimol keyinroq javob berishadi)

    # 1d) Ertalabki tugma orqali bitta hisobning haqiqiy summasi kutilayotgan bo'lsa
    if context.chat_data.get("pending_morning_account"):
        name = context.chat_data["pending_morning_account"]
        m = re.match(r"^(\d[\d\s.,]*)\s*(mln|million|млн|ming|минг)?\b", text.strip(), re.IGNORECASE)
        if not m:
            await update.message.reply_text("Tushunmadim. Faqat summani yozing, masalan: 22830.49")
            return
        actual = parse_amount_token(m.group(1), m.group(2))
        if actual is None:
            await update.message.reply_text("Noto'g'ri raqam. Masalan: 22830.49")
            return
        outcome = apply_morning_account_check(user_id, space_key, name, actual)
        context.chat_data.pop("pending_morning_account", None)
        names = get_account_names(space_key)
        buttons = [[InlineKeyboardButton(n, callback_data=f"morningacct:{n}")] for n in names]
        await update.message.reply_text(
            f"✅ <b>{esc(name)}</b>: {outcome}\n\nYana boshqa hisobni ham tekshirasizmi?",
            reply_markup=InlineKeyboardMarkup(buttons),
        )
        return

    # 1e) Yangi hisob qo'shish kutilayotgan bo'lsa -> endi TASDIQLASH so'raladi
    if context.chat_data.get("pending_new_account"):
        if await maybe_handle_transaction_during_pending(update, context, user_id, space_key, text, "pending_new_account", "yangi hisob nomini kiritish"):
            return
        name = text.strip()
        context.chat_data.pop("pending_new_account", None)
        if not name:
            await update.message.reply_text("Hisob nomi bo'sh bo'lmasligi kerak.")
            await send_main_menu(update.message)
            return
        context.chat_data["pending_new_account_name"] = name
        buttons = [
            [InlineKeyboardButton("✅ Tasdiqlash", callback_data="acctaddconfirm:yes")],
            [InlineKeyboardButton("❌ Bekor qilish", callback_data="acctaddconfirm:no")],
        ]
        await update.message.reply_text(
            f"➕ Yangi hisob qo'shilsinmi: <b>{esc(name)}</b> ?",
            reply_markup=InlineKeyboardMarkup(buttons),
        )
        return

    # 1f) Hisob nomini o'zgartirish (qayta nomlash) kutilayotgan bo'lsa -> endi TASDIQLASH so'raladi
    if context.chat_data.get("pending_rename_account"):
        old_name = context.chat_data["pending_rename_account"]
        if await maybe_handle_transaction_during_pending(update, context, user_id, space_key, text, "pending_rename_account", f"\"{old_name}\" hisobini qayta nomlash"):
            return
        new_name = text.strip()
        context.chat_data.pop("pending_rename_account", None)
        if not new_name:
            await update.message.reply_text("Yangi nom bo'sh bo'lmasligi kerak.")
            await send_main_menu(update.message)
            return
        context.chat_data["pending_rename_new_name"] = (old_name, new_name)
        buttons = [
            [InlineKeyboardButton("✅ Tasdiqlash", callback_data="acctrenameconfirm:yes")],
            [InlineKeyboardButton("❌ Bekor qilish", callback_data="acctrenameconfirm:no")],
        ]
        await update.message.reply_text(
            f"✏️ <b>{esc(old_name)}</b> endi <b>{esc(new_name)}</b> deb nomlansinmi?",
            reply_markup=InlineKeyboardMarkup(buttons),
        )
        return

    # 1g) Hisobdan hisobga o'tkazma miqdori kutilayotgan bo'lsa
    if context.chat_data.get("pending_transfer_amount"):
        stripped = text.strip()
        # Agar foydalanuvchi +/- bilan yozsa - bu ALOHIDA xarajat/daromad,
        # o'tkazma miqdori emas. Uni oddiy tarzda qayd etamiz, o'tkazma
        # jarayoni esa hali tugallanmagan holda davom etadi.
        if stripped.startswith("+") or stripped.startswith("-"):
            if await maybe_handle_transaction_during_pending(
                update, context, user_id, space_key, text,
                "pending_transfer_amount", "hisobdan hisobga o'tkazma miqdorini kiritish",
            ):
                return

        from_name = context.chat_data.get("transfer_from_account")
        amount = parse_plain_amount(stripped)
        accounts = dict(get_accounts(space_key))
        from_balance = accounts.get(from_name, 0.0)

        if amount is None or amount <= 0:
            await update.message.reply_text(
                "Tushunmadim. Faqat miqdorni raqam bilan yozing (belgisiz), masalan: 50000"
            )
            return
        if amount > from_balance + 0.005:
            await update.message.reply_text(
                f"⚠️ <b>{esc(from_name)}</b> hisobida faqat {fmt_num(from_balance, 2)} so'm bor. "
                f"Bundan ko'p miqdorni o'tkaza olmaysiz. Boshqa miqdor yozing:"
            )
            return

        context.chat_data["transfer_amount"] = amount
        context.chat_data.pop("pending_transfer_amount", None)
        names = [n for n in get_account_names(space_key) if n != from_name]
        buttons = [[InlineKeyboardButton(n, callback_data=f"transfer:to:{n}")] for n in names]
        await update.message.reply_text(
            f"💵 Miqdor: <b>{fmt_num(amount, 2)} so'm</b>\n\nQaysi hisobga o'tkazamiz?",
            reply_markup=InlineKeyboardMarkup(buttons),
        )
        return

    # 2) Tugma orqali boshlangan qarz kiritish jarayoni
    pending_dir = context.chat_data.get("pending_debt_direction")
    if pending_dir:
        parsed = parse_debt_args_flexible(text, pending_dir)
        if not parsed:
            question = "Kimdan qarz oldingiz va qancha?" if pending_dir == "oldim" else "Kimga qarz berdingiz va qancha?"
            await update.message.reply_text(f"Tushunmadim. {question}")
            return
        person, amount, currency, note = parsed
        add_debt(space_key, person, pending_dir, amount, currency, note)
        context.chat_data.pop("pending_debt_direction", None)
        unit = "so'm" if currency == "UZS" else "$"
        sign = "-" if pending_dir == "oldim" else "+"
        verb = "siz qarzdorsiz" if pending_dir == "oldim" else "sizga qarzdor"
        await update.message.reply_text(f"✅ Qayd etildi: <b>{esc(person)}</b> — <b>{sign}{fmt_num(amount)} {unit}</b> ({verb})")
        await send_main_menu(update.message)
        return

    # 3) Tabiiy tilda qarz aniqlandi, lekin ism topilmadi -> ism kutilmoqda
    pending_person_info = context.chat_data.get("pending_debt_awaiting_person")
    if pending_person_info:
        direction, amount, currency, note = pending_person_info
        person = text.strip()
        if not person:
            await update.message.reply_text("Ismni yozing, masalan: Ali")
            return
        add_debt(space_key, person, direction, amount, currency, note)
        context.chat_data.pop("pending_debt_awaiting_person", None)
        unit = "so'm" if currency == "UZS" else "$"
        sign = "-" if direction == "oldim" else "+"
        verb = "siz qarzdorsiz" if direction == "oldim" else "sizga qarzdor"
        await update.message.reply_text(f"✅ Qayd etildi: <b>{esc(person)}</b> — <b>{sign}{fmt_num(amount)} {unit}</b> ({verb})")
        await send_main_menu(update.message)
        return

    # 4) Rasmdan summa tanlangandan keyin izoh kutilayotgan bo'lsa
    pending_amount = context.chat_data.get("pending_amount")
    if pending_amount:
        amt, cur = pending_amount
        note = text or "(izohsiz)"
        category = detect_category(note, False)
        method = detect_method(note)
        context.chat_data.pop("pending_amount", None)
        await commit_expense(update, context, user_id, space_key, -abs(amt), cur, category, note, method)
        return

    # 5) Tabiiy tilda yozilgan qarz xabari ("500000 Alidan qarz oldim")
    freeform_debt = try_parse_freeform_debt(text)
    if freeform_debt:
        person, direction, amount, currency, note = freeform_debt
        if person:
            add_debt(space_key, person, direction, amount, currency, note)
            unit = "so'm" if currency == "UZS" else "$"
            sign = "-" if direction == "oldim" else "+"
            verb = "siz qarzdorsiz" if direction == "oldim" else "sizga qarzdor"
            await update.message.reply_text(f"✅ Qayd etildi: <b>{esc(person)}</b> — <b>{sign}{fmt_num(amount)} {unit}</b> ({verb})")
        else:
            context.chat_data["pending_debt_awaiting_person"] = (direction, amount, currency, note)
            question = "Kimdan qarz oldingiz? Ismini yozing." if direction == "oldim" else "Kimga qarz berdingiz? Ismini yozing."
            await update.message.reply_text(question)
        return

    # 6) Oddiy xarajat/daromad yozuvi
    parsed = parse_message(text)
    if not parsed:
        is_q = is_question(text)

        if is_greeting(text) and not is_q:
            # Guruhda oddiy "salom"ga botga aloqasi yo'q bo'lsa jim turamiz
            # (keraksiz javob bermaslik uchun), shaxsiy chatda esa qisqa salom beramiz
            if not silent_if_unrecognized:
                await update.message.reply_text(
                    "Salom! 👋 Xarajat yozmoqchi bo'lsangiz, masalan: -50000 taksi kabi yozing, "
                    "yoki savolingizni yozing."
                )
            return

        # Guruhda botga bevosita murojaat qilinmagan bo'lsa ham, savolga
        # o'xshasa (masalan "ob-havo qanaqa?") AI orqali javob berishga harakat qilamiz
        should_respond = (not silent_if_unrecognized) or is_q
        if not should_respond:
            return

        if AI_AVAILABLE:
            answer = await ask_ai(text)
            if answer:
                await update.message.reply_text(esc(answer))
                return

        if not silent_if_unrecognized:
            await update.message.reply_text(
                "Tushunmadim 🤔 Masalan shunday yozing: -50000 taksi, 13 mln maosh, yoki 500000 Alidan qarz oldim"
            )
        # Guruhda savolga o'xshasa-yu AI yoq/javob berolmasa ham, keraksiz
        # "Tushunmadim" bilan spam qilmaymiz — jim qolamiz
        return

    amount, currency, category, note = parsed
    method = detect_method(note)
    await commit_expense(update, context, user_id, space_key, amount, currency, category, note, method)


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()

    button_routes = {
        BTN_BALANCE: balance,
        BTN_HISTORY: history,
        BTN_STAT: stat,
        BTN_CATEGORIES: categories,
        BTN_DEBTS: debts_menu,
        BTN_KURS: kurs_menu,
        BTN_CARD: card_expenses,
        BTN_CASH: cash_expenses,
        BTN_SALDO: saldo_menu,
        BTN_TRANSFER: transfer_menu,
        BTN_MANAGE_ACCOUNTS: manage_accounts_menu,
        BTN_CALENDAR: calendar_menu,
        BTN_EXCEL: excel_export,
        BTN_RESET: reset_menu,
        BTN_EDIT: edit_menu,
        BTN_START: start,
        BTN_HELP: help_cmd,
    }
    if text in button_routes:
        await button_routes[text](update, context)
        return

    directed = is_directed_at_bot(update, context)
    await process_free_text(update, context, text, silent_if_unrecognized=not directed)


# --- Ovozli xabar ---

async def handle_voice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not VOICE_AVAILABLE:
        await update.message.reply_text(
            "Ovozli xabarni tushunish funksiyasi hozircha o'rnatilmagan. "
            "Yordamdagi yo'riqnomaga qarang, yoki matn bilan yozing."
        )
        return

    voice = update.message.voice
    file = await voice.get_file()
    base = os.path.join(tempfile.gettempdir(), f"voice_{update.effective_user.id}_{now_tz().strftime('%H%M%S%f')}")
    ogg_path = base + ".ogg"
    wav_path = base + ".wav"
    await file.download_to_drive(ogg_path)

    text = None
    try:
        audio = AudioSegment.from_file(ogg_path)
        # Ovoz balandligini me'yorlashtirish (juda past yoki baland ovozni tekislaydi)
        # va aniqlikni oshirish uchun mono + 16kHz formatga o'tkazish
        audio = audio.normalize() if hasattr(audio, "normalize") else audio
        audio = audio.set_channels(1).set_frame_rate(16000)
        audio.export(wav_path, format="wav")
        recognizer = sr.Recognizer()
        recognizer.energy_threshold = 300
        recognizer.dynamic_energy_threshold = True
        with sr.AudioFile(wav_path) as source:
            recognizer.adjust_for_ambient_noise(source, duration=0.3)
            audio_data = recognizer.record(source)
        try:
            text = recognizer.recognize_google(audio_data, language="uz-UZ")
        except Exception as e:
            print(f"[OVOZ XATO] uz-UZ tanib bo'lmadi: {e}")
            try:
                text = recognizer.recognize_google(audio_data, language="ru-RU")
            except Exception as e2:
                print(f"[OVOZ XATO] ru-RU ham tanib bo'lmadi: {e2}")
                text = None
    except Exception as e:
        print(f"[OVOZ XATO] audio qayta ishlashda xato: {e}")
        text = None
    finally:
        for p in (ogg_path, wav_path):
            if os.path.exists(p):
                os.remove(p)

    if not text:
        await update.message.reply_text(
            "Ovozli xabarni tanib bo'lmadi 😕 Iloji bo'lsa jimroq joyda, mikrofonga yaqinroq turib "
            "qayta gapirib ko'ring, yoki matn bilan yozing."
        )
        return

    await update.message.reply_text(f"🎙 Eshitdim: \"{text}\"")
    await process_free_text(update, context, text)


async def send_daily_status(context: ContextTypes.DEFAULT_TYPE):
    """Har kuni 23:59 da (Toshkent vaqti bilan) har bir ma'lum chatga
    o'sha kunning balans holatini avtomatik yuboradi."""
    for chat_id, space_key in get_all_known_chats():
        try:
            text = "🌙 Kun yakuni — bugungi balans holati:\n\n" + build_full_status_text(space_key)
            await context.bot.send_message(chat_id=chat_id, text=text)
        except Exception as e:
            print(f"[KUNLIK XABAR XATO] chat_id={chat_id}: {e}")


async def send_morning_check(context: ContextTypes.DEFAULT_TYPE):
    """Har kuni ertalab har bir ma'lum chatga hisoblar tugma shaklida
    yuboriladi — tugmani bosib, o'sha hisobning HAQIQIY (tiyingacha aniq)
    qoldig'ini kiritish mumkin."""
    for chat_id, space_key in get_all_known_chats():
        try:
            names = get_account_names(space_key)
            buttons = [[InlineKeyboardButton(n, callback_data=f"morningacct:{n}")] for n in names]
            await context.bot.send_message(
                chat_id=chat_id,
                text=(
                    "🌅 Xayrli tong! Hisoblaringizdagi HAQIQIY qoldiqni tekshiramiz.\n"
                    "Pastdagi tugmalardan birini bosib, o'sha hisobning aniq (tiyingacha) "
                    "summasini kiriting. Farq bo'lsa, avtomatik \"Prochee rasxod/daromad\" "
                    "sifatida qayd etib, hisobni to'g'rilab qo'yaman."
                ),
                reply_markup=InlineKeyboardMarkup(buttons),
            )
        except Exception as e:
            print(f"[ERTALABKI SO'ROV XATO] chat_id={chat_id}: {e}")


async def morning_account_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    name = query.data.split(":", 1)[1]
    context.chat_data["pending_morning_account"] = name
    await query.edit_message_text(
        f"<b>{esc(name)}</b> uchun HAQIQIY summani yozing (tiyingacha, masalan: 22830.49):"
    )


def apply_morning_account_check(user_id: int, space_key: str, name: str, actual: float) -> str:
    """Bitta hisob uchun haqiqiy summani qabul qilib, kutilgan qiymat bilan
    solishtiradi, farq bo'lsa avtomatik tuzatuvchi yozuv qo'shadi."""
    current_rows = dict(get_accounts(space_key))
    expected = current_rows.get(name, 0.0)
    diff = expected - actual  # musbat -> pulingiz yo'q bo'lib qolgan (xarajat)
    if abs(diff) > 0.005:
        amount = -diff
        category = "Prochee rasxod" if diff > 0 else "Prochee daromad"
        add_transaction(user_id, space_key, amount, "UZS", category,
                         f"Avtomatik: {name} saldo farqi", None, name)
        result = (f"kutilgan {fmt_num(expected, 2)}, haqiqiy {fmt_num(actual, 2)} — "
                  f"farq {fmt_num(abs(diff), 2)} so'm [{category}] sifatida yozildi")
    else:
        result = f"mos keladi ({fmt_num(actual, 2)} so'm)"
    set_account_balance(space_key, name, actual)
    return result


def process_morning_check_reply(user_id: int, space_key: str, text: str):
    """Ertalabki xabarga ko'p qatorli javobni tahlil qiladi (tugma
    ishlatilmasa ham ishlaydigan zaxira yo'l). Xabar matnini qaytaradi."""
    lines_in = [ln.strip() for ln in text.splitlines() if ln.strip()]
    results = []
    matched_any = False
    for line in lines_in:
        parsed = parse_saldo_input(line, space_key)
        if not parsed:
            continue
        matched_any = True
        name, actual = parsed
        outcome = apply_morning_account_check(user_id, space_key, name, actual)
        results.append(f"  {esc(name)}: {outcome}")

    if not matched_any:
        return None
    return "✅ Ertalabki tekshiruv natijasi:\n" + "\n".join(results) + f"\n\n{build_accounts_summary_text(space_key)}"


async def _post_init(application):
    """Telegram'ning \"/\" menyu tugmasida buyruqlar ro'yxatini ko'rsatadi —
    shunda /start ni qo'lda yozish o'rniga menyudan bir bosishda tanlash mumkin."""
    from telegram import BotCommand
    await application.bot.set_my_commands([
        BotCommand("start", "Botni ishga tushirish / qayta boshlash"),
        BotCommand("help", "Yordam"),
        BotCommand("balance", "Balans"),
        BotCommand("history", "Tarix"),
        BotCommand("stat", "Statistika"),
        BotCommand("categories", "Kategoriyalar"),
        BotCommand("kurs", "Kurs"),
        BotCommand("qarzlar", "Qarzlar ro'yxati"),
        BotCommand("excel", "Excel hisobot"),
        BotCommand("reset", "Balansni tozalash"),
    ])


def main():
    token = os.environ.get("BOT_TOKEN")
    if not token:
        raise SystemExit("BOT_TOKEN muhit o'zgaruvchisi topilmadi. export BOT_TOKEN=... qiling.")

    defaults = Defaults(parse_mode=ParseMode.HTML)
    app = ApplicationBuilder().token(token).defaults(defaults).post_init(_post_init).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_cmd))
    app.add_handler(CommandHandler("balance", balance))
    app.add_handler(CommandHandler("history", history))
    app.add_handler(CommandHandler("stat", stat))
    app.add_handler(CommandHandler("categories", categories))
    app.add_handler(CommandHandler("kurs", kurs_cmd))
    app.add_handler(CommandHandler("group_new", group_new))
    app.add_handler(CommandHandler("group_join", group_join))
    app.add_handler(CommandHandler("group_leave", group_leave))
    app.add_handler(CommandHandler("group_info", group_info))
    app.add_handler(CommandHandler("qarz_oldim", qarz_oldim_cmd))
    app.add_handler(CommandHandler("qarz_berdim", qarz_berdim_cmd))
    app.add_handler(CommandHandler("qarzlar", debts_list_cmd))
    app.add_handler(CommandHandler("qarz_yopish", qarz_yopish_cmd))
    app.add_handler(CommandHandler("reset", reset))
    app.add_handler(CommandHandler("excel", excel_export))
    app.add_handler(CallbackQueryHandler(debt_callback, pattern=r"^debtdir:|^debtlist$"))
    app.add_handler(CallbackQueryHandler(photo_amount_callback, pattern=r"^amtpick:|^amtmanual$"))
    app.add_handler(CallbackQueryHandler(account_callback, pattern=r"^acct:"))
    app.add_handler(CallbackQueryHandler(calendar_callback, pattern=r"^cal:"))
    app.add_handler(CallbackQueryHandler(category_period_callback, pattern=r"^catperiod:"))
    app.add_handler(CallbackQueryHandler(reset_callback, pattern=r"^reset:"))
    app.add_handler(CallbackQueryHandler(edit_pick_callback, pattern=r"^edit:"))
    app.add_handler(CallbackQueryHandler(edit_field_callback, pattern=r"^editfield:"))
    app.add_handler(CallbackQueryHandler(edit_account_callback, pattern=r"^editacct:"))
    app.add_handler(CallbackQueryHandler(account_management_callback, pattern=r"^acctmgmt:|^acctrename:|^acctrenameconfirm:|^acctdel|^acctaddconfirm:"))
    app.add_handler(CallbackQueryHandler(method_link_callback, pattern=r"^methodlink:"))
    app.add_handler(CallbackQueryHandler(method_period_callback, pattern=r"^methodperiod:"))
    app.add_handler(CallbackQueryHandler(pending_continuation_callback, pattern=r"^pendingcont:"))
    app.add_handler(CallbackQueryHandler(transfer_callback, pattern=r"^transfer:"))
    app.add_handler(CallbackQueryHandler(morning_account_callback, pattern=r"^morningacct:"))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.VOICE, handle_voice))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    if app.job_queue is not None:
        try:
            tz = TASHKENT_TZ or None
            app.job_queue.run_daily(send_daily_status, time=dt_time(23, 59, tzinfo=tz))
            app.job_queue.run_daily(send_morning_check, time=dt_time(8, 0, tzinfo=tz))
            print("Kunlik xabar (23:59) va ertalabki so'rov (08:00, Toshkent vaqti) sozlandi.")
        except Exception as e:
            print(f"[OGOHLANTIRISH] Kunlik/ertalabki xabarni sozlab bo'lmadi: {e}")
    else:
        print("[OGOHLANTIRISH] job_queue mavjud emas — 'python-telegram-bot[job-queue]' o'rnatilganini tekshiring.")

    print("Bot ishga tushdi...")
    app.run_polling()


if __name__ == "__main__":
    main()
